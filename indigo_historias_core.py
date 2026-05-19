"""
RPA Vie Cloud / IndiGO: flujo por plantillas PNG numeradas + archivo de texto (mismo criterio que Stone/H40).

Requisitos:
  - Misma sesión Windows / mismo DPI (p. ej. 100 % vs 125 %) para plantillas PNG; RDP o PC lento puede
    cambiar el render del diálogo «Imprimir» — use INDIGO_IMP10_ESPERA_* / INDIGO_IMP10_ADAPTIVE y renueve PNG si hace falta.
  - Vie Cloud visible (p. ej. «Consulta historias» con «Paciente:» y el cuadro de texto). La ventana se ancla por
    plantillas en templates/ (p. ej. «1 …» cédula, «0 …» inicial, PNG con «Vie EHR» en el nombre para logo esquina
    superior derecha), por INDIGO_FRENTE_RAPIDO, INDIGO_HWND o escaneo visual de todas las top-level
    (INDIGO_VENTANA_BUSQUEDA_VISUAL_TODAS=1, defecto). El título GetWindowText es opcional (WebView2/Electron).
    INDIGO_INCLUIR_CHROMIUM_TOPLEVEL_SIN_TITULO=0 solo afecta el modo legado de enumeración; con búsqueda visual
    completa no se excluyen ventanas sin título por ese flag.
  - `flujo_indigo_inicio.txt` (si existe) prepara la pantalla; el menú previo puede hacerse a mano.
  - Carpeta `templates/` con PNG numerados (p. ej. 0–12) e insumos por E.S.E. (ver INSUMOS_ROOT).
  - Entre plantillas 9 y 10: no traer IndiGO/Cursor al frente (INDIGO_BLOQUEO_FOCO_IMPRESION=1);
    INDIGO_FRENTE_RAPIDO no debe anclar Cursor/IDE (Chrome_WidgetWin_1).
"""
from __future__ import annotations

import csv
import ctypes
import io
import json
import os
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import cv2
import numpy as np
import pyautogui
import win32con
import win32gui
from ctypes import wintypes
from pywinauto import keyboard

try:
    import mss
except ImportError as e:
    print("Instale dependencias: pip install -r requirements.txt", file=sys.stderr)
    raise e


def _marcar_dpi_consciente() -> None:
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except (OSError, AttributeError):
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except (OSError, AttributeError):
            pass


_marcar_dpi_consciente()

user32 = ctypes.windll.user32
kernel32 = ctypes.windll.kernel32


def _directorio_raiz_proyecto() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


ROOT = _directorio_raiz_proyecto()
TEMPLATE_DIR = ROOT / "templates"
FLUJO_DEFAULT = ROOT / "flujo_indigo_historias.txt"
# Si existe, se ejecuta una sola vez antes del bucle de cédulas (E.S.E. + menú hasta la pantalla de consulta).
FLUJO_INICIO_DEFAULT = ROOT / "flujo_indigo_inicio.txt"
SALIDAS_DIR = ROOT / "salidas"
# Coordenadas del clic derecho en el listado de folios (plantilla 5); se crea con --calibrar-folios.
CALIBRACION_FOLIOS_DEFAULT = ROOT / "indigo_calibracion_folios.json"

# Misma idea que «Documentos Usuarios x ESE…» del modelo de folios: una subcarpeta por E.S.E. con el Excel de cédulas.
INSUMOS_ROOT = ROOT / "Documentos Usuarios x ESE INDIGO folios de historias clinicas"
# Catálogo: A = código carpeta; D = nº opción del desplegable «motivo» (1 = primera); E = texto de esa opción (se escribe);
# F = observación.
MOTIVO_OBS_ESE_XLSX_DEFAULT = ROOT / "Documentos Usuarios x ESE INDIGO" / "Opciones Motivo de consultas x ESE.xlsx"
MOTIVO_OBS_ESE_XLSX_FOLIOS = (
    ROOT
    / "Documentos Usuarios x ESE INDIGO folios de historias clinicas"
    / "Opciones Motivo de consultas x ESE.xlsx"
)

MATCH_THRESHOLD = 0.48
# Guardar PNG de depuración (plantilla 10 / modal Imprimir) con similitud en banda; no más de 1 cada INDIGO_IMP10_DEBUG_INTERVAL s.
_imp10_debug_last_save = 0.0

TEMPLATE_SCALES = (0.78, 0.86, 0.92, 1.0, 1.08, 1.16, 1.24)
# Plantilla «0» a pantalla completa suele ser mayor que el ROI interno: hacen falta escalas < 0.78.
TEMPLATE_SCALES_VENTANA = (
    0.32,
    0.36,
    0.40,
    0.44,
    0.48,
    0.52,
    0.56,
    0.60,
    0.64,
    0.68,
    0.72,
    0.76,
) + TEMPLATE_SCALES

# Evita spam en encontrar_hwnd_indigo al reintentar cada ~1 s.
_log_fallback_plantilla0_emitido = False
_log_segunda_pasada_ventana_emitido = False
# Tras plantilla 9 (1.er «Imprimir»): evita confundir el mismo PNG con plantilla 10.
_ts_clic_imprimir_9: float = 0.0
_xy_imprimir_9_en_captura: tuple[int, int] | None = None
# Ventana donde hacer clic en plantilla 10 (Vie embebido o diálogo «Imprimir» #32770).
_hwnd_destino_imprimir_10: int | None = None
# Pasos sin traer IndiGO al frente (evita cerrar/perder foco del diálogo «Imprimir» / Guardar PDF).
_pasos_sin_traer_indigo: int = 0
# Tras plantilla 9 hasta terminar 10: no restaurar IndiGO (evita traer Cursor/consola al frente).
_bloqueo_restaurar_indigo_impresion: bool = False
# Microsoft Print to PDF a veces abre «Guardar impresión como» sin mostrar el 2.º Imprimir.
_omitir_clic_plantilla_10: bool = False
# Cédula (columna A) de la iteración actual — para prefijar el PDF en cuanto abre «Guardar impresión como».
_cedula_pdf_iteracion: str = ""
# ROI relativas al cliente al buscar la ventana por plantilla (centro / formulario Paciente).
ROIS_BUSQUEDA_VENTANA_POR_PLANTILLA = (
    (0.0, 0.0, 1.0, 1.0),
    (0.0, 0.12, 1.0, 1.0),
    (0.0, 0.18, 1.0, 1.0),
    (0.04, 0.22, 0.96, 0.98),
    (0.02, 0.28, 0.98, 0.95),
)
# Logo / marca «Vie EHR» (barra o esquina superior derecha): plantillas con «vie»+«ehr» en el nombre usan esto primero.
ROIS_BUSQUEDA_VENTANA_VIE_EHR = (
    (0.52, 0.0, 1.0, 0.12),
    (0.48, 0.0, 1.0, 0.18),
    (0.42, 0.0, 1.0, 0.26),
    (0.0, 0.0, 1.0, 0.14),
)

_indigo_easyocr_reader: object | None = None

if (os.environ.get("INDIGO_PYAUTOGUI_FAILSAFE") or "").strip().lower() in ("0", "false", "no", "off"):
    pyautogui.FAILSAFE = False
else:
    pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.22
_PYAUTOGUI_FAILSAFE_MARGEN_PX = 14


def _clamp_xy_evita_esquina_failsafe_pantalla(cx: int, cy: int) -> tuple[int, int]:
    try:
        sw, sh = pyautogui.size()
    except Exception:
        return cx, cy
    m = _PYAUTOGUI_FAILSAFE_MARGEN_PX
    return max(m, min(sw - 1 - m, int(cx))), max(m, min(sh - 1 - m, int(cy)))


MOUSEEVENTF_LEFTDOWN = 0x0002
MOUSEEVENTF_LEFTUP = 0x0004
MOUSEEVENTF_RIGHTDOWN = 0x0008
MOUSEEVENTF_RIGHTUP = 0x0010
MOUSEEVENTF_MOVE = 0x0001
MOUSEEVENTF_ABSOLUTE = 0x8000
MOUSEEVENTF_VIRTUALDESK = 0x4000
INPUT_MOUSE = 0
ULONG_PTR = ctypes.c_ulonglong if ctypes.sizeof(ctypes.c_void_p) == 8 else ctypes.c_ulong


class _POINT(ctypes.Structure):
    _fields_ = [("x", wintypes.LONG), ("y", wintypes.LONG)]


class _MOUSEINPUT(ctypes.Structure):
    _fields_ = [
        ("dx", wintypes.LONG),
        ("dy", wintypes.LONG),
        ("mouseData", wintypes.DWORD),
        ("dwFlags", wintypes.DWORD),
        ("time", wintypes.DWORD),
        ("dwExtraInfo", ULONG_PTR),
    ]


class _INPUTUNION(ctypes.Union):
    _fields_ = [("mi", _MOUSEINPUT)]


class _INPUT(ctypes.Structure):
    _anonymous_ = ("u",)
    _fields_ = [("type", wintypes.DWORD), ("u", _INPUTUNION)]


def _cursor_pos_pantalla() -> tuple[int, int]:
    pt = _POINT()
    user32.GetCursorPos(ctypes.byref(pt))
    return int(pt.x), int(pt.y)


def _pantalla_a_coords_sendinput_absolutas(cx: int, cy: int) -> tuple[int, int]:
    sw = user32.GetSystemMetrics(0)
    sh = user32.GetSystemMetrics(1)
    ax = int(cx * 65535 / max(1, sw - 1))
    ay = int(cy * 65535 / max(1, sh - 1))
    return ax, ay


def _sendinput_mouse(flags: int, cx: int = 0, cy: int = 0) -> None:
    ax, ay = (cx, cy)
    if flags & MOUSEEVENTF_ABSOLUTE:
        ax, ay = _pantalla_a_coords_sendinput_absolutas(cx, cy)
    inp = _INPUT(
        type=INPUT_MOUSE,
        u=_INPUTUNION(
            mi=_MOUSEINPUT(
                dx=ax,
                dy=ay,
                mouseData=0,
                dwFlags=flags,
                time=0,
                dwExtraInfo=0,
            )
        ),
    )
    user32.SendInput(1, ctypes.byref(inp), ctypes.sizeof(_INPUT))


def _mover_cursor_a_pantalla(cx: int, cy: int) -> tuple[int, int, bool]:
    """Mueve el puntero (SendInput + SetCursorPos + PyAutoGUI) y comprueba si realmente se movió."""
    cx, cy = _clamp_xy_evita_esquina_failsafe_pantalla(cx, cy)
    antes = _cursor_pos_pantalla()
    if _bool_env("INDIGO_MOUSE_USAR_SENDINPUT", True):
        _sendinput_mouse(MOUSEEVENTF_MOVE | MOUSEEVENTF_ABSOLUTE | MOUSEEVENTF_VIRTUALDESK, cx, cy)
        time.sleep(0.05)
    user32.SetCursorPos(int(cx), int(cy))
    time.sleep(0.04)
    dur = _float_env("INDIGO_FOLIOS_MOVER_DURACION_SEG", 0.25)
    pyautogui.moveTo(int(cx), int(cy), duration=max(0.05, dur))
    time.sleep(0.08)
    despues = _cursor_pos_pantalla()
    ok = abs(despues[0] - cx) + abs(despues[1] - cy) <= 20
    if not ok:
        print(
            f"Diagnóstico ratón: antes={antes} → objetivo=({cx},{cy}) → después={despues} "
            f"(no se movió lo suficiente)."
        )
        print(
            "Si usa AnyDesk/RDP: ejecute el RPA en la sesión local de esa PC (monitor físico), "
            "con «Controlar equipo remoto» activo, o no en modo solo lectura."
        )
    return despues[0], despues[1], ok


def _clic_derecho_en_pantalla(cx: int, cy: int) -> bool:
    """Clic derecho en pantalla. Devuelve False si el cursor no llegó al destino (p. ej. AnyDesk)."""
    cx, cy = _clamp_xy_evita_esquina_failsafe_pantalla(cx, cy)
    ax, ay, ok = _mover_cursor_a_pantalla(cx, cy)
    if not ok:
        return False
    time.sleep(0.06)
    if _bool_env("INDIGO_MOUSE_USAR_SENDINPUT", True):
        _sendinput_mouse(MOUSEEVENTF_RIGHTDOWN | MOUSEEVENTF_ABSOLUTE | MOUSEEVENTF_VIRTUALDESK, ax, ay)
        time.sleep(_float_env("INDIGO_CLICK_WIN32_PAUSE_DOWN_UP", 0.03))
        _sendinput_mouse(MOUSEEVENTF_RIGHTUP | MOUSEEVENTF_ABSOLUTE | MOUSEEVENTF_VIRTUALDESK, ax, ay)
        return True
    if _bool_env("INDIGO_FOLIOS_CLIC_PYAUTOGUI", True) or _bool_env("INDIGO_CLICK_PYAUTOGUI", False):
        pyautogui.rightClick(ax, ay)
        return True
    user32.mouse_event(MOUSEEVENTF_RIGHTDOWN, 0, 0, 0, 0)
    time.sleep(_float_env("INDIGO_CLICK_WIN32_PAUSE_DOWN_UP", 0.03))
    user32.mouse_event(MOUSEEVENTF_RIGHTUP, 0, 0, 0, 0)
    return True


def _hwnd_es_descendiente_o_igual(hwnd_hijo: int, hwnd_ancestro: int) -> bool:
    p = int(hwnd_hijo)
    raiz = int(hwnd_ancestro)
    while p:
        if p == raiz:
            return True
        p = win32gui.GetParent(p) or 0
    return False


def _lparam_punto_cliente(x: int, y: int) -> int:
    return ((int(y) & 0xFFFF) << 16) | (int(x) & 0xFFFF)


def _clic_derecho_hwnd_en_pantalla(
    hwnd_raiz: int,
    cx: int,
    cy: int,
    etiqueta: str = "",
) -> bool:
    """Clic derecho por mensajes Win32 (no mueve el cursor; útil con AnyDesk/RDP)."""
    if not _bool_env("INDIGO_FOLIOS_CLIC_HWND", True):
        return False
    restaurar_y_traer_al_frente(hwnd_raiz)
    time.sleep(0.06)
    hwnd_tgt = win32gui.WindowFromPoint((int(cx), int(cy)))
    if not hwnd_tgt or not _hwnd_es_descendiente_o_igual(hwnd_tgt, hwnd_raiz):
        hwnd_tgt = hwnd_raiz
    try:
        cx_c, cy_c = win32gui.ScreenToClient(hwnd_tgt, (int(cx), int(cy)))
    except Exception:
        hwnd_tgt = hwnd_raiz
        cx_c, cy_c = win32gui.ScreenToClient(hwnd_tgt, (int(cx), int(cy)))
    lp = _lparam_punto_cliente(cx_c, cy_c)
    usar_send = _bool_env("INDIGO_FOLIOS_HWND_SENDMESSAGE", False)
    enviar = win32gui.SendMessage if usar_send else win32gui.PostMessage
    try:
        enviar(hwnd_tgt, win32con.WM_MOUSEMOVE, 0, lp)
        time.sleep(0.02)
        enviar(hwnd_tgt, win32con.WM_RBUTTONDOWN, win32con.MK_RBUTTON, lp)
        time.sleep(_float_env("INDIGO_FOLIOS_HWND_CLICK_PAUSE", 0.06))
        enviar(hwnd_tgt, win32con.WM_RBUTTONUP, 0, lp)
        if _bool_env("INDIGO_FOLIOS_WM_CONTEXTMENU", True):
            time.sleep(0.03)
            enviar(hwnd_tgt, win32con.WM_CONTEXTMENU, hwnd_tgt, lp)
    except Exception as ex:
        print(f"{etiqueta}: clic derecho por HWND falló ({ex})")
        return False
    pref = "SendMessage" if usar_send else "PostMessage"
    print(
        f"{etiqueta}: clic derecho por ventana ({pref}, hwnd={hwnd_tgt}, "
        f"cliente {cx_c},{cy_c}) — el ratón físico puede no moverse"
    )
    return True


def _clic_derecho_en_coordenadas_folio(
    hwnd: int,
    cx: int,
    cy: int,
    etiqueta: str,
) -> bool:
    """Mueve el ratón y hace clic derecho; si falla (p. ej. AnyDesk), reintenta por mensaje HWND."""
    cx, cy = _clamp_xy_evita_esquina_failsafe_pantalla(cx, cy)
    restaurar_y_traer_al_frente(hwnd)
    time.sleep(0.05)
    if _bool_env("INDIGO_FOLIOS_MOVER_RATON_VISIBLE", True):
        if _clic_derecho_en_pantalla(cx, cy):
            _programar_sin_traer_indigo_tras_menu_folios()
            return True
        print(
            f"Flujo: {etiqueta} — el ratón no llegó a ({cx},{cy}); "
            "se intenta clic derecho por mensaje de ventana (INDIGO_FOLIOS_CLIC_HWND)."
        )
    ok = _clic_derecho_hwnd_en_pantalla(hwnd, cx, cy, etiqueta)
    if ok:
        _programar_sin_traer_indigo_tras_menu_folios()
    return ok


def _captura_monitor_principal_bgr() -> tuple[np.ndarray, int, int]:
    with mss.mss() as sct:
        mon = sct.monitors[1]
    lo, to = int(mon["left"]), int(mon["top"])
    ww, hh = int(mon["width"]), int(mon["height"])
    return capturar_ventana_bgr(lo, to, ww, hh), lo, to


def _clic_izquierdo_menu_contexto_en_pantalla(cx: int, cy: int, etiqueta: str = "") -> bool:
    """Clic en ítem de menú contextual sin traer IndiGO al frente (PostMessage a #32768 o coordenadas)."""
    cx, cy = _clamp_xy_evita_esquina_failsafe_pantalla(cx, cy)
    hwnd_pt = win32gui.WindowFromPoint((int(cx), int(cy)))
    cls = (win32gui.GetClassName(hwnd_pt) or "").strip() if hwnd_pt else ""
    if hwnd_pt and cls in ("#32768", "SysListView32"):
        try:
            cx_c, cy_c = win32gui.ScreenToClient(hwnd_pt, (int(cx), int(cy)))
            lp = _lparam_punto_cliente(cx_c, cy_c)
            win32gui.PostMessage(hwnd_pt, win32con.WM_LBUTTONDOWN, win32con.MK_LBUTTON, lp)
            time.sleep(_float_env("INDIGO_MENU_FOLIO_CLIC_PAUSE", 0.05))
            win32gui.PostMessage(hwnd_pt, win32con.WM_LBUTTONUP, 0, lp)
            print(
                f"{etiqueta}: clic en menú contextual (hwnd={hwnd_pt}, clase {cls!r}, "
                f"cliente {cx_c},{cy_c})"
            )
            return True
        except Exception as ex:
            print(f"{etiqueta}: PostMessage en menú falló ({ex}); se prueba clic en pantalla.")
    if _bool_env("INDIGO_MENU_FOLIO_CLIC_PYAUTOGUI", True):
        pyautogui.click(int(cx), int(cy))
        print(f"{etiqueta}: clic PyAutoGUI en ({cx},{cy})")
        return True
    user32.SetCursorPos(int(cx), int(cy))
    time.sleep(0.05)
    user32.mouse_event(MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    time.sleep(0.03)
    user32.mouse_event(MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    return True


def _mover_y_clic_pantalla(cx: int, cy: int, boton: str = "left") -> None:
    """Clic en pantalla; por defecto Win32 (evita fail-safe de PyAutoGUI al mover el ratón)."""
    cx, cy = _clamp_xy_evita_esquina_failsafe_pantalla(cx, cy)
    if _bool_env("INDIGO_CLICK_PYAUTOGUI", False):
        pyautogui.moveTo(cx, cy, duration=0.06)
        if (boton or "").lower() == "right":
            pyautogui.rightClick(cx, cy)
        else:
            pyautogui.click(cx, cy)
        return
    settle = _float_env("INDIGO_CLICK_WIN32_SETTLE_BEFORE_CLICK", 0.09)
    gap_down_up = _float_env("INDIGO_CLICK_WIN32_PAUSE_DOWN_UP", 0.03)
    user32.SetCursorPos(int(cx), int(cy))
    time.sleep(max(0.0, settle))
    if (boton or "").lower() == "right":
        user32.mouse_event(MOUSEEVENTF_RIGHTDOWN, 0, 0, 0, 0)
        time.sleep(max(0.0, gap_down_up))
        user32.mouse_event(MOUSEEVENTF_RIGHTUP, 0, 0, 0, 0)
    else:
        user32.mouse_event(MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
        time.sleep(max(0.0, gap_down_up))
        user32.mouse_event(MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)


class RECT(ctypes.Structure):
    _fields_ = [
        ("left", wintypes.LONG),
        ("top", wintypes.LONG),
        ("right", wintypes.LONG),
        ("bottom", wintypes.LONG),
    ]


def _norm_etiqueta(s: str) -> str:
    s = (s or "").strip().lower()
    return "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")


def _cv2_imread_path(path: Path) -> np.ndarray | None:
    try:
        data = path.read_bytes()
    except OSError:
        return None
    if not data:
        return None
    arr = np.frombuffer(data, dtype=np.uint8)
    return cv2.imdecode(arr, cv2.IMREAD_UNCHANGED)


def _enum_windows() -> list[tuple[int, str]]:
    out: list[tuple[int, str]] = []

    @ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
    def cb(hwnd, _):
        if user32.IsWindowVisible(hwnd):
            buf = ctypes.create_unicode_buffer(512)
            user32.GetWindowTextW(hwnd, buf, 512)
            t = buf.value
            if t:
                out.append((hwnd, t))
        return True

    user32.EnumWindows(cb, 0)
    return out


def _rect_plausible(left: int, top: int, right: int, bottom: int) -> bool:
    w, h = right - left, bottom - top
    if w < 80 or h < 80:
        return False
    if left < -5000 or top < -5000:
        return False
    return True


def _rect_actual(hwnd: int) -> tuple[int, int, int, int]:
    r = RECT()
    user32.GetWindowRect(hwnd, ctypes.byref(r))
    return int(r.left), int(r.top), int(r.right), int(r.bottom)


def _activar_bloqueo_foco_impresion_9_10() -> None:
    """Impide restaurar_y_traer_al_frente(IndiGO) mientras debe estar al frente Imprimir/Vie."""
    global _bloqueo_restaurar_indigo_impresion, _pasos_sin_traer_indigo
    _bloqueo_restaurar_indigo_impresion = True
    _programar_sin_traer_indigo_para_flujo_impresion_pdf()


def _desactivar_bloqueo_foco_impresion_9_10() -> None:
    global _bloqueo_restaurar_indigo_impresion
    _bloqueo_restaurar_indigo_impresion = False


def restaurar_y_traer_al_frente(hwnd: int) -> None:
    if _bloqueo_restaurar_indigo_impresion and _bool_env("INDIGO_BLOQUEO_FOCO_IMPRESION", True):
        return
    tit_chk = (win32gui.GetWindowText(int(hwnd)) or "").strip()
    if _hwnd_es_ide_cursor_o_editor(int(hwnd), tit_chk):
        print(
            f"Flujo: no se trae al frente IDE/consola «{tit_chk[:60]}» (hwnd={hwnd}); "
            "use la ventana Vie/IndiGO o INDIGO_HWND."
        )
        return
    for _ in range(6):
        left, top, right, bottom = _rect_actual(hwnd)
        if not _rect_plausible(left, top, right, bottom) or user32.IsIconic(hwnd):
            user32.ShowWindow(hwnd, win32con.SW_RESTORE)
            time.sleep(0.35)
            user32.ShowWindow(hwnd, win32con.SW_SHOWNORMAL)
            time.sleep(0.2)
        user32.BringWindowToTop(hwnd)
        user32.SetForegroundWindow(hwnd)
        time.sleep(0.2)
        left, top, right, bottom = _rect_actual(hwnd)
        if _rect_plausible(left, top, right, bottom):
            return
        user32.SetWindowPos(
            hwnd,
            win32con.HWND_TOP,
            80,
            80,
            1200,
            820,
            win32con.SWP_SHOWWINDOW,
        )
        time.sleep(0.25)
    raise RuntimeError("No se pudo dejar la ventana IndiGO visible en pantalla.")


def activar_ventana_modal(hwnd: int) -> None:
    if user32.IsIconic(hwnd):
        user32.ShowWindow(hwnd, win32con.SW_RESTORE)
        time.sleep(0.2)
    user32.BringWindowToTop(hwnd)
    user32.SetForegroundWindow(hwnd)
    time.sleep(0.12)


def _hwnd_ancestro_o_mismo(hwnd: int, ancestro: int) -> bool:
    p = int(hwnd or 0)
    objetivo = int(ancestro or 0)
    while p:
        if p == objetivo:
            return True
        p = int(win32gui.GetParent(p) or 0)
    return False


def _ventana_esta_al_frente(hwnd: int) -> bool:
    if not hwnd or not user32.IsWindow(hwnd):
        return False
    fg = int(user32.GetForegroundWindow() or 0)
    return fg == hwnd or _hwnd_ancestro_o_mismo(fg, hwnd)


def _force_foreground_window(hwnd: int) -> bool:
    """Activa una ventana aunque otro modal (p. ej. Vie) esté encima."""
    if not hwnd or not user32.IsWindow(hwnd):
        return False
    if _ventana_esta_al_frente(hwnd):
        return True
    if user32.IsIconic(hwnd):
        user32.ShowWindow(hwnd, win32con.SW_RESTORE)
        time.sleep(0.15)
    if _bool_env("INDIGO_FORCE_FG_ALT", False):
        try:
            pyautogui.press("alt")
            time.sleep(0.05)
        except Exception:
            pass
    fg = int(user32.GetForegroundWindow() or 0)
    fg_tid = ctypes.c_ulong(0)
    tgt_tid = ctypes.c_ulong(0)
    user32.GetWindowThreadProcessId(fg, ctypes.byref(fg_tid))
    user32.GetWindowThreadProcessId(hwnd, ctypes.byref(tgt_tid))
    attached = False
    if fg_tid.value and tgt_tid.value and fg_tid.value != tgt_tid.value:
        attached = bool(user32.AttachThreadInput(fg_tid.value, tgt_tid.value, True))
    user32.ShowWindow(hwnd, win32con.SW_SHOW)
    user32.BringWindowToTop(hwnd)
    user32.SetForegroundWindow(hwnd)
    if attached:
        user32.AttachThreadInput(fg_tid.value, tgt_tid.value, False)
    time.sleep(0.12)
    return _ventana_esta_al_frente(hwnd)


def _enviar_ventana_detras(hwnd: int) -> None:
    flags = win32con.SWP_NOMOVE | win32con.SWP_NOSIZE | win32con.SWP_NOACTIVATE
    user32.SetWindowPos(int(hwnd), win32con.HWND_BOTTOM, 0, 0, 0, 0, flags)


def _enviar_modales_vie_impresion_detras(hwnd_guardar: int) -> int:
    """El modal «Consultar / Imprimir Historias» suele tapar «Guardar impresión como»."""
    n = 0
    for h, tit in _enum_hwnds_modal_vie_consultar_imprimir():
        if int(h) == int(hwnd_guardar):
            continue
        _enviar_ventana_detras(int(h))
        n += 1
        print(
            f"Flujo: modal Vie enviado detrás del Guardar PDF — «{tit}» (hwnd={h})"
        )
    for h, tit in _enum_hwnds_dialogo_impresion_sistema():
        if int(h) == int(hwnd_guardar):
            continue
        nt = _norm_etiqueta(tit)
        if nt in ("imprimir", "print", "imprimiendo", "printing"):
            _enviar_ventana_detras(int(h))
            print(f"Flujo: diálogo impresión enviado detrás — «{tit}» (hwnd={h})")
    return n


def _restaurar_hwnd_si_minimizado(hwnd: int, etiqueta: str = "") -> bool:
    """Restaura ventanas en bandeja/minimizadas para que el RPA pueda interactuar."""
    if hwnd <= 0 or not user32.IsWindow(hwnd):
        return False
    cambio = False
    if user32.IsIconic(hwnd):
        user32.ShowWindow(hwnd, win32con.SW_RESTORE)
        time.sleep(0.2)
        cambio = True
    left, top, right, bottom = _rect_actual(hwnd)
    if not _rect_plausible(left, top, right, bottom):
        user32.ShowWindow(hwnd, win32con.SW_SHOWNORMAL)
        time.sleep(0.15)
        cambio = True
    if cambio and etiqueta:
        tit = (win32gui.GetWindowText(hwnd) or "").strip()
        print(f"Flujo: ventana restaurada ({etiqueta}) — «{tit[:55]}» (hwnd={hwnd})")
    return cambio


def _hwnd_top_level_utilizable(hwnd: int) -> bool:
    if hwnd <= 0 or not user32.IsWindow(hwnd):
        return False
    if user32.IsIconic(hwnd):
        return True
    return bool(user32.IsWindowVisible(hwnd))


def _activar_dialogo_guardar_pdf_forzado(hwnd_guardar: int) -> bool:
    """Pone «Guardar impresión como» delante del modal Vie y le da foco real."""
    if not hwnd_guardar or not user32.IsWindow(hwnd_guardar):
        return False
    tit_g = (win32gui.GetWindowText(int(hwnd_guardar)) or "").strip()
    if _titulo_es_dialogo_imprimiendo_progreso(tit_g):
        alterno = _seleccionar_hwnd_dialogo_guardar_pdf()
        if alterno is not None and int(alterno) != int(hwnd_guardar):
            print(
                f"Flujo: aviso — se pidió activar «Imprimiendo» (hwnd={hwnd_guardar}); "
                f"se usa «Guardar impresión como» hwnd={alterno}"
            )
            hwnd_guardar = int(alterno)
    _restaurar_hwnd_si_minimizado(int(hwnd_guardar), "Guardar PDF")
    flags = win32con.SWP_NOMOVE | win32con.SWP_NOSIZE | win32con.SWP_SHOWWINDOW
    if _bool_env("INDIGO_GUARDAR_PDF_ENVIAR_VIE_DETRAS", True):
        _enviar_modales_vie_impresion_detras(hwnd_guardar)
    _enviar_imprimiendo_detras_del_guardar(int(hwnd_guardar))
    user32.SetWindowPos(int(hwnd_guardar), win32con.HWND_TOPMOST, 0, 0, 0, 0, flags)
    time.sleep(0.06)
    ok = _force_foreground_window(int(hwnd_guardar))
    user32.SetWindowPos(
        int(hwnd_guardar),
        win32con.HWND_NOTOPMOST,
        0,
        0,
        0,
        0,
        win32con.SWP_NOMOVE | win32con.SWP_NOSIZE | win32con.SWP_SHOWWINDOW,
    )
    user32.BringWindowToTop(int(hwnd_guardar))
    time.sleep(0.08)
    tit = (win32gui.GetWindowText(int(hwnd_guardar)) or "").strip()
    if ok:
        print(f"Flujo: Guardar PDF al frente — «{tit}» (hwnd={hwnd_guardar})")
    else:
        print(
            f"Flujo: aviso — no se confirmó foco en Guardar PDF «{tit}» (hwnd={hwnd_guardar}); "
            "se reintenta escribir la cédula."
        )
    return ok


def rect_pantalla(hwnd: int) -> tuple[int, int, int, int]:
    r = RECT()
    user32.GetWindowRect(hwnd, ctypes.byref(r))
    return int(r.left), int(r.top), int(r.right), int(r.bottom)


def tamano_cliente(hwnd: int) -> tuple[int, int]:
    r = RECT()
    user32.GetClientRect(hwnd, ctypes.byref(r))
    return int(r.right - r.left), int(r.bottom - r.top)


def cliente_a_pantalla(hwnd: int, cx: int, cy: int) -> tuple[int, int]:
    pt = wintypes.POINT(cx, cy)
    user32.ClientToScreen(hwnd, ctypes.byref(pt))
    return int(pt.x), int(pt.y)


def capturar_ventana_bgr(left: int, top: int, width: int, height: int) -> np.ndarray:
    if width < 2 or height < 2:
        raise RuntimeError("Rectángulo de ventana inválido.")
    with mss.MSS() as sct:
        region = {"left": left, "top": top, "width": width, "height": height}
        shot = np.array(sct.grab(region))
    return cv2.cvtColor(shot, cv2.COLOR_BGRA2BGR)


def _origen_y_captura_para_match_hwnd(hwnd: int, etiqueta_log: str = "") -> tuple[np.ndarray, int, int]:
    pref = f"{etiqueta_log}: " if etiqueta_log else ""
    for intento in range(6):
        left, top, right, bottom = rect_pantalla(hwnd)
        w, h = right - left, bottom - top
        plausible = w >= 32 and h >= 32 and _rect_plausible(left, top, right, bottom)
        if plausible:
            return capturar_ventana_bgr(left, top, w, h), left, top
        print(f"{pref}rect anómalo ({w}×{h}); restaurar intento {intento + 1}/6…")
        try:
            restaurar_y_traer_al_frente(hwnd)
        except RuntimeError:
            activar_ventana_modal(hwnd)
        time.sleep(0.28)
    with mss.MSS() as sct:
        mon = sct.monitors[1]
    lo, to = int(mon["left"]), int(mon["top"])
    ww, hh = int(mon["width"]), int(mon["height"])
    print(f"{pref}búsqueda en monitor principal ({ww}×{hh}).")
    return capturar_ventana_bgr(lo, to, ww, hh), lo, to


def _recorte_roi(hay: np.ndarray, roi_rel: tuple[float, float, float, float]) -> tuple[np.ndarray, int, int]:
    H, W = hay.shape[:2]
    x0, y0, x1, y1 = roi_rel
    rx0 = max(0, int(W * x0))
    ry0 = max(0, int(H * y0))
    rx1 = min(W, int(W * x1))
    ry1 = min(H, int(H * y1))
    if rx1 - rx0 < 32 or ry1 - ry0 < 32:
        return hay, 0, 0
    return hay[ry0:ry1, rx0:rx1], rx0, ry0


def _tpl_gray_encajada_en_roi(tpl_g: np.ndarray, rw: int, rh: int) -> np.ndarray:
    """Si la plantilla no cabe en el ROI, la reduce (matchTemplate exige tpl < imagen en ambos lados)."""
    th0, tw0 = tpl_g.shape[:2]
    if rh <= 16 or rw <= 16:
        return tpl_g
    if tw0 < rw - 1 and th0 < rh - 1:
        return tpl_g
    shrink = min((rw - 2) / max(tw0, 8), (rh - 2) / max(th0, 8), 1.0) * 0.995
    if shrink >= 1.0:
        return tpl_g
    tw1 = max(8, int(round(tw0 * shrink)))
    th1 = max(8, int(round(th0 * shrink)))
    return cv2.resize(tpl_g, (tw1, th1), interpolation=cv2.INTER_AREA)


def _mejor_match_multiescala(
    hay_bgr: np.ndarray,
    tpl_bgr: np.ndarray,
    roi_rel: tuple[float, float, float, float],
    orden_candidatos: str,
    min_score: float | None,
    *,
    cx_min_frac: float | None = None,
    cx_max_frac: float | None = None,
    cy_max_frac: float | None = None,
    cy_min_frac: float | None = None,
    exclude_xy: tuple[int, int] | None = None,
    exclude_radio_px: int = 0,
    escalas: tuple[float, ...] | None = None,
) -> tuple[int, int, float, int, int] | None:
    thr = MATCH_THRESHOLD if min_score is None else float(min_score)
    filtro_pos = (
        cx_min_frac is not None
        or cx_max_frac is not None
        or cy_max_frac is not None
        or cy_min_frac is not None
        or (exclude_xy is not None and exclude_radio_px > 0)
    )
    h_hay, w_hay = hay_bgr.shape[0], hay_bgr.shape[1]
    roi, ox, oy = _recorte_roi(hay_bgr, roi_rel)
    roi_g = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
    tpl_g = cv2.cvtColor(tpl_bgr, cv2.COLOR_BGR2GRAY)
    rh, rw = roi_g.shape[:2]
    tpl_g = _tpl_gray_encajada_en_roi(tpl_g, rw, rh)
    th0, tw0 = tpl_g.shape[:2]
    mejor: tuple[float, int, int, int, int] | None = None

    for escala in escalas if escalas is not None else TEMPLATE_SCALES:
        tw = max(8, int(round(tw0 * escala)))
        th = max(8, int(round(th0 * escala)))
        if tw >= roi_g.shape[1] or th >= roi_g.shape[0]:
            continue
        interp = cv2.INTER_AREA if escala < 1.0 else cv2.INTER_LINEAR
        t = cv2.resize(tpl_g, (tw, th), interpolation=interp)
        res = cv2.matchTemplate(roi_g, t, cv2.TM_CCOEFF_NORMED)
        _mn, max_v, _min_loc, max_loc = cv2.minMaxLoc(res)
        if max_v < thr * 0.85:
            continue
        umbral_pico = max(thr * 0.92, max_v * 0.96)
        loc_y, loc_x = np.where(res >= umbral_pico)
        if loc_x.size == 0:
            candidatos = [(int(max_loc[0]), int(max_loc[1]), float(max_v))]
        else:
            candidatos = [(int(xi), int(yi), float(res[yi, xi])) for yi, xi in zip(loc_y, loc_x)]
        if filtro_pos:
            candidatos.sort(key=lambda p: (-p[2], -p[0]) if orden_candidatos == "right" else (-p[2], p[0]))
        elif orden_candidatos == "right":
            candidatos.sort(key=lambda p: (-p[2], -p[0]))
        elif orden_candidatos == "top":
            candidatos.sort(key=lambda p: (p[1], p[0], -p[2]))
        else:
            candidatos.sort(key=lambda p: (p[0], -p[2]))
        validos: list[tuple[int, int, float]] = []
        for lx, ly, sc_c in candidatos:
            cx_abs = ox + lx + tw // 2
            cy_abs = oy + ly + th // 2
            if cx_min_frac is not None and cx_abs < w_hay * cx_min_frac:
                continue
            if cx_max_frac is not None and cx_abs >= w_hay * cx_max_frac:
                continue
            if cy_max_frac is not None and cy_abs >= h_hay * cy_max_frac:
                continue
            if cy_min_frac is not None and cy_abs < h_hay * cy_min_frac:
                continue
            if exclude_xy is not None and exclude_radio_px > 0:
                ex, ey = exclude_xy
                if (cx_abs - ex) ** 2 + (cy_abs - ey) ** 2 < exclude_radio_px**2:
                    continue
            validos.append((cx_abs, cy_abs, sc_c))
        if not validos:
            continue
        if orden_candidatos == "bottom_right":
            max_sc = max(v[2] for v in validos)
            pool = [v for v in validos if v[2] >= max_sc - 0.05]
            cx, cy, sc = max(pool, key=lambda v: (v[1], v[0]))  # más abajo, más a la derecha
        elif orden_candidatos == "right":
            cx, cy, sc = max(validos, key=lambda v: (-v[2], -v[0]))
        elif orden_candidatos == "top":
            cx, cy, sc = min(validos, key=lambda v: (v[1], -v[2]))
        else:
            cx, cy, sc = min(validos, key=lambda v: (v[0], -v[2]))
        if mejor is None or sc > mejor[0]:
            mejor = (sc, cx, cy, tw, th)
        elif abs(sc - mejor[0]) < 0.02:
            if orden_candidatos == "bottom_right":
                if cy > mejor[2] or (cy == mejor[2] and cx > mejor[1]):
                    mejor = (sc, cx, cy, tw, th)
            elif orden_candidatos == "top":
                if cy < mejor[2]:
                    mejor = (sc, cx, cy, tw, th)
            elif orden_candidatos == "right":
                if cx > mejor[1]:
                    mejor = (sc, cx, cy, tw, th)
            else:
                if cx < mejor[1]:
                    mejor = (sc, cx, cy, tw, th)

    if mejor is None:
        return None
    sc, cx, cy, tw, th = mejor
    if sc < thr:
        return None
    return cx, cy, sc, tw, th


def clic_plantilla_en_hwnd(
    hwnd: int,
    tpl_path: Path,
    rois: list[tuple[float, float, float, float]],
    orden: str,
    min_score: float | None,
    etiqueta_log: str,
    *,
    pantalla_completa_si_falla: bool = False,
    umbrales_pantalla_completa: tuple[float, ...] = (0.30, 0.24, 0.20),
    fraccion_click_en_plantilla: tuple[float, float] | None = None,
    click_nudge_px: tuple[int, int] | None = None,
    boton: str = "left",
    cy_max_frac: float | None = None,
    cy_min_frac: float | None = None,
    cx_min_frac: float | None = None,
    exclude_xy: tuple[int, int] | None = None,
    exclude_radio_px: int = 0,
    tpl_bgr_override: np.ndarray | None = None,
) -> tuple[int, int, float]:
    if tpl_bgr_override is not None:
        tpl = tpl_bgr_override
    else:
        tpl = _cv2_imread_path(tpl_path)
    if tpl is None or tpl.size == 0:
        raise RuntimeError(f"No se pudo leer la plantilla {tpl_path}")
    if tpl.ndim == 3 and tpl.shape[2] == 4:
        tpl = cv2.cvtColor(tpl, cv2.COLOR_BGRA2BGR)
    pantalla, left, top = _origen_y_captura_para_match_hwnd(hwnd, etiqueta_log)
    hit = None
    for roi in rois:
        hit = _mejor_match_multiescala(
            pantalla,
            tpl,
            roi,
            orden,
            min_score,
            cx_min_frac=cx_min_frac,
            cx_max_frac=None,
            cy_max_frac=cy_max_frac,
            cy_min_frac=cy_min_frac,
            exclude_xy=exclude_xy,
            exclude_radio_px=exclude_radio_px,
            escalas=None,
        )
        if hit is not None:
            break
    if hit is None and pantalla_completa_si_falla:
        for thr in umbrales_pantalla_completa:
            hit = _mejor_match_multiescala(
                pantalla,
                tpl,
                (0.0, 0.0, 1.0, 1.0),
                orden,
                thr,
                cx_min_frac=cx_min_frac,
                cy_max_frac=cy_max_frac,
                cy_min_frac=cy_min_frac,
                exclude_xy=exclude_xy,
                exclude_radio_px=exclude_radio_px,
                escalas=None,
            )
            if hit is not None:
                print(f"{etiqueta_log}: match pantalla completa (umbral {thr:.2f}).")
                break
    if hit is None:
        raise RuntimeError(
            f"{etiqueta_log}: sin coincidencia de plantilla {tpl_path.name}. Revise ROI o el PNG."
        )
    cx_img, cy_img, sc, tw, th = hit
    if fraccion_click_en_plantilla is not None and tw >= 2 and th >= 2:
        fxp, fyp = fraccion_click_en_plantilla
        tl_x = int(round(cx_img - tw / 2.0))
        tl_y = int(round(cy_img - th / 2.0))
        ox = max(0, min(tw - 1, int(round((tw - 1) * fxp))))
        oy = max(0, min(th - 1, int(round((th - 1) * fyp))))
        cx_img, cy_img = tl_x + ox, tl_y + oy
    cx = left + int(cx_img)
    cy = top + int(cy_img)
    cx, cy = _clamp_xy_evita_esquina_failsafe_pantalla(cx, cy)
    if click_nudge_px is not None:
        ndx, ndy = int(click_nudge_px[0]), int(click_nudge_px[1])
        if ndx != 0 or ndy != 0:
            cx += ndx
            cy += ndy
            cx, cy = _clamp_xy_evita_esquina_failsafe_pantalla(cx, cy)
    suf = "clic derecho" if (boton or "").lower() == "right" else "clic"
    print(f"{etiqueta_log}: {suf} {tpl_path.name} (similitud {sc:.2f}) → ({cx},{cy})")
    _mover_y_clic_pantalla(cx, cy, boton)
    return cx, cy, sc


def _max_sim_ventana_entre_plantillas_ancla(hwnd: int, rutas_tpl: list[Path]) -> float:
    """Mayor similitud entre las plantillas usadas para anclar HWND (0 inicial, 1 cédula, …)."""
    roi_def = (0.02, 0.02, 0.98, 0.98)
    mejor = 0.0
    for tpl_path in rutas_tpl:
        rois_list = _rois_ampliados_plantilla_8_paciente(tpl_path.name, roi_def, roi_def)
        mejor = max(mejor, _max_similitud_plantilla_hwnd(hwnd, tpl_path, rois_list))
    return mejor


def _max_similitud_plantilla_haystack(
    hay_bgr: np.ndarray,
    tpl_path: Path,
    rois: list[tuple[float, float, float, float]] | tuple[tuple[float, float, float, float], ...],
) -> float:
    """Mayor TM_CCOEFF_NORMED entre ROIs y escalas sobre una captura ya obtenida."""
    tpl = _cv2_imread_path(tpl_path)
    if tpl is None or tpl.size == 0:
        return 0.0
    if tpl.ndim == 3 and tpl.shape[2] == 4:
        tpl = cv2.cvtColor(tpl, cv2.COLOR_BGRA2BGR)
    tpl_g = cv2.cvtColor(tpl, cv2.COLOR_BGR2GRAY)
    global_max = 0.0
    for roi_rel in rois:
        roi, _ox, _oy = _recorte_roi(hay_bgr, roi_rel)
        roi_g = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        rh, rw = roi_g.shape[:2]
        tpl_loc = _tpl_gray_encajada_en_roi(tpl_g, rw, rh)
        loc_th, loc_tw = tpl_loc.shape[:2]
        for escala in TEMPLATE_SCALES_VENTANA:
            tw = max(8, int(round(loc_tw * escala)))
            th = max(8, int(round(loc_th * escala)))
            if tw >= roi_g.shape[1] or th >= roi_g.shape[0]:
                continue
            interp = cv2.INTER_AREA if escala < 1.0 else cv2.INTER_LINEAR
            t = cv2.resize(tpl_loc, (tw, th), interpolation=interp)
            res = cv2.matchTemplate(roi_g, t, cv2.TM_CCOEFF_NORMED)
            _mn, max_v, _min_loc, _max_loc = cv2.minMaxLoc(res)
            global_max = max(global_max, float(max_v))
    return global_max


def _max_similitud_tpl_haystack_detalle(
    hay_bgr: np.ndarray,
    tpl_path: Path,
    rois: list[tuple[float, float, float, float]] | tuple[tuple[float, float, float, float], ...],
) -> tuple[float, tuple[int, int, int, int] | None]:
    """Igual que _max_similitud_plantilla_haystack pero devuelve rectángulo del mejor match en coords. de hay_bgr."""
    tpl = _cv2_imread_path(tpl_path)
    if tpl is None or tpl.size == 0:
        return 0.0, None
    if tpl.ndim == 3 and tpl.shape[2] == 4:
        tpl = cv2.cvtColor(tpl, cv2.COLOR_BGRA2BGR)
    tpl_g = cv2.cvtColor(tpl, cv2.COLOR_BGR2GRAY)
    best_sc = 0.0
    best_rect: tuple[int, int, int, int] | None = None
    for roi_rel in rois:
        roi, ox, oy = _recorte_roi(hay_bgr, roi_rel)
        roi_g = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        rh, rw = roi_g.shape[:2]
        tpl_loc = _tpl_gray_encajada_en_roi(tpl_g, rw, rh)
        loc_th, loc_tw = tpl_loc.shape[:2]
        for escala in TEMPLATE_SCALES_VENTANA:
            tw = max(8, int(round(loc_tw * escala)))
            th = max(8, int(round(loc_th * escala)))
            if tw >= roi_g.shape[1] or th >= roi_g.shape[0]:
                continue
            interp = cv2.INTER_AREA if escala < 1.0 else cv2.INTER_LINEAR
            t = cv2.resize(tpl_loc, (tw, th), interpolation=interp)
            res = cv2.matchTemplate(roi_g, t, cv2.TM_CCOEFF_NORMED)
            _mn, max_v, _min_loc, max_loc = cv2.minMaxLoc(res)
            max_v = float(max_v)
            if max_v > best_sc:
                lx, ly = int(max_loc[0]), int(max_loc[1])
                best_sc = max_v
                best_rect = (ox + lx, oy + ly, tw, th)
    return best_sc, best_rect


def _max_similitud_plantilla_hwnd(
    hwnd: int,
    tpl_path: Path,
    rois: list[tuple[float, float, float, float]],
) -> float:
    """Mayor TM_CCOEFF_NORMED entre ROIs y escalas (sin clic; para esperas / validación)."""
    hay, _l, _t = _origen_y_captura_para_match_hwnd(hwnd, "")
    return _max_similitud_plantilla_haystack(hay, tpl_path, rois)


def _max_similitud_plantilla_hwnd_excluyendo(
    hwnd: int,
    tpl_path: Path,
    rois: list[tuple[float, float, float, float]],
    exclude_xy: tuple[int, int] | None = None,
    exclude_radio_px: int = 0,
) -> float:
    """Mayor similitud ignorando coincidencias sobre el 1.er botón Imprimir (no anula todo el match)."""
    hay, _l, _t = _origen_y_captura_para_match_hwnd(hwnd, "")
    tpl = _cv2_imread_path(tpl_path)
    if tpl is None or tpl.size == 0:
        return 0.0
    if tpl.ndim == 3 and tpl.shape[2] == 4:
        tpl = cv2.cvtColor(tpl, cv2.COLOR_BGRA2BGR)
    mejor = 0.0
    for roi in rois:
        hit = _mejor_match_multiescala(
            hay,
            tpl,
            roi,
            "left",
            0.06,
            exclude_xy=exclude_xy,
            exclude_radio_px=exclude_radio_px,
        )
        if hit is not None:
            mejor = max(mejor, float(hit[2]))
    return mejor


def _rois_dialogo_imprimir_windows() -> list[tuple[float, float, float, float]]:
    """Solo franja inferior derecha: evita otro texto «Imprimir» (título/etiquetas) en el modal."""
    return [
        (0.50, 0.76, 0.99, 0.99),
        (0.42, 0.70, 0.99, 0.99),
        (0.55, 0.80, 0.98, 0.98),
        (0.35, 0.62, 0.99, 0.99),
    ]


def _rois_boton_imprimir_modal_vie() -> list[tuple[float, float, float, float]]:
    return [
        (0.55, 0.82, 0.99, 0.99),
        (0.48, 0.76, 0.99, 0.99),
    ]


def _match_boton_imprimir_10_contexto(
    hwnd: int,
    titulo: str,
) -> tuple[str, float, float, list[tuple[float, float, float, float]]]:
    """Orden y ROIs para plantilla 10: botón real abajo a la derecha del modal."""
    es_vie = _titulo_ventana_es_modal_vie_consultar_imprimir(titulo)
    es_win = _hwnd_es_dialogo_impresion_sistema(hwnd, titulo) or _norm_etiqueta(titulo) in (
        "imprimir",
        "print",
    )
    if es_vie:
        return (
            "bottom_right",
            _float_env("INDIGO_IMP10_VIE_CX_MIN", 0.52),
            _float_env("INDIGO_IMP10_VIE_CY_MIN", 0.78),
            _rois_boton_imprimir_modal_vie(),
        )
    if es_win:
        return (
            "bottom_right",
            _float_env("INDIGO_IMP10_CX_MIN", 0.48),
            _float_env("INDIGO_IMP10_CY_MIN", 0.70),
            _rois_dialogo_imprimir_windows(),
        )
    return ("left", None, None, [(0.02, 0.02, 0.98, 0.98)])


def _detalle_match_boton_imprimir_10_en_hwnd(
    hwnd: int,
    tpl_path: Path,
    titulo: str,
) -> tuple[float, tuple[int, int] | None]:
    """Mayor similitud del botón (grayscale + TM_CCOEFF_NORMED); opcional Canny si INDIGO_IMP10_MATCH_EDGES=1.
    Devuelve (score, centro del match en coords. de la captura haystack del hwnd)."""
    orden, cx_min, cy_min, rois = _match_boton_imprimir_10_contexto(hwnd, titulo)
    hay, _l, _t = _origen_y_captura_para_match_hwnd(hwnd, "")
    tpl = _cv2_imread_path(tpl_path)
    if tpl is None or tpl.size == 0:
        return 0.0, None
    if tpl.ndim == 3 and tpl.shape[2] == 4:
        tpl = cv2.cvtColor(tpl, cv2.COLOR_BGRA2BGR)

    def _best_pair(h_bgr: np.ndarray, t_bgr: np.ndarray) -> tuple[float, tuple[int, int] | None]:
        mejor_s = 0.0
        mejor_p: tuple[int, int] | None = None
        for roi_rel in rois:
            hit = _mejor_match_multiescala(
                h_bgr,
                t_bgr,
                roi_rel,
                orden,
                0.06,
                cx_min_frac=cx_min,
                cy_min_frac=cy_min,
            )
            if hit is not None and float(hit[2]) > mejor_s:
                mejor_s = float(hit[2])
                mejor_p = (int(hit[0]), int(hit[1]))
        return mejor_s, mejor_p

    sc, pos = _best_pair(hay, tpl)
    if _bool_env("INDIGO_IMP10_MATCH_EDGES", False):
        g_h = cv2.cvtColor(hay, cv2.COLOR_BGR2GRAY)
        g_t = cv2.cvtColor(tpl, cv2.COLOR_BGR2GRAY)
        t1 = int(_float_env("INDIGO_IMP10_EDGES_CANNY_T1", 40))
        t2 = int(_float_env("INDIGO_IMP10_EDGES_CANNY_T2", 120))
        e_h = cv2.Canny(g_h, t1, t2)
        e_t = cv2.Canny(g_t, t1, t2)
        hay_e = cv2.cvtColor(e_h, cv2.COLOR_GRAY2BGR)
        tpl_e = cv2.cvtColor(e_t, cv2.COLOR_GRAY2BGR)
        sc_e, pos_e = _best_pair(hay_e, tpl_e)
        if sc_e > sc:
            sc, pos = sc_e, pos_e
    return sc, pos


def _max_similitud_boton_imprimir_10_en_hwnd(
    hwnd: int,
    tpl_path: Path,
    titulo: str,
) -> float:
    s, _ = _detalle_match_boton_imprimir_10_en_hwnd(hwnd, tpl_path, titulo)
    return s


def _tpl_es_captura_ventana_completa(tpl_path: Path) -> bool:
    tpl = _cv2_imread_path(tpl_path)
    if tpl is None or tpl.size == 0:
        return False
    h, w = tpl.shape[:2]
    return w >= int(_float_env("INDIGO_PLANTILLA_10_ANCHO_VENTANA_MIN", 280)) or h >= int(
        _float_env("INDIGO_PLANTILLA_10_ALTO_VENTANA_MIN", 100)
    )


def _fraccion_click_boton_imprimir_plantilla_10(hwnd: int, tpl_path: Path) -> tuple[float, float] | None:
    """Punto de clic dentro del rectángulo coincidente (0–1). Ventana completa: hacia la esquina del botón.
    Recorte pequeño del botón: suele haber padding asimétrico en el PNG (p. ej. hueco a la izquierda) y el centro
    geométrico cae fuera del control; se desplaza hacia la etiqueta con FRACCION_PEQUENO_* (diálogo Windows «Imprimir»)."""
    raw = (os.environ.get("INDIGO_PLANTILLA_10_FRACCION_CLICK") or "").strip().replace(",", " ").split()
    if len(raw) >= 2:
        try:
            return float(raw[0]), float(raw[1])
        except ValueError:
            pass
    tit = (win32gui.GetWindowText(hwnd) or "").strip()
    if _tpl_es_captura_ventana_completa(tpl_path):
        if _titulo_ventana_es_modal_vie_consultar_imprimir(tit):
            return (
                _float_env("INDIGO_PLANTILLA_10_FRACCION_VIE_X", 0.82),
                _float_env("INDIGO_PLANTILLA_10_FRACCION_VIE_Y", 0.94),
            )
        return (
            _float_env("INDIGO_PLANTILLA_10_FRACCION_WIN_X", 0.83),
            _float_env("INDIGO_PLANTILLA_10_FRACCION_WIN_Y", 0.90),
        )
    # Plantilla = solo el botón (o recorte estrecho): no usar el centro del match como clic por defecto.
    if _hwnd_es_dialogo_impresion_sistema(hwnd, tit) or _norm_etiqueta(tit) in ("imprimir", "print"):
        return (
            _float_env("INDIGO_PLANTILLA_10_FRACCION_PEQUENO_WIN_X", 0.72),
            _float_env("INDIGO_PLANTILLA_10_FRACCION_PEQUENO_WIN_Y", 0.50),
        )
    if _titulo_ventana_es_modal_vie_consultar_imprimir(tit):
        return (
            _float_env("INDIGO_PLANTILLA_10_FRACCION_PEQUENO_VIE_X", 0.58),
            _float_env("INDIGO_PLANTILLA_10_FRACCION_PEQUENO_VIE_Y", 0.52),
        )
    return None


def _hwnd_es_dialogo_impresion_o_guardar_excluir_indigo(hwnd: int, title: str) -> bool:
    """Diálogos Imprimir / Guardar PDF / Imprimiendo: no son la ventana de trabajo IndiGO."""
    tit = title or ""
    if _hwnd_es_dialogo_impresion_sistema(hwnd, tit):
        return True
    if _titulo_es_guardar_pdf_o_impresion(tit):
        return True
    if _titulo_ventana_es_dialogo_impresion_windows(tit) and not _titulo_ventana_es_modal_vie_consultar_imprimir(tit):
        return True
    n = _norm_etiqueta(tit)
    if n in ("imprimir", "print"):
        if _clase_ventana(hwnd).lower() == "#32770":
            return True
    return False


def _preparar_siguiente_cedula_tras_fallo_impresion() -> None:
    """Cierra diálogos de impresión/guardar abiertos para no anclar la siguiente cédula en «Imprimir»."""
    global _hwnd_destino_imprimir_10, _pasos_sin_traer_indigo, _omitir_clic_plantilla_10
    _hwnd_destino_imprimir_10 = None
    _pasos_sin_traer_indigo = 0
    _omitir_clic_plantilla_10 = False
    _desactivar_bloqueo_foco_impresion_9_10()
    h_guardar = _buscar_hwnd_guardar_pdf_abierto(restore=True)
    ced = _cedula_pdf_actual()
    if h_guardar is not None and ced and _bool_env("INDIGO_GUARDAR_PDF_INTENTAR_TRAS_FALLO_IMP", True):
        print(
            f"Flujo: tras fallo impresión — intentando poner cédula {ced!r} "
            "en «Guardar impresión como» antes de cerrar."
        )
        try:
            _escribir_cedula_en_guardar_pdf(ced)
        except Exception as ex:
            print(f"Flujo: aviso — no se pudo escribir cédula tras fallo: {ex}")
    if not _bool_env("INDIGO_CERRAR_DIALOGOS_TRAS_FALLO", True):
        return
    cerrados: list[str] = []
    for h, tit in _enum_hwnds_dialogo_impresion_sistema():
        try:
            activar_ventana_modal(h)
            keyboard.send_keys("{ESC}")
            time.sleep(0.12)
            cerrados.append(tit or "?")
        except Exception:
            pass
    for h, tit in _enum_hwnds_guardar_pdf_impresion():
        n = _norm_etiqueta(tit)
        if "guardar" in n or "save" in n:
            try:
                activar_ventana_modal(h)
                keyboard.send_keys("{ESC}")
                time.sleep(0.12)
                cerrados.append(tit or "?")
            except Exception:
                pass
    if cerrados:
        print(f"Flujo: cerrados diálogos auxiliares antes de la siguiente cédula: {', '.join(cerrados[:4])}")


def _programar_sin_traer_indigo_para_flujo_impresion_pdf() -> None:
    """Evita restaurar_y_traer_al_frente(IndiGO) mientras el usuario debe interactuar con diálogos de Windows."""
    global _pasos_sin_traer_indigo
    extra = int(_float_env("INDIGO_IMP10_PASOS_SIN_TRAER_INDIGO", 40))
    extra = max(12, min(80, extra))
    _pasos_sin_traer_indigo = max(_pasos_sin_traer_indigo, extra)


def _enviar_modal_vie_detras_dialogo_imprimir_windows(hwnd_imprimir: int) -> None:
    """El modal Vie «Consultar / Imprimir Historias» no debe tapar el #32770 «Imprimir»."""
    for h, tit in _enum_hwnds_modal_vie_consultar_imprimir():
        if int(h) == int(hwnd_imprimir):
            continue
        _enviar_ventana_detras(int(h))
        print(f"Flujo: modal Vie detrás del diálogo Imprimir — «{tit}» (hwnd={h})")


def _alejar_ide_consola_del_frente_si_tapa() -> None:
    """Si Cursor/consola quedó al frente durante la espera 9→10, la envía detrás sin activar Vie."""
    if not _bool_env("INDIGO_IMP10_ALEJAR_IDE_DEL_FRENTE", True):
        return
    try:
        fg = int(user32.GetForegroundWindow() or 0)
    except Exception:
        return
    if fg <= 0 or not user32.IsWindow(fg):
        return
    tit_fg = (win32gui.GetWindowText(fg) or "").strip()
    if not _hwnd_es_ide_cursor_o_editor(fg, tit_fg) and not _titulo_es_terminal_o_consola_rpa(tit_fg):
        return
    # No activar otra app: solo quitar IDE/consola del frente para que aparezca Imprimir/Vie.
    _enviar_ventana_detras(fg)
    print(
        f"Flujo: IDE/consola «{tit_fg[:50]}» enviada detrás (no debe tapar Imprimir/Vie durante 9→10)."
    )


def _mantener_foco_flujo_imprimir_9_10(hwnd_vie_modal: int | None = None) -> None:
    """Mantiene foco en «Imprimir» de Windows; el modal Vie solo si aún no abrió el diálogo del sistema."""
    _alejar_ide_consola_del_frente_si_tapa()
    h_win = _buscar_hwnd_dialogo_imprimir_windows_abierto()
    if h_win is not None:
        _enviar_modal_vie_detras_dialogo_imprimir_windows(int(h_win))
        _activar_dialogo_impresion_windows(int(h_win))
        return
    if hwnd_vie_modal and user32.IsWindow(int(hwnd_vie_modal)):
        if not _hwnd_es_ide_cursor_o_editor(int(hwnd_vie_modal), win32gui.GetWindowText(int(hwnd_vie_modal))):
            activar_ventana_modal(int(hwnd_vie_modal))
        return
    for h, _tit in _enum_hwnds_modal_vie_consultar_imprimir():
        activar_ventana_modal(int(h))
        return


def _imp10_requiere_dialogo_imprimir_windows() -> bool:
    return _bool_env("INDIGO_IMP10_SOLO_DIALOGO_WINDOWS", True)


def _esperar_dialogo_imprimir_windows_aparecer(
    timeout_sec: float,
    etiqueta: str = "tras plantilla 9",
) -> int:
    """Espera el #32770 «Imprimir» del sistema (no el botón del modal Vie)."""
    timeout = max(15.0, float(timeout_sec))
    deadline = time.time() + timeout
    ultimo_log = 0.0
    print(
        f"Flujo: esperar diálogo «Imprimir» de Windows {etiqueta} "
        f"(máx {timeout:.0f}s; INDIGO_IMP10_SOLO_DIALOGO_WINDOWS)"
    )
    while time.time() < deadline:
        h = _buscar_hwnd_dialogo_imprimir_windows_abierto()
        if h is not None:
            tit = (win32gui.GetWindowText(int(h)) or "").strip()
            _enviar_modal_vie_detras_dialogo_imprimir_windows(int(h))
            _activar_dialogo_impresion_windows(int(h))
            print(f"Flujo: diálogo Imprimir Windows listo — «{tit}» (hwnd={h})")
            return int(h)
        if time.time() - ultimo_log >= 10.0:
            vie = [t for _h, t in _enum_hwnds_modal_vie_consultar_imprimir()]
            extra = f"; modal Vie visible: {vie[0]!r}" if vie else ""
            print(
                f"Flujo: … aún sin diálogo Imprimir Windows (quedan "
                f"{max(0, int(deadline - time.time()))}s){extra}"
            )
            ultimo_log = time.time()
        time.sleep(0.4)
    raise TimeoutError(
        f"No apareció el diálogo «Imprimir» de Windows {etiqueta} en {timeout:.0f}s. "
        "El RPA no debe hacer clic en el 2.º «Imprimir» del modal Vie (plantilla 9); "
        "espere a que Windows abra su ventana de impresión. Suba INDIGO_ESPERA_MODAL_IMPRIMIR_TIMEOUT "
        "o revise la impresora PDF."
    )


def _set_cedula_pdf_iteracion(cedula: str) -> None:
    global _cedula_pdf_iteracion
    _cedula_pdf_iteracion = _normalizar_cedula_nombre_pdf(cedula) or (cedula or "").strip()


def _cedula_pdf_actual() -> str:
    return _normalizar_cedula_nombre_pdf(_cedula_pdf_iteracion)


def _nombre_es_titulo_spooler_motivo(texto_campo: str, ced: str) -> bool:
    """True si el nombre parece «Verificación Tic» (motivo) y no la cédula del Excel."""
    if _nombre_archivo_coincide_cedula(texto_campo, ced):
        return False
    n = _norm_etiqueta(texto_campo or "")
    if not n:
        return False
    if any(x in n for x in ("verificacion", "verificación")):
        return True
    if re.search(r"\btic\b", n) or n.endswith("tic") or "tic.pdf" in n:
        return True
    return False


def _hwnd_boton_por_texto_en_dialogo(hwnd_root: int, textos_buscar: tuple[str, ...]) -> int | None:
    objetivos = {_norm_etiqueta(t) for t in textos_buscar}
    encontrados: list[int] = []

    @ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
    def cb(ch, _):
        if not user32.IsWindowVisible(ch):
            return True
        t = _norm_etiqueta((win32gui.GetWindowText(ch) or "").strip())
        if t in objetivos:
            encontrados.append(int(ch))
        return True

    try:
        win32gui.EnumChildWindows(int(hwnd_root), cb, None)
    except Exception:
        pass
    return encontrados[0] if encontrados else None


def _pulsar_boton_guardar_en_dialogo(hwnd_dialog: int) -> bool:
    """Clic en «Guardar» del diálogo #32770 (más fiable que Enter con foco en otro control)."""
    _activar_dialogo_guardar_pdf_forzado(int(hwnd_dialog))
    time.sleep(0.08)
    for textos in (
        ("guardar", "&guardar"),
        ("save", "&save"),
    ):
        h_btn = _hwnd_boton_por_texto_en_dialogo(int(hwnd_dialog), textos)
        if h_btn is None:
            continue
        try:
            win32gui.SendMessage(int(h_btn), win32con.BM_CLICK, 0, 0)
            print(f"Flujo: BM_CLICK en botón Guardar (hwnd={h_btn})")
            time.sleep(0.2)
            return True
        except Exception as ex:
            print(f"Flujo: BM_CLICK Guardar falló ({ex})")
    try:
        win32gui.SendMessage(int(hwnd_dialog), win32con.WM_COMMAND, 1, 0)
        print(f"Flujo: WM_COMMAND IDOK en diálogo Guardar (hwnd={hwnd_dialog})")
        return True
    except Exception:
        return False


def _prefijar_nombre_guardar_pdf_si_posible(hwnd_guardar: int | None = None) -> bool:
    """Escribe la cédula en «Guardar impresión como» en cuanto el diálogo está listo."""
    ced = _cedula_pdf_actual()
    if not ced:
        return False
    h = int(hwnd_guardar or _hwnd_guardar_pdf_si_visible() or 0)
    if h <= 0:
        return False
    _activar_dialogo_guardar_pdf_forzado(h)
    time.sleep(0.12)
    if _escribir_nombre_en_todos_edits_guardar(h, ced) or _escribir_nombre_archivo_guardar_pdf_wm(h, ced):
        leido = _leer_nombre_archivo_guardar_pdf(h)
        if _nombre_archivo_coincide_cedula(leido, ced):
            print(f"Flujo: prefijado nombre PDF → {leido!r} (antes de confirmar guardado)")
            return True
    return False


def _programar_sin_traer_indigo_tras_menu_folios() -> None:
    """Tras clic derecho en folios: no traer IndiGO al frente o se cierra el menú contextual (paso 6)."""
    global _pasos_sin_traer_indigo
    extra = int(_float_env("INDIGO_MENU_FOLIO_PASOS_SIN_TRAER_INDIGO", 8))
    extra = max(3, min(20, extra))
    _pasos_sin_traer_indigo = max(_pasos_sin_traer_indigo, extra)


def _activar_dialogo_impresion_windows(hwnd_imp: int) -> None:
    """Refuerzo de foco en el diálogo «Imprimir» antes del clic plantilla 10."""
    if hwnd_imp <= 0 or not user32.IsWindow(hwnd_imp):
        return
    veces = int(_float_env("INDIGO_IMP10_ACTIVAR_DIALOGO_VECES", 3))
    veces = max(1, min(8, veces))
    pausa = _float_env("INDIGO_IMP10_ACTIVAR_DIALOGO_PAUSA", 0.25)
    for _ in range(veces):
        if not _force_foreground_window(int(hwnd_imp)):
            activar_ventana_modal(hwnd_imp)
        time.sleep(pausa)


def _buscar_hwnd_dialogo_imprimir_windows_abierto() -> int | None:
    for hi, tit in _enum_hwnds_dialogo_impresion_sistema():
        if _norm_etiqueta(tit) in ("imprimir", "print"):
            h = int(hi)
            _restaurar_hwnd_si_minimizado(h, "Imprimir Windows")
            return h
    return None


def _titulo_es_dialogo_imprimiendo_progreso(title: str) -> bool:
    """Ventana de progreso «Imprimiendo» — no tiene el campo Nombre del PDF."""
    n = _norm_etiqueta(title or "")
    if not n:
        return False
    if n in ("imprimiendo", "printing"):
        return True
    if ("imprimiendo" in n or "printing" in n) and "guardar" not in n:
        return True
    return False


def _titulo_es_dialogo_guardar_pdf_nombre(title: str) -> bool:
    """«Guardar impresión como» / Guardar como — aquí va la cédula en el campo Nombre."""
    if _titulo_es_dialogo_imprimiendo_progreso(title):
        return False
    n = _norm_etiqueta(title or "")
    compact = n.replace(" ", "")
    if "guardar" in n and (
        "como" in n or "pdf" in n or "salida" in n or "impresion" in n or "resultado" in n
    ):
        return True
    if "guardar" in n and "impresion" in n:
        return True
    if "guardar" in n and "imprimir" in n:
        return True
    if "save print" in n or "print output" in n or "guardarimpresion" in compact:
        return True
    if "saveprintoutput" in compact or "guardarcomo" in compact:
        return True
    if "save as" in n or "saveas" in compact:
        return True
    if "microsoft" in n and "pdf" in n:
        return True
    if "print" in n and "pdf" in n and ("file" in n or "archivo" in n or "output" in n or "salida" in n):
        return True
    return _titulo_es_dialogo_guardar_archivo(title)


def _titulo_es_guardar_pdf_o_impresion(title: str) -> bool:
    """Cualquier diálogo del flujo PDF (progreso o guardar nombre)."""
    return _titulo_es_dialogo_imprimiendo_progreso(title) or _titulo_es_dialogo_guardar_pdf_nombre(title)


def _prioridad_titulo_dialogo_guardar_pdf(titulo: str) -> int:
    n = _norm_etiqueta(titulo or "")
    if "guardar" in n and "impresion" in n and "como" in n:
        return 100
    if "guardar" in n and ("como" in n or "impresion" in n):
        return 90
    if "guardar" in n:
        return 70
    if "save" in n and "as" in n:
        return 85
    return 10


def _enum_hwnds_top_level_por_filtro(
    acepta_titulo,
) -> list[tuple[int, str]]:
    out: list[tuple[int, str]] = []

    @ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
    def cb(hwnd, _):
        try:
            hi = int(hwnd)
        except (TypeError, ValueError):
            return True
        if not _hwnd_top_level_utilizable(hi):
            return True
        tit = (win32gui.GetWindowText(hwnd) or "").strip()
        cls = _clase_ventana(hi).lower()
        if acepta_titulo(tit):
            out.append((hi, tit))
            return True
        if cls == "#32770" and _titulo_es_dialogo_guardar_archivo(tit) and acepta_titulo(tit):
            out.append((hi, tit))
        return True

    user32.EnumWindows(cb, 0)
    return out


def _enum_hwnds_dialogo_guardar_pdf() -> list[tuple[int, str]]:
    """Solo «Guardar impresión como» (campo Nombre), no «Imprimiendo»."""
    return _enum_hwnds_top_level_por_filtro(_titulo_es_dialogo_guardar_pdf_nombre)


def _enum_hwnds_dialogo_imprimiendo() -> list[tuple[int, str]]:
    return _enum_hwnds_top_level_por_filtro(_titulo_es_dialogo_imprimiendo_progreso)


def _enum_hwnds_guardar_pdf_impresion() -> list[tuple[int, str]]:
    """Alias: diálogo donde se escribe el nombre del PDF (sin ventana de progreso)."""
    return _enum_hwnds_dialogo_guardar_pdf()


def _seleccionar_hwnd_dialogo_guardar_pdf(cands: list[tuple[int, str]] | None = None) -> int | None:
    """Elige «Guardar impresión como», nunca «Imprimiendo»."""
    if cands is None:
        cands = _enum_hwnds_dialogo_guardar_pdf()
    if not cands:
        return None
    ordenados = sorted(cands, key=lambda x: _prioridad_titulo_dialogo_guardar_pdf(x[1]), reverse=True)
    return int(ordenados[0][0])


def _enviar_imprimiendo_detras_del_guardar(hwnd_guardar: int) -> None:
    """La ventana «Imprimiendo» no debe tapar «Guardar impresión como»."""
    if not _bool_env("INDIGO_GUARDAR_PDF_IMPRIMIENDO_DETRAS", True):
        return
    for h, tit in _enum_hwnds_dialogo_imprimiendo():
        if int(h) == int(hwnd_guardar):
            continue
        _enviar_ventana_detras(int(h))
        print(f"Flujo: «Imprimiendo» detrás del Guardar PDF — «{tit}» (hwnd={h})")


def _titulo_es_dialogo_guardar_archivo(title: str) -> bool:
    n = _norm_etiqueta(title or "")
    if not n:
        return False
    if "guardar" in n and ("como" in n or "impresion" in n or "pdf" in n):
        return True
    if "save" in n and ("as" in n or "print" in n):
        return True
    return False


def _buscar_hwnd_guardar_pdf_abierto(restore: bool = True) -> int | None:
    hwnd = _seleccionar_hwnd_dialogo_guardar_pdf()
    if hwnd is None:
        return None
    if restore:
        _restaurar_hwnd_si_minimizado(hwnd, "Guardar PDF")
        _activar_dialogo_guardar_pdf_forzado(hwnd)
    return hwnd


def _imp10_saltar_si_guardar_pdf_abierto() -> bool:
    """Print to PDF saltó el #32770 «Imprimir»: ya está «Guardar impresión como»."""
    global _omitir_clic_plantilla_10, _hwnd_destino_imprimir_10
    if not _bool_env("INDIGO_IMP10_SALTAR_SI_GUARDAR_PDF", True):
        return False
    h = _buscar_hwnd_guardar_pdf_abierto(restore=True)
    if h is None:
        return False
    _omitir_clic_plantilla_10 = True
    _hwnd_destino_imprimir_10 = None
    _programar_sin_traer_indigo_para_flujo_impresion_pdf()
    tit = (win32gui.GetWindowText(int(h)) or "").strip()
    print(
        "Flujo: «Guardar impresión como» detectado — la impresión ya avanzó "
        f"(no hace falta 2.º Imprimir en pantalla). «{tit}» hwnd={h}"
    )
    if _cedula_pdf_actual():
        try:
            _preparar_campo_nombre_guardar_pdf_para_escritura(0)
            _prefijar_nombre_guardar_pdf_si_posible(int(h))
        except Exception as ex:
            print(f"Flujo: aviso — prefijar cédula al detectar Guardar PDF: {ex}")
    return True


def _traer_dialogos_guardar_pdf_al_frente() -> int | None:
    cands_guardar = _enum_hwnds_dialogo_guardar_pdf()
    if not cands_guardar:
        return None
    hwnd_save = _seleccionar_hwnd_dialogo_guardar_pdf(cands_guardar)
    if hwnd_save is None:
        return None
    imprimiendo = _enum_hwnds_dialogo_imprimiendo()
    if imprimiendo:
        nombres_imp = ", ".join(repr(t) for _h, t in imprimiendo[:2])
        print(
            f"Flujo: diálogo Guardar PDF al frente — «{win32gui.GetWindowText(hwnd_save)!r}» "
            f"(hwnd={hwnd_save}); progreso aparte: {nombres_imp}"
        )
    _activar_dialogo_guardar_pdf_forzado(int(hwnd_save))
    return int(hwnd_save)


def _clic_10_produjo_avance_impresion() -> bool:
    """True si tras el clic 10 apareció Imprimiendo, Guardar PDF o cerró el diálogo Imprimir."""
    if _enum_hwnds_dialogo_guardar_pdf() or _enum_hwnds_dialogo_imprimiendo():
        return True
    imprimir_abiertos = [
        (h, t)
        for h, t in _enum_hwnds_dialogo_impresion_sistema()
        if _norm_etiqueta(t) in ("imprimir", "print")
    ]
    if not imprimir_abiertos:
        return True
    return False


def _similitud_imprimir_10_y_hwnd_destino(
    hwnd_vie: int,
    tpl_path: Path,
    rois_vie: list[tuple[float, float, float, float]],
    exclude_xy: tuple[int, int] | None,
    exclude_radio_px: int,
) -> tuple[float, int | None, str]:
    """Busca plantilla 10 en diálogo Imprimir de Windows y en la ventana Vie (excl. botón 9)."""
    mejor = 0.0
    mejor_hwnd: int | None = None
    origen = "ninguno"
    for h, tit in _enum_hwnds_dialogo_impresion_windows():
        if _titulo_ventana_es_modal_vie_consultar_imprimir(tit):
            sc = _max_similitud_boton_imprimir_10_en_hwnd(h, tpl_path, tit)
            if sc > mejor:
                mejor, mejor_hwnd, origen = sc, h, f"modal Vie ({tit!r})"
            continue
        if not _hwnd_es_dialogo_impresion_sistema(h, tit) and _norm_etiqueta(tit) not in (
            "imprimir",
            "print",
        ):
            continue
        sc = _max_similitud_boton_imprimir_10_en_hwnd(h, tpl_path, tit)
        if sc > mejor:
            mejor, mejor_hwnd, origen = sc, h, f"diálogo Windows ({tit!r})"
    sc_vie = _max_similitud_plantilla_hwnd_excluyendo(
        hwnd_vie, tpl_path, rois_vie, exclude_xy, exclude_radio_px
    )
    if sc_vie > mejor:
        mejor, mejor_hwnd, origen = sc_vie, hwnd_vie, "ventana Vie/IndiGO"
    return mejor, mejor_hwnd, origen


def _max_similitud_imprimir_segundo_multiventana(
    hwnd_vie: int,
    tpl_path: Path,
    rois_vie: list[tuple[float, float, float, float]],
    exclude_xy: tuple[int, int] | None = None,
    exclude_radio_px: int = 0,
) -> tuple[float, int | None, str]:
    return _similitud_imprimir_10_y_hwnd_destino(
        hwnd_vie, tpl_path, rois_vie, exclude_xy, exclude_radio_px
    )


def _float_env(name: str, default: float) -> float:
    raw = (os.environ.get(name) or "").strip().replace(",", ".")
    if not raw:
        return default
    try:
        return float(raw)
    except ValueError:
        return default


def _bool_env(name: str, default: bool) -> bool:
    raw = (os.environ.get(name) or "").strip().lower()
    if not raw:
        return default
    if raw in ("1", "true", "yes", "on", "si", "sí"):
        return True
    if raw in ("0", "false", "no", "off"):
        return False
    return default


def _titulo_es_terminal_o_consola_rpa(title: str) -> bool:
    """Excluye consolas, IDEs y editores: no deben usarse para matchTemplate de Vie Cloud."""
    n = _norm_etiqueta(title)
    # Título vacío: no excluir (WPF / Electron a veces no rellenan la barra de título).
    if not n:
        return False
    if "cmd.exe" in n or "powershell" in n or "pwsh" in n:
        return True
    if "windowsterminal" in n.replace(" ", "") or "windows terminal" in n:
        return True
    if "system32" in n and "cmd" in n:
        return True
    if "run_indigo" in n.replace(" ", "") or "indigo_historias" in n.replace(" ", ""):
        return True
    if "rpa indigo" in n and (".py" in n or "py " in n):
        return True
    if "python" in n and ".py" in n:
        return True
    # Cursor: título «… - Cursor»; sin .py (no confundir con Vie Cloud).
    if " - cursor" in n or n.rstrip().endswith("- cursor"):
        return True
    if "cursor" in n and (".py" in n or "py " in n):
        return True
    if "vscode" in n or "visual studio code" in n or "visual studio" in n:
        return True
    if "jetbrains" in n or "pycharm" in n or "webstorm" in n:
        return True
    # Bloc de notas / Notepad: «INDIGO.txt» dispara la pista «indigo» por error.
    if "bloc de notas" in n or "notepad" in n or "wordpad" in n:
        return True
    # SQL Server Management Studio / clientes SQL: dan falsos positivos altos con plantillas 0 y 8.
    if "microsoft sql server management studio" in n:
        return True
    if "sql server management studio" in n:
        return True
    if "management studio" in n and "sql" in n:
        return True
    if ".sql" in n and ("no conectado" in n or "disconnected" in n or "query" in n):
        return True
    if "azure data studio" in n:
        return True
    compact = n.replace(" ", "")
    if "azuredatastudio" in compact or "dbeaver" in n or "pgadmin" in n:
        return True
    return False


def _hwnd_es_ide_cursor_o_editor(hwnd: int, title: str) -> bool:
    """Cursor, VS Code y editores Chromium: no deben anclarse como Vie ni restaurarse al frente."""
    if _titulo_es_terminal_o_consola_rpa(title):
        return True
    n = _norm_etiqueta(title or "")
    cls = _clase_ventana(hwnd).lower()
    if " - cursor" in n or n.endswith("- cursor") or n == "cursor":
        return True
    if "cursor" in n and (".py" in n or " - " in n):
        return True
    if "visual studio code" in n or n.endswith("- visual studio code"):
        return True
    if cls == "chrome_widgetwin_1" and (
        "cursor" in n
        or "visual studio code" in n
        or "vscode" in n
        or (not n and _bool_env("INDIGO_EXCLUIR_CHROME_SIN_TITULO_IDE", False))
    ):
        return True
    return False


def _clase_ventana(hwnd: int) -> str:
    try:
        return (win32gui.GetClassName(int(hwnd)) or "").strip()
    except Exception:
        return ""


def _hwnd_es_explorador_archivos_windows(hwnd: int) -> bool:
    """Explorador de carpetas (CabinetWClass): puede coincidir con plantillas y al traerlo al frente tapa IndiGO."""
    cls = _clase_ventana(hwnd).lower()
    return cls in ("cabinetwclass", "explorewclass")


def _hwnd_es_cliente_navegador_o_correo(hwnd: int, title: str) -> bool:
    """Chrome, Edge, Firefox y clientes de correo: suelen dar match visual con plantillas clínicas por error."""
    n = _norm_etiqueta(title or "")
    cls = _clase_ventana(hwnd).lower()
    if cls == "mozillawindowclass":
        return True
    if not n:
        return False
    compact = n.replace(" ", "")
    if "microsoftteams" in compact or ("teams" in n and "|" in (title or "")):
        return True
    marcas = (
        "googlechrome",
        "chromium",
        "mozillafirefox",
        "microsoftedge",
        "operasoftware",
        "bravesoftware",
        "microsoftoutlook",
        "outlook",
        "thunderbird",
        "windowsmail",
        "gmail",
        "hotmail",
        "yahoomail",
        "correo -",
        " - mail",
        " - correo",
    )
    if any(m in compact or m in n for m in marcas):
        return True
    if "correo" in n and ("windows" in n or "microsoft" in n):
        return True
    if cls == "chrome_widgetwin_1" and (
        "googlechrome" in compact or "microsoftedge" in compact or "chromium" in compact
    ):
        return True
    return False


def _hwnd_es_excel_u_hoja_calculo(hwnd: int, title: str) -> bool:
    """Excel (y libros .xlsx abiertos): no son Vie Cloud; «Pacientes_….xlsx» activaba la pista «paciente» por subcadena."""
    cls = _clase_ventana(hwnd).lower()
    if cls == "xlmain":
        return True
    n = _norm_etiqueta(title or "")
    compact = n.replace(" ", "")
    for suf in (".xlsx", ".xlsm", ".xlsb", ".xls"):
        if suf in compact:
            return True
    if "protectedview" in compact and "excel" in compact:
        return True
    if "microsoftexcel" in compact:
        return True
    if n.rstrip().endswith("excel") or " - excel" in n:
        return True
    if "libreoffice" in n and "calc" in n:
        return True
    return False


def _titulo_ventana_es_modal_vie_consultar_imprimir(title: str) -> bool:
    """Modal de Vie «Consultar / Imprimir Historias» — no es el diálogo Imprimir del sistema."""
    n = _norm_etiqueta(title or "")
    if not n:
        return False
    if "consultar" in n and "imprimir" in n:
        return True
    if "historias" in n and "imprimir" in n:
        return True
    if "historia" in n and "clinica" in n and "imprimir" in n:
        return True
    return False


def _enum_hwnds_modal_vie_consultar_imprimir() -> list[tuple[int, str]]:
    """Ventanas top-level del modal Vie (plantillas 9 y 10 del 1.er Imprimir)."""
    out: list[tuple[int, str]] = []

    @ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
    def cb(hwnd, _):
        if not user32.IsWindowVisible(hwnd):
            return True
        try:
            hi = int(hwnd)
        except (TypeError, ValueError):
            return True
        tit = (win32gui.GetWindowText(hwnd) or "").strip()
        if _titulo_ventana_es_modal_vie_consultar_imprimir(tit):
            out.append((hi, tit))
        return True

    user32.EnumWindows(cb, 0)
    return out


def _esperar_hwnd_modal_vie_consultar_imprimir(timeout_sec: float) -> tuple[int, str] | None:
    deadline = time.time() + max(1.0, timeout_sec)
    ultimo_info = 0.0
    while time.time() < deadline:
        cands = _enum_hwnds_modal_vie_consultar_imprimir()
        if cands:
            return cands[0]
        if time.time() - ultimo_info >= 3.0:
            print(
                f"Flujo: … esperando modal «Consultar / Imprimir Historias» "
                f"(quedan {max(0, int(deadline - time.time()))}s)"
            )
            ultimo_info = time.time()
        time.sleep(0.28)
    return None


def _titulo_ventana_es_dialogo_impresion_windows(title: str) -> bool:
    """Diálogo Imprimir / Guardar PDF del sistema (no el modal grande de Vie)."""
    n = _norm_etiqueta(title or "")
    if not n:
        return False
    if _titulo_ventana_es_modal_vie_consultar_imprimir(title):
        return False
    compact = n.replace(" ", "")
    if n in ("imprimir", "print") or compact in ("imprimir", "print"):
        return True
    if "guardar como" in n or "guardarcomo" in compact:
        return True
    if "save as" in n or "saveas" in compact:
        return True
    if "saveprintoutput" in compact or "microsoftprinttopdf" in compact or "printtopdf" in compact:
        return True
    if "propiedades de impresora" in n or "printerproperties" in compact:
        return True
    if "impresora" in n and len(n) < 55:
        return True
    if n.startswith("imprimir ") and len(n) < 48 and "historia" not in n and "consultar" not in n:
        return True
    return False


def _hwnd_es_dialogo_impresion_sistema(hwnd: int, title: str) -> bool:
    """#32770 + título corto «Imprimir»; excluye ventana Vie «Consultar / Imprimir Historias»."""
    if not _titulo_ventana_es_dialogo_impresion_windows(title):
        return False
    if _titulo_ventana_es_modal_vie_consultar_imprimir(title):
        return False
    n = _norm_etiqueta(title or "")
    cls = _clase_ventana(hwnd).lower()
    if cls == "#32770":
        return True
    if n in ("imprimir", "print"):
        return True
    if "guardar como" in n or "save as" in n:
        return True
    return False


def _enum_hwnds_dialogo_impresion_windows() -> list[tuple[int, str]]:
    out: list[tuple[int, str]] = []

    @ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
    def cb(hwnd, _):
        try:
            hi = int(hwnd)
        except (TypeError, ValueError):
            return True
        if not _hwnd_top_level_utilizable(hi):
            return True
        tit = (win32gui.GetWindowText(hwnd) or "").strip()
        if _titulo_ventana_es_dialogo_impresion_windows(tit):
            out.append((hi, tit))
        return True

    user32.EnumWindows(cb, 0)
    return out


def _enum_hwnds_dialogo_impresion_sistema() -> list[tuple[int, str]]:
    return [(h, t) for h, t in _enum_hwnds_dialogo_impresion_windows() if _hwnd_es_dialogo_impresion_sistema(h, t)]


def _dialogo_impresion_sistema_visible() -> bool:
    for _h, tit in _enum_hwnds_dialogo_impresion_sistema():
        n = _norm_etiqueta(tit)
        if n in ("imprimir", "print"):
            return True
    return len(_enum_hwnds_dialogo_impresion_sistema()) > 0


def _esperar_dialogo_impresion_windows(timeout: float | None = None) -> None:
    """Tras click 10: espera aparición y cierre del diálogo Imprimir de Windows antes de Deshacer/guardar."""
    if timeout is None:
        timeout = _float_env("INDIGO_IMPRESION_TIMEOUT", 180.0)
    timeout = max(15.0, min(900.0, float(timeout)))
    t0 = time.time()
    aparecer_max = _float_env("INDIGO_IMPRESION_APARECER_TIMEOUT", 55.0)
    hwnd_imp: int | None = None
    tit_imp = ""

    print(
        f"Flujo: esperar ventana de impresión de Windows (máx {timeout:.0f}s; "
        f"INDIGO_IMPRESION_TIMEOUT / INDIGO_IMPRESION_APARECER_TIMEOUT)"
    )

    while time.time() - t0 < aparecer_max:
        cands = _enum_hwnds_dialogo_impresion_sistema()
        for hi, tit in cands:
            if _norm_etiqueta(tit) in ("imprimir", "print"):
                hwnd_imp, tit_imp = hi, tit
                print(f"Flujo: diálogo Imprimir de Windows → {tit_imp!r} (hwnd={hwnd_imp})")
                break
        if hwnd_imp is not None:
            break
        guard = _enum_hwnds_dialogo_guardar_pdf()
        if guard:
            tit_g = guard[0][1]
            print(
                f"Flujo: Impresión a PDF — diálogo Guardar detectado ({tit_g!r}); "
                "se estabiliza antes de escribir la cédula."
            )
            h_g = _esperar_dialogo_guardar_pdf_estable(
                timeout_sec=_float_env("INDIGO_GUARDAR_PDF_APARECER_TIMEOUT", 90.0),
                pausa_estabilizar=_float_env(
                    "INDIGO_IMPRESION_TRAS_GUARDAR_VISIBLE_SEG",
                    _float_env("INDIGO_GUARDAR_PDF_ESTABILIZAR_SEG", 2.5),
                ),
            )
            _programar_sin_traer_indigo_para_flujo_impresion_pdf()
            _prefijar_nombre_guardar_pdf_si_posible(int(h_g))
            return
        vie = [t for _h, t in _enum_hwnds_dialogo_impresion_windows() if _titulo_ventana_es_modal_vie_consultar_imprimir(t)]
        if vie and int(time.time() - t0) % 8 < 1:
            print(
                f"Flujo: modal Vie aún visible ({vie[0]!r}); "
                "esperando diálogo «Imprimir» del sistema (no confundir con Consultar/Imprimir Historias)."
            )
        time.sleep(0.35)

    if hwnd_imp is None:
        guard2 = _enum_hwnds_dialogo_guardar_pdf()
        if guard2:
            print(
                f"Flujo: diálogo Guardar detectado al final de la espera inicial "
                f"({guard2[0][1]!r}); estabilizando."
            )
            h_g2 = _esperar_dialogo_guardar_pdf_estable()
            _programar_sin_traer_indigo_para_flujo_impresion_pdf()
            _prefijar_nombre_guardar_pdf_si_posible(int(h_g2))
            return
        fijo = _float_env("INDIGO_IMPRESION_SLEEP_SI_SIN_VENTANA", 18.0)
        print(
            f"Flujo: no se detectó «Imprimir» en la barra de título; "
            f"espera fija {fijo:.1f}s (INDIGO_IMPRESION_SLEEP_SI_SIN_VENTANA)."
        )
        time.sleep(fijo)
        return

    ultimo_log = time.time()
    while time.time() - t0 < timeout:
        cands = _enum_hwnds_dialogo_impresion_sistema()
        if not cands:
            extra = _float_env("INDIGO_IMPRESION_TRAS_CIERRE_SEG", 1.5)
            print(
                f"Flujo: sin diálogos de impresión/guardar PDF (esperó {time.time() - t0:.1f}s); "
                f"pausa {extra:.1f}s."
            )
            time.sleep(extra)
            return
        if time.time() - ultimo_log >= 15.0:
            titulos = ", ".join(repr(t) for _h, t in cands[:3])
            print(
                f"Flujo: … aún en impresión/guardar ({titulos}; "
                f"quedan {max(0, int(timeout - (time.time() - t0)))}s)"
            )
            ultimo_log = time.time()
        time.sleep(0.45)

    print(f"Flujo: aviso — timeout ({timeout:.0f}s) esperando cierre del diálogo de impresión; se continúa.")


def _enum_candidatos_hwnd_para_plantilla_indigo() -> list[tuple[int, str]]:
    """Modo legado (INDIGO_VENTANA_BUSQUEDA_VISUAL_TODAS=0): títulos no vacíos + Chromium sin título según env."""
    out: list[tuple[int, str]] = []
    seen: set[int] = set()

    @ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
    def cb(hwnd, _):
        if not user32.IsWindowVisible(hwnd):
            return True
        try:
            hi = int(hwnd)
        except (TypeError, ValueError):
            return True
        if hi in seen:
            return True
        buf = ctypes.create_unicode_buffer(512)
        user32.GetWindowTextW(hwnd, buf, 512)
        tit = (buf.value or "").strip()
        if tit:
            seen.add(hi)
            out.append((hi, buf.value))
            return True
        # Chromium/WebView2 sin título en la barra (Vie frecuente). Por defecto sí se incluye; opt-out con =0.
        raw_ch = (os.environ.get("INDIGO_INCLUIR_CHROMIUM_TOPLEVEL_SIN_TITULO") or "").strip().lower()
        if raw_ch in ("0", "false", "no", "off"):
            incluir = False
        elif raw_ch in ("1", "true", "yes", "on", "si", "sí"):
            incluir = True
        else:
            incluir = True
        cls = _clase_ventana(hi).lower()
        if not incluir or cls != "chrome_widgetwin_1":
            return True
        left, top, right, bottom = rect_pantalla(hi)
        w, hgt = right - left, bottom - top
        if w < 680 or hgt < 420:
            return True
        seen.add(hi)
        out.append((hi, "(sin titulo · WebView/Chromium)"))
        return True

    user32.EnumWindows(cb, 0)
    return out


def _hwnd_es_shell_escritorio_o_falso_match_plantilla(hwnd: int, title: str) -> bool:
    """Progman / Program Manager / WorkerW capturan el escritorio: suelen dar 0.9+ de similitud con plantilla 0 por error."""
    n = _norm_etiqueta(title or "")
    if n == "program manager":
        return True
    cls = _clase_ventana(hwnd).lower()
    if cls == "progman":
        return True
    if cls == "workerw" and len(n) < 2:
        return True
    if cls == "shell_traywnd":
        return True
    return False


def _hwnd_es_ventana_sistema_excluir_de_match_indigo(_hwnd: int, title: str) -> bool:
    """Paneles de Windows que pueden dar matchTemplate 1.0 contra UI clínica por error (no son Vie Cloud)."""
    n = _norm_etiqueta(title or "")
    if not n:
        return False
    compact = n.replace(" ", "")
    # «Experiencia de entrada de Windows» / Windows Input Experience (Ink, dictado, etc.)
    if "experienciadeentradadewindows" in compact or "windowsinputexperience" in compact:
        return True
    if "entrada de texto" in n or "textinput" in compact:
        return True
    if "tecladotactil" in compact or "touchkeyboard" in compact or "tecladoenpantalla" in compact:
        return True
    if "onscreenkeyboard" in compact or "osk.exe" in compact:
        return True
    return False


def _prioridad_titulo_ventana_ve_indigo(title: str) -> int:
    """Mayor = más probable que sea Vie / IndiGO (desempate si varias ventanas tienen similitud alta)."""
    n = _norm_etiqueta(title or "")
    if not n:
        return 0
    if "consulta historias" in n or "consulta de historias" in n:
        return 4
    if "vie ehr" in n or "vie cloud" in n or "viecloud" in n.replace(" ", ""):
        return 3
    if "vie" in n and "ehr" in n:
        return 3
    if "indigo" in n and len(n) >= 12:
        return 2
    if "historia" in n and "clinica" in n:
        return 1
    return 0


def _titulo_bloquea_busqueda_plantilla_ventana_indigo(title: str) -> bool:
    """Igual que Stone/KACTUS: no comparar plantilla contra Teams, Slack, etc. (evitan falsos positivos)."""
    n = _norm_etiqueta(title or "")
    if not n:
        return False
    compact = n.replace(" ", "")
    if "microsoftteams" in compact or "msteams" in compact:
        return True
    if "teams" in n and ("microsoft" in n or "|" in (title or "")):
        return True
    bloqueos = (
        "skype",
        "slack -",
        "slack |",
        "zoom meeting",
        "zoom -",
        "discord",
        "telegram",
        "spotify",
        "whatsapp",
        "google meet",
        "webex",
    )
    if any(b in n for b in bloqueos):
        return True
    if n.startswith("teams ") or n == "teams":
        return True
    return False


def _puntuacion_titulo_favorece_vie_indigo(title: str) -> int:
    """Puntuación tipo Stone: solo ventanas con pistas fuertes de Vie/IndiGO pasan el umbral por defecto."""
    n = _norm_etiqueta(title or "")
    if not n:
        return 0
    compact = n.replace(" ", "")
    score = 0
    if "consultahistorias" in compact or "consultadehistorias" in compact or "consulta historias" in n:
        score += 14
    if "vieehr" in compact or "vie ehr" in n:
        score += 10
    if "viecloud" in compact or "vie cloud" in n or "viecloudplatform" in compact:
        score += 10
    if "vie" in n and ("clinical" in n or "ehr" in n or "cloud" in n):
        score += 6
    if "indigointernational" in compact or "indigo international" in n:
        score += 8
    if "indigo" in n and len(n) >= 14:
        score += 4
    if "historia" in n and "clinica" in n:
        score += 5
    if "cloudplatform" in compact and "vie" in n:
        score += 5
    return score


def _dimensiones_minimas_ventana_para_plantilla() -> tuple[int, int]:
    """Evita anclar la plantilla a ventanas diminutas del sistema (INDIGO_VENTANA_MIN_ANCHO / _ALTO)."""
    try:
        mw = int((os.environ.get("INDIGO_VENTANA_MIN_ANCHO") or "480").strip())
    except ValueError:
        mw = 480
    try:
        mh = int((os.environ.get("INDIGO_VENTANA_MIN_ALTO") or "300").strip())
    except ValueError:
        mh = 300
    return max(380, mw), max(260, mh)


def _plantilla_es_pantalla_inicial_busqueda_ventana(p: Path) -> bool:
    """PNG de la pantalla de arranque (Consulta historias): ancla la búsqueda de HWND y evita Chrome/Excel."""
    low = p.name.lower()
    return low.startswith("0 ") or low.startswith("0.") or "pantalla inicial" in low


def _plantilla_es_marca_vie_ehr(p: Path) -> bool:
    """Plantilla pequeña del logo / texto «Vie EHR» (esquina superior derecha). Nombre de archivo con vie + ehr."""
    low = p.name.lower()
    if "vie" in low and "ehr" in low:
        return True
    compact = low.replace(" ", "").replace("_", "").replace("-", "")
    return "vieehr" in compact


def _rois_busqueda_para_plantilla_ventana(p: Path) -> tuple[tuple[float, float, float, float], ...]:
    """ROI: logo Vie EHR (arriba-derecha) primero si el PNG lo sugiere; si no, ROI general del formulario."""
    vrois = tuple(ROIS_BUSQUEDA_VENTANA_VIE_EHR)
    prois = tuple(ROIS_BUSQUEDA_VENTANA_POR_PLANTILLA)
    if _plantilla_es_marca_vie_ehr(p):
        return vrois + prois
    if _bool_env("INDIGO_VENTANA_ROI_ESQUINA_TODAS_PLANTILLAS", False):
        return vrois + prois
    return prois


def _priorizar_rutas_plantilla_vie_ehr(rutas: list[Path]) -> list[Path]:
    vie = [p for p in rutas if _plantilla_es_marca_vie_ehr(p)]
    rest = [p for p in rutas if not _plantilla_es_marca_vie_ehr(p)]
    return vie + rest


def _sanitizar_fragmento_nombre_archivo(s: str, max_len: int = 48) -> str:
    t = re.sub(r"[^\w.\-]+", "_", (s or "").strip(), flags=re.UNICODE)
    t = t.strip("._") or "x"
    return t[:max_len]


def _enum_toplevel_para_match_visual_indigo() -> list[tuple[int, str]]:
    """Todas las ventanas top-level visibles, no minimizadas, tamaño ≥ mínimo; título opcional (WebView2, Java, etc.)."""
    out: list[tuple[int, str]] = []
    mw, mh = _dimensiones_minimas_ventana_para_plantilla()
    try:
        max_c = int((os.environ.get("INDIGO_VENTANA_VISUAL_MAX_CANDIDATOS") or "100").strip())
    except ValueError:
        max_c = 100
    max_c = max(8, min(300, max_c))

    @ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
    def cb(hwnd, _):
        if len(out) >= max_c:
            return False
        if not user32.IsWindowVisible(hwnd):
            return True
        if user32.IsIconic(hwnd):
            return True
        try:
            hi = int(hwnd)
        except (TypeError, ValueError):
            return True
        left, top, right, bottom = rect_pantalla(hi)
        w, hgt = right - left, bottom - top
        if w < mw or hgt < mh:
            return True
        if not _rect_plausible(left, top, right, bottom):
            return True
        try:
            ex = int(win32gui.GetWindowLong(hi, win32con.GWL_EXSTYLE))
        except Exception:
            ex = 0
        if (ex & win32con.WS_EX_TOOLWINDOW) and w * hgt < 450_000:
            return True
        buf = ctypes.create_unicode_buffer(512)
        user32.GetWindowTextW(hwnd, buf, 512)
        title = (buf.value or "").strip()
        if _hwnd_es_shell_escritorio_o_falso_match_plantilla(hi, title):
            return True
        if _hwnd_es_ventana_sistema_excluir_de_match_indigo(hi, title):
            return True
        if _hwnd_es_dialogo_impresion_o_guardar_excluir_indigo(hi, title):
            return True
        if _hwnd_es_explorador_archivos_windows(hi):
            return True
        if _hwnd_es_excel_u_hoja_calculo(hi, title):
            return True
        if _titulo_es_terminal_o_consola_rpa(title):
            return True
        if _hwnd_es_cliente_navegador_o_correo(hi, title):
            return True
        if _titulo_bloquea_busqueda_plantilla_ventana_indigo(title):
            return True
        out.append((hi, title if title else "(sin título)"))
        return True

    user32.EnumWindows(cb, 0)
    fg = int(user32.GetForegroundWindow())
    if fg > 0 and user32.IsWindowVisible(fg):
        rest = [(h, t) for h, t in out if h != fg]
        hit = next(((h, t) for h, t in out if h == fg), None)
        if hit:
            return [hit] + rest
    return out


def _dir_debug_escaneo_ventana_indigo() -> Path | None:
    if not _bool_env("INDIGO_VENTANA_DEBUG_SCAN", False):
        return None
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    d = SALIDAS_DIR / "debug_ventana_indigo" / ts
    d.mkdir(parents=True, exist_ok=True)
    return d


def _resize_para_debug_captura(img: np.ndarray, max_ancho: int = 1280) -> np.ndarray:
    h, w = img.shape[:2]
    if w <= max_ancho:
        return img
    s = max_ancho / float(w)
    return cv2.resize(img, (max_ancho, max(1, int(round(h * s)))), interpolation=cv2.INTER_AREA)


def _indigo_ocr_raw_en_banda_superior(img_bgr: np.ndarray) -> str:
    global _indigo_easyocr_reader
    modo = (os.environ.get("INDIGO_VENTANA_OCR") or "off").strip().lower()
    if modo in ("", "0", "off", "no", "false"):
        return ""
    if modo == "auto":
        modo = "tesseract"
    roi, _ox, _oy = _recorte_roi(img_bgr, (0.0, 0.0, 1.0, 0.42))
    if roi.size == 0:
        return ""
    if modo == "easyocr":
        try:
            import easyocr  # type: ignore[import-untyped]

            if _indigo_easyocr_reader is None:
                _indigo_easyocr_reader = easyocr.Reader(["es", "en"], gpu=False, verbose=False)
            reader = _indigo_easyocr_reader
            parts: list[str] = []
            for t in reader.readtext(roi, detail=0, paragraph=True):
                if isinstance(t, str):
                    parts.append(t)
            return "\n".join(parts)
        except Exception:
            return ""
    if modo in ("tesseract", "1", "on", "yes", "si", "sí"):
        try:
            import pytesseract  # type: ignore[import-untyped]
            from PIL import Image  # type: ignore[import-untyped]

            rgb = cv2.cvtColor(roi, cv2.COLOR_BGR2RGB)
            pil_img = Image.fromarray(rgb)
            return pytesseract.image_to_string(pil_img, lang="spa+eng") or ""
        except Exception:
            return ""
    return ""


def _indigo_ocr_bonus_y_etiquetas(img_bgr: np.ndarray) -> tuple[float, list[str]]:
    """Bonus 0..~0.18 sumado al score de plantilla si OCR encuentra textos típicos de Vie (opcional)."""
    raw = _indigo_ocr_raw_en_banda_superior(img_bgr)
    if not raw.strip():
        return 0.0, []
    n = _norm_etiqueta(raw)
    compact = n.replace(" ", "")
    hits: list[str] = []
    bonus = 0.0
    checks = (
        (("consultahistorias", "consultadehistorias"), 0.07, "consulta historias"),
        (("vieehr", "vie ehr"), 0.07, "Vie EHR"),
        (("paciente:", "paciente :"), 0.05, "Paciente"),
    )
    for keys, b, label in checks:
        ok = any(k in compact or k in n for k in keys)
        if ok:
            hits.append(label)
            bonus += b
    return min(0.22, bonus), hits


def _mejor_hwnd_visual_todas_top_level(rutas_tpl: list[Path]) -> tuple[int, float, str, Path] | None:
    """Enumera top-level visibles, captura cada una, max matchTemplate entre plantillas; sin filtro GetWindowText."""
    if not rutas_tpl:
        return None
    min_acept = _float_env("INDIGO_VENTANA_ACEPTAR_SIM_MIN", 0.10)
    need_tpl = _float_env("INDIGO_VENTANA_MATCH_MIN", 0.24)
    ocr_fb = _bool_env("INDIGO_VENTANA_OCR_FALLBACK", False)
    try:
        ocr_fb_min = int((os.environ.get("INDIGO_VENTANA_OCR_FALLBACK_MIN") or "2").strip())
    except ValueError:
        ocr_fb_min = 2
    ocr_fb_min = max(1, min(5, ocr_fb_min))
    debug_dir = _dir_debug_escaneo_ventana_indigo()
    debug_log = _bool_env("INDIGO_VENTANA_DEBUG_SCAN", False)
    anotar = _bool_env("INDIGO_VENTANA_DEBUG_ANOTAR", False)

    mejores: list[
        tuple[float, float, float, int, str, Path, list[str], tuple[int, int, int, int] | None, np.ndarray | None]
    ] = []
    for hwnd, title in _enum_toplevel_para_match_visual_indigo():
        try:
            h = int(hwnd)
        except (TypeError, ValueError):
            continue
        try:
            img, _lx, _ty = _origen_y_captura_para_match_hwnd(h, "")
        except (RuntimeError, OSError):
            continue
        cls = _clase_ventana(h)
        best_sc = 0.0
        best_tpl = rutas_tpl[0]
        best_rect: tuple[int, int, int, int] | None = None
        for tpl_path in rutas_tpl:
            rois = _rois_busqueda_para_plantilla_ventana(tpl_path)
            if debug_dir is not None and anotar:
                sc_d, rect_d = _max_similitud_tpl_haystack_detalle(img, tpl_path, rois)
            else:
                sc_d = _max_similitud_plantilla_haystack(img, tpl_path, rois)
                rect_d = None
            if sc_d > best_sc:
                best_sc = sc_d
                best_tpl = tpl_path
                best_rect = rect_d
        ocr_b, ocr_hits = _indigo_ocr_bonus_y_etiquetas(img)
        comb = min(1.0, float(best_sc) + float(ocr_b))
        img_dbg = img if debug_dir is not None else None
        mejores.append((comb, float(best_sc), float(ocr_b), h, title, best_tpl, ocr_hits, best_rect, img_dbg))
        if debug_log:
            print(
                f"  INDIGO escaneo ventana: hwnd={h} class={cls!r} title={title!r} "
                f"tpl_sim={best_sc:.3f} ocr_bonus={ocr_b:.3f} comb={comb:.3f} tpl={best_tpl.name!r}"
            )
        if debug_dir is not None and img_dbg is not None:
            shot = _resize_para_debug_captura(img_dbg)
            if anotar and best_rect is not None:
                x, y, tw, th = best_rect
                cv2.rectangle(shot, (x, y), (x + tw, y + th), (0, 220, 0), 2)
                for i, line in enumerate(
                    (
                        f"hwnd={h}",
                        f"tpl_sim={best_sc:.3f}",
                        f"ocr={ocr_b:.3f}",
                        f"cls={_sanitizar_fragmento_nombre_archivo(cls)}",
                    )
                ):
                    cv2.putText(
                        shot,
                        line,
                        (6, 22 + i * 22),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        0.52,
                        (40, 255, 40),
                        2,
                        cv2.LINE_AA,
                    )
            fn = (
                f"cand_{h}_{_sanitizar_fragmento_nombre_archivo(cls)}_"
                f"{_sanitizar_fragmento_nombre_archivo(title)}_{best_sc:.3f}.png"
            )
            try:
                cv2.imwrite(str(debug_dir / fn), shot)
            except cv2.error:
                pass

    if not mejores:
        return None
    mejores.sort(key=lambda r: (-r[0], -r[1], -r[2]))
    comb, raw_sc, ocr_b, hwnd, title, best_tpl, ocr_hits, _rect, _img = mejores[0]
    if raw_sc <= 1e-4 and ocr_b <= 1e-4:
        return None
    ocr_ok = ocr_fb and len(ocr_hits) >= ocr_fb_min
    ocr_min_tpl = _float_env("INDIGO_VENTANA_OCR_FALLBACK_PLANTILLA_MIN", 0.05)
    if raw_sc < min_acept:
        return None
    if raw_sc < need_tpl and not (ocr_ok and raw_sc >= ocr_min_tpl):
        return None
    if debug_log:
        print(
            f"INDIGO escaneo ventana: ELEGIDO hwnd={hwnd} tpl_sim={raw_sc:.3f} comb={comb:.3f} "
            f"tpl={best_tpl.name!r} title={title!r}"
        )
    return hwnd, raw_sc, title, best_tpl


def _rutas_plantillas_ventana_indigo() -> list[Path]:
    """PNG para localizar la ventana por captura (barra de título a veces vacía o genérica).

    Orden: opcional INDIGO_PLANTILLA_VENTANA (se fusiona, no sustituye toda la lista salvo modo exclusivo)
    → «1 *.png» (campo Paciente/cédula; anclaje principal, estilo H40 primer clic)
    → «0 *.png» y *Pantalla inicial* (opcional) → PNG con «Vie»+«EHR» en el nombre (logo esquina)
    → «8 *.png» solo si el nombre sugiere Paciente/cédula (compat.).

    INDIGO_PLANTILLA_VENTANA_VIE_EHR: ruta opcional a PNG pequeño del logo Vie EHR (se antepone a la lista).

    INDIGO_VENTANA_PLANTILLA_1_PRIMERO=0: orden clásico (0 e inicial antes que 1).

    INDIGO_PLANTILLA_VENTANA_EXCLUSIVA=1: solo se usa la ruta de INDIGO_PLANTILLA_VENTANA (comportamiento antiguo).
    INDIGO_VENTANA_SOLO_PLANTILLA_8: ignorado para cargar archivos (antes omitía la plantilla 0 y rompía el anclaje).
    """
    out: list[Path] = []

    def add(p: Path) -> None:
        q = p.resolve()
        if q.is_file() and q not in out:
            out.append(q)

    exclusiva = (os.environ.get("INDIGO_PLANTILLA_VENTANA_EXCLUSIVA") or "").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
        "si",
        "sí",
    )
    env = (os.environ.get("INDIGO_PLANTILLA_VENTANA") or "").strip().strip('"')
    if env:
        p = Path(env)
        if p.is_file():
            add(p)
        else:
            q = (ROOT / env) if not Path(env).is_absolute() else Path(env)
            if q.is_file():
                add(q)
        if exclusiva and out:
            return _priorizar_rutas_plantilla_vie_ehr(out)

    env_ve = (os.environ.get("INDIGO_PLANTILLA_VENTANA_VIE_EHR") or "").strip().strip('"')
    if env_ve:
        pve = Path(env_ve) if Path(env_ve).is_absolute() else (ROOT / env_ve)
        if pve.is_file():
            out.insert(0, pve.resolve())

    if not TEMPLATE_DIR.is_dir():
        return _priorizar_rutas_plantilla_vie_ehr(out)
    raw_1p = (os.environ.get("INDIGO_VENTANA_PLANTILLA_1_PRIMERO") or "1").strip().lower()
    uno_primero = raw_1p not in ("0", "false", "no", "off")

    def add_cero_e_inicial() -> None:
        for f in sorted(TEMPLATE_DIR.glob("0 *.png")):
            add(f)
        for pat in ("*Pantalla inicial*.png", "*pantalla inicial*.png"):
            for f in sorted(TEMPLATE_DIR.glob(pat)):
                add(f)

    def add_unos() -> None:
        for f in sorted(TEMPLATE_DIR.glob("1 *.png")):
            add(f)

    if uno_primero:
        add_unos()
        add_cero_e_inicial()
    else:
        add_cero_e_inicial()
        add_unos()
    for f in sorted(TEMPLATE_DIR.glob("8 *.png")):
        low = f.name.lower()
        if any(
            x in low
            for x in (
                "paciente",
                "cedula",
                "cédula",
                "excel",
                "input",
                "binocular",
                "documento",
            )
        ):
            add(f)
    for pat in ("*Vie*EHR*.png", "*vie*ehr*.png"):
        for f in sorted(TEMPLATE_DIR.glob(pat)):
            if f.is_file() and f.suffix.lower() == ".png":
                add(f)
    return _priorizar_rutas_plantilla_vie_ehr(out)


def _mejor_hwnd_entre_plantillas_candidatas(rutas: list[Path]) -> tuple[int, float, str, Path] | None:
    """Elige HWND por la mejor similitud entre plantillas de anclaje (0, 1, …).

    Por defecto se evalúan **todas** las rutas a la vez (típico: ya está el formulario y la 0 no
    coincide). Con INDIGO_VENTANA_COMBINAR_PLANTILLAS_ANCLA=0 se conserva el modo antiguo: solo
    plantilla(s) «0 / Pantalla inicial» primero, luego el resto (y el aviso si la 0 no matchea).
    """
    global _log_fallback_plantilla0_emitido
    prim = [p for p in rutas if _plantilla_es_pantalla_inicial_busqueda_ventana(p)]
    mezclar_desde_inicio = (os.environ.get("INDIGO_VENTANA_BUSCAR_TAMBIEN_CON_PLANTILLA_8") or "").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
        "si",
        "sí",
    )
    raw_comb = (os.environ.get("INDIGO_VENTANA_COMBINAR_PLANTILLAS_ANCLA") or "").strip().lower()
    if raw_comb in ("0", "false", "no", "off"):
        combinar_todas = False
    elif raw_comb in ("1", "true", "yes", "on", "si", "sí"):
        combinar_todas = True
    else:
        combinar_todas = True  # defecto: evita aviso engañoso si solo la 1 es visible
    combinar_todas = combinar_todas or mezclar_desde_inicio
    if _bool_env("INDIGO_VENTANA_BUSQUEDA_VISUAL_TODAS", True):
        hit_v = _mejor_hwnd_visual_todas_top_level(rutas)
        if hit_v is not None:
            return hit_v
    if prim and not combinar_todas:
        best: tuple[float, int, str, Path] | None = None
        for tpl_path in prim:
            hit = _mejor_hwnd_por_plantilla_ventana(tpl_path)
            if hit is None:
                continue
            hwnd, sc, tit = hit
            if best is None or sc > best[0]:
                best = (sc, hwnd, tit, tpl_path)
        if best is not None:
            sc, hwnd, tit, tpl_path = best
            return hwnd, sc, tit, tpl_path
        if not _log_fallback_plantilla0_emitido:
            print(
                "Ventana IndiGO: la plantilla «0 / Pantalla inicial» no coincide con ninguna ventana visible "
                "(si IndiGO ya está en el formulario de cédula, es normal). Se prueban plantilla 1 y demás "
                "anclajes. Renueve el PNG 0 o baje INDIGO_VENTANA_MATCH_MIN. "
                "El defecto ya ancla con 0+1 a la vez (INDIGO_VENTANA_COMBINAR_PLANTILLAS_ANCLA); "
                "este aviso solo aparece con COMBINAR=0."
            )
            _log_fallback_plantilla0_emitido = True
    best_all: tuple[float, int, str, Path] | None = None
    for tpl_path in rutas:
        hit = _mejor_hwnd_por_plantilla_ventana(tpl_path)
        if hit is None:
            continue
        hwnd, sc, tit = hit
        if best_all is None or sc > best_all[0]:
            best_all = (sc, hwnd, tit, tpl_path)
    if best_all is None:
        return None
    sc, hwnd, tit, tpl_path = best_all
    return hwnd, sc, tit, tpl_path


def _mejor_hwnd_por_plantilla_ventana(tpl_path: Path) -> tuple[int, float, str] | None:
    """Entre ventanas visibles, elige la que mejor coincide con la plantilla (captura del HWND)."""
    global _log_segunda_pasada_ventana_emitido
    tpl = _cv2_imread_path(tpl_path)
    if tpl is None or tpl.size == 0:
        return None
    if tpl.ndim == 3 and tpl.shape[2] == 4:
        tpl = cv2.cvtColor(tpl, cv2.COLOR_BGRA2BGR)
    raw_min = (os.environ.get("INDIGO_VENTANA_MATCH_MIN") or "").strip().replace(",", ".")
    try:
        umbral = float(raw_min) if raw_min else 0.24
    except ValueError:
        umbral = 0.24

    floor = _float_env("INDIGO_VENTANA_MATCH_FLOOR", 0.10)
    rois_busqueda = tuple(ROIS_BUSQUEDA_VENTANA_POR_PLANTILLA)
    umbrales = (
        umbral,
        umbral - 0.05,
        umbral - 0.1,
        0.22,
        0.20,
        0.17,
        0.15,
        0.13,
        0.11,
        0.10,
        0.09,
        0.08,
        0.07,
        0.065,
    )

    try:
        min_pt = int((os.environ.get("INDIGO_VENTANA_MIN_PUNTUACION_TITULO") or "4").strip())
    except ValueError:
        min_pt = 4
    raw_re = (os.environ.get("INDIGO_VENTANA_REINTENTAR_SIN_PUNTUACION_TITULO") or "1").strip().lower()
    reintentar_sin_pt = raw_re not in ("0", "false", "no", "off")

    def _recolectar_hwnd_por_plantilla(filtrar_puntuacion_titulo: bool) -> list[tuple[float, int, int, int, str]]:
        out: list[tuple[float, int, int, int, str]] = []
        for hwnd, title in _enum_candidatos_hwnd_para_plantilla_indigo():
            if _titulo_es_terminal_o_consola_rpa(title):
                continue
            try:
                h = int(hwnd)
            except (TypeError, ValueError):
                continue
            if _hwnd_es_shell_escritorio_o_falso_match_plantilla(h, title):
                continue
            if _hwnd_es_ventana_sistema_excluir_de_match_indigo(h, title):
                continue
            if _hwnd_es_explorador_archivos_windows(h):
                continue
            if _hwnd_es_excel_u_hoja_calculo(h, title):
                continue
            if _hwnd_es_cliente_navegador_o_correo(h, title):
                continue
            raw_tit = (title or "").strip()
            if _titulo_bloquea_busqueda_plantilla_ventana_indigo(raw_tit):
                continue
            pt = _puntuacion_titulo_favorece_vie_indigo(raw_tit)
            chromium_exempt = "WebView" in raw_tit or "Chromium" in raw_tit
            if filtrar_puntuacion_titulo and raw_tit and not chromium_exempt and pt < min_pt:
                continue
            left, top, right, bottom = rect_pantalla(h)
            w, hgt = right - left, bottom - top
            mw, mh = _dimensiones_minimas_ventana_para_plantilla()
            if w < mw or hgt < mh:
                continue
            try:
                img, _lx, _ty = _origen_y_captura_para_match_hwnd(h, "")
            except (RuntimeError, OSError):
                continue
            hit = None
            for thr in umbrales:
                t_eff = max(floor, float(thr))
                for roi_rel in rois_busqueda:
                    hit_try = _mejor_match_multiescala(
                        img, tpl, roi_rel, "left", t_eff, escalas=TEMPLATE_SCALES_VENTANA
                    )
                    if hit_try is not None:
                        hit = hit_try
                        break
                if hit is not None:
                    break
            if hit is not None:
                sc = float(hit[2])
                area = max(1, w * hgt)
                prio = _prioridad_titulo_ventana_ve_indigo(title or "")
                out.append((sc, pt, prio, area, h, title or "(sin titulo)"))
        return out

    scored = _recolectar_hwnd_por_plantilla(True)
    if not scored and reintentar_sin_pt:
        scored = _recolectar_hwnd_por_plantilla(False)
        if scored and not _log_segunda_pasada_ventana_emitido:
            _log_segunda_pasada_ventana_emitido = True
            print(
                "Ventana IndiGO: 2.ª pasada de anclaje por plantilla (sin filtro "
                "INDIGO_VENTANA_MIN_PUNTUACION_TITULO). Desactive con INDIGO_VENTANA_REINTENTAR_SIN_PUNTUACION_TITULO=0 "
                "si hubiera falso positivo."
            )
    if not scored:
        return None
    scored.sort(key=lambda x: (-x[0], -x[1], -x[2], -x[3]))
    sc, _pt, _prio, _area, hwnd, tit = scored[0]
    min_acept = _float_env("INDIGO_VENTANA_ACEPTAR_SIM_MIN", 0.10)
    if sc < min_acept:
        return None
    return hwnd, sc, tit


def _buscar_hwnd_por_titulo_indigo() -> int | None:
    pistas_fuertes = (
        "seleccione la empresa",
        "vie cloud",
        "viecloud",
        "vie ehr",
        "vieehr",
        "cloud platform",
        "viecloudplatform",
        "indigo international",
        "consulta historias",
        "consulta de historias",
        "historia clinica",
        "historias clinicas",
    )
    pistas_secundarias = ("area de trabajo", "consulta externa")
    candidatos: list[tuple[int, str, int]] = []
    for hwnd, title in _enum_windows():
        if _titulo_es_terminal_o_consola_rpa(title):
            continue
        try:
            hi = int(hwnd)
        except (TypeError, ValueError):
            continue
        if _hwnd_es_shell_escritorio_o_falso_match_plantilla(hi, title):
            continue
        if _hwnd_es_ventana_sistema_excluir_de_match_indigo(hi, title):
            continue
        if _hwnd_es_dialogo_impresion_o_guardar_excluir_indigo(hi, title):
            continue
        if _hwnd_es_explorador_archivos_windows(hi):
            continue
        if _hwnd_es_excel_u_hoja_calculo(hi, title):
            continue
        if _hwnd_es_cliente_navegador_o_correo(hi, title):
            continue
        if _titulo_bloquea_busqueda_plantilla_ventana_indigo(title or ""):
            continue
        n = _norm_etiqueta(title)
        prio = 0
        if n:
            compact_t = n.replace(" ", "")
            if re.search(r"indigo\d{2,}", compact_t):
                prio = max(prio, 2)
            if any(p in n for p in pistas_fuertes):
                prio = max(prio, 2)
            elif any(p in n for p in pistas_secundarias):
                prio = max(prio, 1)
            elif re.search(r"\bpaciente\b", n):
                # «Pacientes_….xlsx» contiene «paciente» como subcadena; exigir palabra suelta.
                prio = max(prio, 1)
            elif "indigo" in n and len(n) >= 12:
                # Títulos tipo «… IndiGO …» en el módulo (evitar consola/RPA y editores).
                if "cursor" in n or "notepad" in n or ".py" in n:
                    pass
                elif "run_indigo" in n.replace(" ", "") or "indigo_historias" in n.replace(" ", ""):
                    pass
                elif "powershell" in n or "cmd.exe" in n:
                    pass
                else:
                    prio = 1
            # No usar «indigo» suelto en títulos cortos o de herramientas (sigue arriba).
        if prio > 0:
            candidatos.append((hi, title or "", prio))
    if not candidatos:
        return None
    candidatos.sort(key=lambda x: (-x[2], -len(x[1])))
    return candidatos[0][0]


def _hwnd_tras_pausa_foco_si_plantilla_encaja(rutas_tpl: list[Path]) -> int | None:
    """INDIGO_FOCO_INDIGO_SEG: espera N s, toma la ventana al frente y valida con las plantillas de ventana."""
    raw = (os.environ.get("INDIGO_FOCO_INDIGO_SEG") or "").strip()
    if not raw:
        return None
    try:
        seg = float(raw.replace(",", "."))
    except ValueError:
        return None
    if seg <= 0:
        return None
    umin = _float_env("INDIGO_FOCO_PLANTILLA_MIN", 0.14)
    rois = list(ROIS_BUSQUEDA_VENTANA_POR_PLANTILLA)
    print(
        f"Ventana IndiGO: espere {seg:.0f} s con Vie/IndiGO al frente y haga clic en su barra de título "
        f"(INDIGO_FOCO_INDIGO_SEG). Se validará similitud ≥ {umin:.2f} (INDIGO_FOCO_PLANTILLA_MIN)."
    )
    time.sleep(seg)
    hwnd = int(user32.GetForegroundWindow())
    if hwnd <= 0 or not user32.IsWindow(hwnd) or not user32.IsWindowVisible(hwnd):
        print("Ventana al frente: HWND inválido o no visible.")
        return None
    tit = win32gui.GetWindowText(hwnd)
    if _hwnd_es_ide_cursor_o_editor(hwnd, tit):
        print(
            f"La ventana al frente es un IDE/editor («{tit[:70]}»), no Vie Cloud. "
            "Ponga IndiGO al frente o use INDIGO_HWND."
        )
        return None
    if _hwnd_es_explorador_archivos_windows(hwnd):
        return None
    if _hwnd_es_excel_u_hoja_calculo(hwnd, tit):
        print(
            "La ventana al frente es Excel/LibreOffice Calc (no Vie Cloud). "
            "Cierre el libro o ponga IndiGO al frente y vuelva a intentar."
        )
        return None
    if _hwnd_es_cliente_navegador_o_correo(hwnd, tit):
        print(
            "La ventana al frente es navegador o correo (no Vie Cloud). "
            "Ponga IndiGO al frente o use INDIGO_HWND."
        )
        return None
    if _titulo_bloquea_busqueda_plantilla_ventana_indigo(tit):
        print(
            "La ventana al frente parece Teams u otra app de colaboración (no Vie). "
            "Ponga IndiGO al frente o use INDIGO_HWND."
        )
        return None
    raw_fg = (tit or "").strip()
    pt_fg = _puntuacion_titulo_favorece_vie_indigo(raw_fg)
    try:
        min_pt_fg = int((os.environ.get("INDIGO_VENTANA_MIN_PUNTUACION_TITULO") or "4").strip())
    except ValueError:
        min_pt_fg = 4
    if raw_fg and not ("WebView" in raw_fg or "Chromium" in raw_fg) and pt_fg < min_pt_fg:
        print(
            f"La ventana al frente no tiene pistas de título de Vie/IndiGO (título {tit!r}). "
            "Haga clic en la barra de título de IndiGO o use INDIGO_HWND."
        )
        return None
    if _hwnd_es_shell_escritorio_o_falso_match_plantilla(hwnd, tit):
        return None
    if _hwnd_es_ventana_sistema_excluir_de_match_indigo(hwnd, tit):
        return None
    left, top, right, bottom = rect_pantalla(hwnd)
    w, hgt = right - left, bottom - top
    mw, mh = _dimensiones_minimas_ventana_para_plantilla()
    if w < mw or hgt < mh:
        print(f"Ventana al frente demasiado pequeña ({w}×{hgt}); mínimo {mw}×{mh}.")
        return None
    mejor = 0.0
    mejor_tpl: Path | None = None
    for tpl_path in rutas_tpl:
        s = _max_similitud_plantilla_hwnd(hwnd, tpl_path, rois)
        if s > mejor:
            mejor = s
            mejor_tpl = tpl_path
    if mejor_tpl is None or mejor < umin:
        print(
            f"Ventana al frente no coincide con plantillas de IndiGO (mejor similitud {mejor:.2f} < {umin:.2f}; "
            f"título {tit!r})."
        )
        return None
    print(
        f"Ventana IndiGO por foco del usuario (similitud {mejor:.2f} con {mejor_tpl.name!r}; "
        f"título {tit!r}; hwnd={hwnd})"
    )
    return hwnd


def _intentar_hwnd_frente_si_plantilla_indigo(rutas_tpl: list[Path]) -> int | None:
    """Si IndiGO está al frente, ancla por similitud con plantillas 0/1 sin el filtro estricto de título.

    Útil cuando el título de la barra no coincide con las pistas o está vacío en otro HWND padre.
    Desactivar: INDIGO_FRENTE_RAPIDO=0
    """
    raw = (os.environ.get("INDIGO_FRENTE_RAPIDO") or "1").strip().lower()
    if raw in ("0", "false", "no", "off"):
        return None
    if not rutas_tpl:
        return None
    hwnd = int(user32.GetForegroundWindow())
    if hwnd <= 0 or not user32.IsWindow(hwnd) or not user32.IsWindowVisible(hwnd):
        return None
    tit = win32gui.GetWindowText(hwnd)
    if _hwnd_es_ide_cursor_o_editor(hwnd, tit):
        return None
    if _hwnd_es_shell_escritorio_o_falso_match_plantilla(hwnd, tit):
        return None
    if _hwnd_es_ventana_sistema_excluir_de_match_indigo(hwnd, tit):
        return None
    if _hwnd_es_dialogo_impresion_o_guardar_excluir_indigo(hwnd, tit):
        return None
    if _hwnd_es_explorador_archivos_windows(hwnd):
        return None
    if _hwnd_es_excel_u_hoja_calculo(hwnd, tit):
        return None
    if _hwnd_es_cliente_navegador_o_correo(hwnd, tit):
        return None
    if _titulo_bloquea_busqueda_plantilla_ventana_indigo(tit):
        return None
    left, top, right, bottom = rect_pantalla(hwnd)
    w, hgt = right - left, bottom - top
    mw, mh = _dimensiones_minimas_ventana_para_plantilla()
    if w < mw or hgt < mh:
        return None
    mx = _max_sim_ventana_entre_plantillas_ancla(hwnd, rutas_tpl)
    umin = _float_env("INDIGO_FRENTE_RAPIDO_PLANTILLA_MIN", 0.08)
    if mx < umin:
        return None
    print(
        f"Ventana IndiGO por ventana al frente (similitud máx. {mx:.2f} ≥ {umin:.2f}; título {tit!r}; hwnd={hwnd}). "
        "Deje IndiGO activo antes de iniciar; INDIGO_FRENTE_RAPIDO_PLANTILLA_MIN / =0 si molesta."
    )
    return hwnd


def _debug_log_titulos_ventanas_visibles() -> None:
    """INDIGO_LOG_TITULOS_TOPLEVEL=1: imprime GetWindowText de ventanas visibles (diagnóstico de anclaje)."""
    raw = (os.environ.get("INDIGO_LOG_TITULOS_TOPLEVEL") or "").strip().lower()
    if raw not in ("1", "true", "yes", "on", "si", "sí"):
        return
    print("INDIGO_LOG_TITULOS_TOPLEVEL: ventanas visibles con título no vacío (GetWindowText):")
    n = 0
    vio_vie = False
    for _hwnd, title in _enum_windows():
        t = (title or "").strip()
        if not t:
            continue
        nt = _norm_etiqueta(t)
        if any(
            x in nt
            for x in (
                "consulta historias",
                "consulta de historias",
                "vie ehr",
                "viecloud",
                "vie cloud",
                "indigo",
                "historia clinica",
                "historias clinicas",
            )
        ):
            vio_vie = True
        if len(t) > 100:
            t = t[:97] + "..."
        print(f"  {t!r}")
        n += 1
        if n >= 40:
            break
    if n == 0:
        print(
            "  (ninguna; Vie/IndiGO a veces usa barra vacía — use INDIGO_HWND o deje IndiGO al frente "
            "tras INDIGO_FOCO_INDIGO_SEG.)"
        )
    elif not vio_vie:
        print(
            "  — Ningún título de la lista coincide con Vie/IndiGO típico («Consulta historias», Vie EHR, IndiGO…). "
            "Abra Vie Cloud, entre a «Consulta historias», haga clic en la barra de esa ventana y vuelva a ejecutar; "
            "o use INDIGO_HWND. Si la barra va vacía (solo WebView), el anclaje usa ventanas sin título salvo que "
            "INDIGO_INCLUIR_CHROMIUM_TOPLEVEL_SIN_TITULO=0 las excluya."
        )


def encontrar_hwnd_indigo(timeout: float | None = None) -> int:
    """Ventana Vie Cloud / IndiGO: INDIGO_HWND, foco, frente rápido, escaneo visual de todas las top-level
    (defecto), plantilla en templates/, y opcionalmente título GetWindowText."""
    global _log_fallback_plantilla0_emitido
    global _log_segunda_pasada_ventana_emitido
    _log_fallback_plantilla0_emitido = False
    _log_segunda_pasada_ventana_emitido = False
    if timeout is None:
        timeout = _float_env("INDIGO_ENCONTRAR_HWND_TIMEOUT", 60.0)
    raw = (os.environ.get("INDIGO_HWND") or "").strip()
    if raw.isdigit() and user32.IsWindow(int(raw)):
        print(f"Usando INDIGO_HWND={raw}")
        return int(raw)

    rutas_tpl = _rutas_plantillas_ventana_indigo()
    h_foco = _hwnd_tras_pausa_foco_si_plantilla_encaja(rutas_tpl)
    if h_foco is not None:
        return h_foco
    t0 = time.time()
    ultimo_scan_tpl = -100.0
    ultimo_frente = -100.0
    while time.time() - t0 < timeout:
        now = time.time()
        if rutas_tpl and (now - ultimo_frente) >= 0.35:
            ultimo_frente = now
            h_fg = _intentar_hwnd_frente_si_plantilla_indigo(rutas_tpl)
            if h_fg is not None:
                return h_fg
        if rutas_tpl and (now - ultimo_scan_tpl) >= 1.0:
            ultimo_scan_tpl = now
            hit = _mejor_hwnd_entre_plantillas_candidatas(rutas_tpl)
            if hit is not None:
                h2, sc, tit, tplv = hit
                print(
                    f"Ventana IndiGO por plantilla {tplv.name!r} "
                    f"(similitud {sc:.2f}; título barra: {tit!r}; hwnd={h2})"
                )
                return h2
        h = _buscar_hwnd_por_titulo_indigo()
        if h is not None:
            sin_val = (os.environ.get("INDIGO_SIN_VALIDAR_PLANTILLA_TRAS_TITULO") or "").strip().lower() in (
                "1",
                "true",
                "yes",
                "on",
                "si",
                "sí",
            )
            if sin_val or not rutas_tpl:
                print(
                    "Ventana IndiGO por título de barra (INDIGO_SIN_VALIDAR_PLANTILLA_TRAS_TITULO: "
                    "sin comprobar plantilla 0/1 en la captura)."
                )
                return h
            min_vis = _float_env("INDIGO_VALIDAR_TRAS_TITULO_MIN", 0.10)
            mx = _max_sim_ventana_entre_plantillas_ancla(h, rutas_tpl)
            if mx >= min_vis:
                print(
                    f"Ventana IndiGO por título de barra (coincidencia visual máx. {mx:.2f} ≥ {min_vis:.2f} "
                    "con plantillas de anclaje)."
                )
                return h
            print(
                f"Ventana con título reconocido pero captura poco compatible con plantillas 0/1 "
                f"(máx. similitud {mx:.2f} < {min_vis:.2f}). Se ignora ese HWND y se sigue por plantilla. "
                "Traiga IndiGO al frente, actualice los PNG, o use INDIGO_SIN_VALIDAR_PLANTILLA_TRAS_TITULO=1."
            )
        time.sleep(0.25)

    msg = (
        "No se encontró Vie Cloud: ninguna ventana encajó por escaneo visual top-level (plantillas + ROI), "
        "ni por título de Windows (GetWindowText; pistas p. ej. «Consulta historias», «Vie EHR», «IndiGO…») "
        "con validación visual, con similitud suficiente. "
        "No se exige que el título diga literalmente «Vie Cloud»."
    )
    if not rutas_tpl:
        msg += (
            " Coloque al menos «1 …cédula/Paciente….png» en templates/ (la plantilla «0» es opcional), "
            "o INDIGO_PLANTILLA_VENTANA=ruta\\archivo.png (fusiona con templates salvo INDIGO_PLANTILLA_VENTANA_EXCLUSIVA=1), "
            "o INDIGO_HWND."
        )
    else:
        nombres = ", ".join(p.name for p in rutas_tpl[:6])
        if len(rutas_tpl) > 6:
            nombres += ", …"
        msg += (
            f" Revisó: {nombres}. Añada un PNG pequeño con «Vie» y «EHR» en el nombre (logo arriba-derecha) o "
            "INDIGO_PLANTILLA_VENTANA_VIE_EHR=ruta.png; traiga IndiGO al frente; actualice plantillas; "
            "INDIGO_VENTANA_DEBUG_SCAN=1 (guarda candidatos en salidas/debug_ventana_indigo/…), "
            "INDIGO_VENTANA_DEBUG_ANOTAR=1 (rectángulo y texto en el PNG), INDIGO_VENTANA_BUSQUEDA_VISUAL_TODAS=0 "
            "solo si necesita el modo antiguo (título + Chromium sin título); "
            "INDIGO_VENTANA_MATCH_MIN=0.20, INDIGO_VENTANA_MATCH_FLOOR=0.08, INDIGO_FOCO_INDIGO_SEG=4, "
            "INDIGO_ENCONTRAR_HWND_TIMEOUT=60, INDIGO_VENTANA_COMBINAR_PLANTILLAS_ANCLA=1 (defecto: anclar con 0+1+… "
            "a la vez), INDIGO_VENTANA_BUSCAR_TAMBIEN_CON_PLANTILLA_8=1 (alias antiguo «mezclar»), "
            "INDIGO_VENTANA_REINTENTAR_SIN_PUNTUACION_TITULO=1 (defecto: 2.ª pasada sin filtro de título; =0 la desactiva), "
            "INDIGO_INCLUIR_CHROMIUM_TOPLEVEL_SIN_TITULO=0 solo afecta el modo legado con INDIGO_VENTANA_BUSQUEDA_VISUAL_TODAS=0; "
            "INDIGO_FRENTE_RAPIDO=1 (defecto: ancla con la ventana al frente si las plantillas encajan), "
            "INDIGO_VENTANA_MIN_PUNTUACION_TITULO=0 (solo para saltar el filtro tipo Stone/Teams), "
            "INDIGO_VENTANA_OCR=tesseract (requiere Tesseract en PATH; opcional), INDIGO_VENTANA_OCR_FALLBACK=1, "
            "INDIGO_VENTANA_ACEPTAR_SIM_MIN=0.10 (anclaje por plantilla más laxo), INDIGO_LOG_TITULOS_TOPLEVEL=1 "
            "(lista títulos de ventanas visibles en consola), o INDIGO_HWND."
        )
    _debug_log_titulos_ventanas_visibles()
    raise TimeoutError(msg)


def _plantilla_png_prefiere_sobre(prev: Path, nuevo: Path) -> bool:
    """True si «nuevo» debe reemplazar «prev» (botón pequeño mejor que captura de ventana entera)."""
    pn, nn = prev.name.lower(), nuevo.name.lower()
    if "ventana" in pn and "ventana" not in nn:
        return True
    if "ventana" in nn and "ventana" not in pn:
        return False
    try:
        return nuevo.stat().st_size < prev.stat().st_size * 0.55
    except OSError:
        return False


def _map_id_str_a_plantilla_path() -> dict[str, Path]:
    out: dict[str, Path] = {}
    if not TEMPLATE_DIR.is_dir():
        return out
    for f in sorted(TEMPLATE_DIR.iterdir()):
        if not f.is_file() or f.suffix.lower() != ".png":
            continue
        m = re.match(r"^(\d+(?:\.\d*)?)\s+", f.name)
        if m:
            kid = m.group(1)
            if kid in out:
                if _plantilla_png_prefiere_sobre(out[kid], f):
                    out[kid] = f
            else:
                out[kid] = f
    return out


def _plantillas_png_por_numero() -> dict[int, str]:
    out: dict[int, str] = {}
    for kid, p in _map_id_str_a_plantilla_path().items():
        if "." in kid:
            continue
        if kid.isdigit():
            out[int(kid)] = p.name
    return out


def _parse_click_line(rest: str) -> tuple[str, tuple[float, float, float, float]]:
    ts = rest.split()
    if len(ts) >= 5:
        try:
            roi = (float(ts[-4]), float(ts[-3]), float(ts[-2]), float(ts[-1]))
            name = " ".join(ts[:-4]).strip()
            if name:
                return name, roi
        except ValueError:
            pass
    return rest.strip(), (0.02, 0.02, 0.98, 0.98)


def _resolver_ruta_plantilla_png(nombre: str) -> Path:
    nombre = (nombre or "").strip()
    if not nombre:
        raise ValueError("Nombre de plantilla vacío")
    m = re.match(r"^(\d+(?:\.\d*)?)\s+", nombre)
    if m:
        sid = _map_id_str_a_plantilla_path()
        kid = m.group(1)
        if kid in sid:
            return sid[kid].resolve()
    head = nombre.split(maxsplit=1)[0]
    if head.isdigit():
        n = int(head)
        mp = _plantillas_png_por_numero()
        if n in mp:
            p = TEMPLATE_DIR / mp[n]
            if p.is_file():
                return p.resolve()
        cand = sorted(TEMPLATE_DIR.glob(f"{n} *.png"))
        if cand:
            return cand[0].resolve()
    p = (TEMPLATE_DIR / nombre).resolve()
    if p.is_file():
        return p
    raise FileNotFoundError(
        f"No se encontró plantilla {nombre!r} en {TEMPLATE_DIR}. Use «1 nombre.png», «click 1», etc."
    )


def extraer_codigo_numerico_carpeta_ese(nombre_carpeta: str) -> str:
    """Prefijo numérico del nombre de carpeta E.S.E.: «009 Guarne» → «009», «014 Santamaria» → «014»."""
    s = (nombre_carpeta or "").strip()
    if not s:
        return ""
    m = re.match(r"^(\d+)", s)
    if m:
        return m.group(1)
    for part in s.split():
        if part.isdigit():
            return part
    return ""


def _ruta_tabla_motivo_obs_ese(xlsx_explicito: Path | None = None) -> Path:
    if xlsx_explicito is not None:
        return xlsx_explicito.expanduser().resolve()
    raw = (os.environ.get("INDIGO_TABLA_MOTIVO_OBS_XLSX") or "").strip().strip('"')
    if raw:
        p = Path(raw)
        return p.resolve() if p.is_absolute() else (ROOT / p).resolve()
    candidatos = (
        MOTIVO_OBS_ESE_XLSX_DEFAULT,
        INSUMOS_ROOT / "Opciones Motivo de consultas x ESE.xlsx",
    )
    for p in candidatos:
        if p.is_file():
            return p.resolve()
    return MOTIVO_OBS_ESE_XLSX_DEFAULT.resolve()


def _codigo_ese_coincide_celda_a(valor_celda: object, codigo_carpeta: str) -> bool:
    """True si la celda A de la tabla corresponde al código de la carpeta (009 ≈ 9)."""
    sc = (codigo_carpeta or "").strip()
    if not sc:
        return False
    sa = str(valor_celda if valor_celda is not None else "").strip()
    if not sa:
        return False
    if sa == sc:
        return True
    try:
        return int(float(sa.replace(",", "."))) == int(float(sc.replace(",", ".")))
    except ValueError:
        pass
    za, zc = sa.lstrip("0") or "0", sc.lstrip("0") or "0"
    return za == zc


def _fila_es_probable_encabezado_motivo_obs(row: tuple) -> bool:
    a = str(row[0] if row and len(row) > 0 else "").strip().lower()
    if not a:
        return False
    if "codigo" in a or "código" in a:
        return True
    return a in ("a", "code", "cod", "ese", "n°", "no", "num", "numero", "número")


def _parse_numero_opcion_motivo_columna_d(val: object) -> int | None:
    """Entero ≥ 1: posición de la opción en el desplegable (1 = primera fila). Celda vacía → None."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        n = int(float(s.replace(",", ".")))
    except ValueError:
        return None
    return n if n >= 1 else None


def motivo_y_observacion_desde_excel_ese(
    codigo_ese: str,
    nombre_carpeta: str,
    xlsx: Path | None = None,
) -> tuple[str, str, int | None]:
    """Lee catálogo: D = nº opción select motivo; E = texto de la opción (motivo); F = observación. A = código carpeta."""
    path = _ruta_tabla_motivo_obs_ese(xlsx)
    if not path.is_file():
        raise FileNotFoundError(
            f"No existe el catálogo de motivo/observación por E.S.E.:\n  {path}\n"
            "Créelo o defina INDIGO_TABLA_MOTIVO_OBS_XLSX (o use --tabla-motivo-obs-xlsx)."
        )
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("Para leer el catálogo .xlsx instale: pip install openpyxl") from e

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        filas = list(ws.iter_rows(min_row=1, max_col=6, values_only=True))
    finally:
        wb.close()

    if not filas:
        raise ValueError(f"{path.name}: la hoja activa está vacía.")

    codigo_ese = (codigo_ese or "").strip()
    nombre_carpeta = (nombre_carpeta or "").strip()
    inicio = 1 if len(filas) > 1 and _fila_es_probable_encabezado_motivo_obs(filas[0]) else 0

    def _cel(row: tuple, i: int) -> object:
        if not row or i >= len(row):
            return None
        return row[i]

    for row in filas[inicio:]:
        if not row:
            continue
        a = _cel(row, 0)
        if _fila_es_probable_encabezado_motivo_obs(row):
            continue
        coincide = False
        if codigo_ese and _codigo_ese_coincide_celda_a(a, codigo_ese):
            coincide = True
        elif nombre_carpeta and str(a or "").strip().casefold() == nombre_carpeta.casefold():
            coincide = True
        if not coincide:
            continue
        d_num = _parse_numero_opcion_motivo_columna_d(_cel(row, 3))
        texto_motivo = str(_cel(row, 4) or "").strip()
        obs = str(_cel(row, 5) or "").strip()
        if not texto_motivo and d_num is None:
            continue
        return texto_motivo, obs, d_num

    quien = f"código «{codigo_ese}»" if codigo_ese else f"nombre de carpeta «{nombre_carpeta}»"
    raise ValueError(
        f"No hay fila en {path.name} con columna A que coincida con {quien}. "
        "Revise el prefijo numérico de la carpeta (p. ej. «009 Nombre») o el texto en A."
    )


def _resolver_ruta_plantilla_codigo_ese(code: str) -> Path:
    """PNG del recorte del código en la tabla (un archivo por código o ruta con variable de entorno)."""
    code = (code or "").strip()
    if not code:
        raise ValueError("Código E.S.E. vacío: use carpeta tipo «009 NombreHospital».")
    envpat = (os.environ.get("INDIGO_PLANTILLA_CODIGO_ESE") or "").strip()
    if envpat:
        rel = envpat.replace("{code}", code).replace("{CODE}", code)
        p = Path(rel)
        if not p.is_absolute():
            p = (TEMPLATE_DIR / rel).resolve()
        else:
            p = p.resolve()
        if p.is_file():
            return p
    candidatos: list[Path] = [
        TEMPLATE_DIR / f"1_codigo_{code}.png",
        TEMPLATE_DIR / f"1 codigo {code}.png",
        TEMPLATE_DIR / f"codigo_ese_{code}.png",
        TEMPLATE_DIR / f"{code} codigo ese.png",
    ]
    sub = TEMPLATE_DIR / "codigos_ese"
    if sub.is_dir():
        candidatos.extend([sub / f"{code}.png", sub / f"{code}_codigo.png"])
    for p in candidatos:
        if p.is_file():
            return p.resolve()
    for g in sorted(TEMPLATE_DIR.glob(f"1_codigo_{code}*.png")):
        return g.resolve()
    # Compatibilidad: plantilla única antigua solo para 009 (mismo recorte del código en tabla).
    legacy = TEMPLATE_DIR / "1 seleccionar ese según carpeta escogida inicialmente.png"
    if code == "009" and legacy.is_file():
        return legacy.resolve()
    raise FileNotFoundError(
        f"No hay plantilla para el código {code!r}. Cree p. ej.:\n"
        f"  {TEMPLATE_DIR / f'1_codigo_{code}.png'}\n"
        f"(recorte pequeño del número {code} en la grilla, como la columna Código). "
        "O defina INDIGO_PLANTILLA_CODIGO_ESE=codigos_ese/{{code}}.png"
    )


def _resolver_click_desde_linea(rest: str, num_map: dict[int, str]) -> tuple[str, tuple[float, float, float, float]]:
    ts = rest.split()
    roi_def = (0.02, 0.02, 0.98, 0.98)
    sid = _map_id_str_a_plantilla_path()
    if ts and re.fullmatch(r"\d+(?:\.\d*)?", ts[0]):
        key = ts[0]
        if key in sid:
            archivo = sid[key].name
            if len(ts) == 5:
                try:
                    roi = (float(ts[1]), float(ts[2]), float(ts[3]), float(ts[4]))
                    return archivo, roi
                except ValueError:
                    pass
            return archivo, roi_def
    if ts and ts[0].isdigit():
        n = int(ts[0])
        if n in num_map and len(ts) == 1:
            return num_map[n], roi_def
        if n in num_map and len(ts) == 5:
            try:
                roi = (float(ts[1]), float(ts[2]), float(ts[3]), float(ts[4]))
                return num_map[n], roi
            except ValueError:
                pass
    return _parse_click_line(rest)


def cargar_pasos_flujo_indigo(path: Path) -> list[tuple]:
    if not path.is_file():
        raise FileNotFoundError(
            f"No existe {path}. Cree la secuencia (clics numerados, type_cedula, type_motivo_consulta, …)."
        )
    num_map = _plantillas_png_por_numero()
    pasos: list[tuple] = []
    for raw in path.read_text(encoding="utf-8-sig").splitlines():
        line = (raw or "").strip()
        if not line or line.startswith("#"):
            continue
        parts = line.split(None, 1)
        cmd = parts[0].lower()
        rest = parts[1].strip() if len(parts) > 1 else ""
        if cmd == "click":
            archivo, roi = _resolver_click_desde_linea(rest, num_map)
            rp = _resolver_ruta_plantilla_png(archivo)
            pasos.append(("click", rp.name, roi))
        elif cmd in ("click_right", "click_derecho", "right_click"):
            archivo, roi = _resolver_click_desde_linea(rest, num_map)
            rp = _resolver_ruta_plantilla_png(archivo)
            pasos.append(("click_right", rp.name, roi))
        elif cmd in ("click_folios", "clic_folios", "click_folio_calibrado"):
            pasos.append(("click_folios",))
        elif cmd in ("click_menu_folio", "clic_menu_folio"):
            archivo, roi = _resolver_click_desde_linea(rest, num_map)
            rp = _resolver_ruta_plantilla_png(archivo)
            pasos.append(("click_menu_folio", rp.name, roi))
        elif cmd in ("wait_template", "esperar_plantilla"):
            if not rest:
                raise ValueError(f"Falta plantilla en {path.name}: {line!r}")
            ts = rest.split()
            rest_tokens = ts[:]
            timeout_sec = 90.0
            min_sim: float | None = None
            while len(rest_tokens) >= 2:
                try:
                    last = float(rest_tokens[-1].replace(",", "."))
                except ValueError:
                    break
                if 0.08 <= last <= 0.99:
                    min_sim = last
                    rest_tokens = rest_tokens[:-1]
                    continue
                if 1.0 <= last <= 900.0:
                    timeout_sec = last
                    rest_tokens = rest_tokens[:-1]
                    continue
                break
            rest_id = " ".join(rest_tokens).strip()
            if not rest_id:
                raise ValueError(f"ID de plantilla inválido en {path.name}: {line!r}")
            archivo, roi = _resolver_click_desde_linea(rest_id, num_map)
            rp = _resolver_ruta_plantilla_png(archivo)
            pasos.append(("wait_template", rp.name, roi, timeout_sec, min_sim))
        elif cmd in ("wait_template_any", "esperar_cualquier_plantilla"):
            if not rest:
                raise ValueError(f"Falta plantilla(s) en {path.name}: {line!r}")
            ts = rest.split()
            rest_tokens = ts[:]
            timeout_sec = 90.0
            min_sim_any: float | None = None
            while len(rest_tokens) >= 2:
                try:
                    last = float(rest_tokens[-1].replace(",", "."))
                except ValueError:
                    break
                if 0.08 <= last <= 0.99:
                    min_sim_any = last
                    rest_tokens = rest_tokens[:-1]
                    continue
                if 1.0 <= last <= 900.0:
                    timeout_sec = last
                    rest_tokens = rest_tokens[:-1]
                    continue
                break
            if len(rest_tokens) == 1 and "," in rest_tokens[0]:
                ids_any = [x.strip() for x in rest_tokens[0].split(",") if x.strip()]
            else:
                ids_any = [t for t in rest_tokens if t.strip()]
            if not ids_any:
                raise ValueError(f"Sin IDs de plantilla en {path.name}: {line!r}")
            specs_any: list[tuple[str, tuple[float, float, float, float]]] = []
            for tid in ids_any:
                archivo_a, roi_a = _resolver_click_desde_linea(tid, num_map)
                rp_a = _resolver_ruta_plantilla_png(archivo_a)
                specs_any.append((rp_a.name, roi_a))
            pasos.append(("wait_template_any", tuple(specs_any), timeout_sec, min_sim_any))
        elif cmd == "click_optional":
            archivo, roi = _resolver_click_desde_linea(rest, num_map)
            try:
                rp = _resolver_ruta_plantilla_png(archivo)
                pasos.append(("click_optional", rp.name, roi))
            except FileNotFoundError:
                print(
                    f"Aviso ({path.name}): click_optional omitido — no existe plantilla para {archivo!r} "
                    f"(añada el PNG en templates/ si necesita rellenar combos)."
                )
        elif cmd == "type_cedula":
            pasos.append(("type_cedula",))
        elif cmd in ("type_cedula_guardar_pdf", "cedula_guardar_pdf", "nombre_pdf_cedula"):
            pasos.append(("type_cedula_guardar_pdf",))
        elif cmd in (
            "guardar_pdf_cedula",
            "guardar_pdf",
            "confirmar_guardar_pdf",
            "cerrar_impresion_tras_pdf",
            "cerrar_impresion",
        ):
            pasos.append((cmd,))
        elif cmd in ("esperar_guardar_pdf", "wait_guardar_pdf", "esperar_guardar_pdf_estable"):
            timeout_g = _float_env("INDIGO_GUARDAR_PDF_APARECER_TIMEOUT", 90.0)
            if rest:
                try:
                    timeout_g = float(rest.replace(",", "."))
                except ValueError:
                    pass
            pasos.append(("esperar_guardar_pdf", timeout_g))
        elif cmd in ("click_codigo_ese", "click_ese_carpeta"):
            pasos.append(("click_codigo_ese",))
        elif cmd == "type_motivo_consulta":
            pasos.append(("type_motivo_consulta",))
        elif cmd == "type_observacion":
            pasos.append(("type_observacion",))
        elif cmd in ("esperar_modal_imprimir", "esperar_imprimir_modal"):
            timeout_modal = _float_env("INDIGO_ESPERA_MODAL_IMPRIMIR_TIMEOUT", 300.0)
            min_sim_modal: float | None = None
            if rest:
                ts_m = rest.split()
                rest_tokens_m = ts_m[:]
                while len(rest_tokens_m) >= 2:
                    try:
                        last_m = float(rest_tokens_m[-1].replace(",", "."))
                    except ValueError:
                        break
                    if 0.08 <= last_m <= 0.99:
                        min_sim_modal = last_m
                        rest_tokens_m = rest_tokens_m[:-1]
                        continue
                    if 1.0 <= last_m <= 900.0:
                        timeout_modal = last_m
                        rest_tokens_m = rest_tokens_m[:-1]
                        continue
                    break
                if len(rest_tokens_m) == 1:
                    try:
                        timeout_modal = float(rest_tokens_m[0].replace(",", "."))
                    except ValueError:
                        pass
            pasos.append(("esperar_modal_imprimir", timeout_modal, min_sim_modal))
        elif cmd in ("wait_impresion_windows", "esperar_impresion", "wait_print"):
            timeout_imp = _float_env("INDIGO_IMPRESION_TIMEOUT", 180.0)
            if rest:
                try:
                    timeout_imp = float(rest.replace(",", "."))
                except ValueError:
                    pass
            pasos.append(("wait_impresion_windows", timeout_imp))
        elif cmd == "sleep" and rest:
            pasos.append(("sleep", float(rest.replace(",", "."))))
        elif cmd == "key" and rest:
            pasos.append(("key", rest.strip().lower()))
        else:
            raise ValueError(f"Línea no reconocida en {path.name}: {line!r}")
    if not pasos:
        raise ValueError(f"{path.name} no tiene pasos activos.")
    for p in pasos:
        if p[0] in ("click", "click_right", "wait_template", "wait_template_any"):
            if p[0] == "wait_template_any":
                for nombre_p, _roi_p in p[1]:
                    _resolver_ruta_plantilla_png(nombre_p)
            else:
                _resolver_ruta_plantilla_png(p[1])
        elif p[0] == "click_optional":
            try:
                _resolver_ruta_plantilla_png(p[1])
            except FileNotFoundError:
                pass
    return pasos


@dataclass
class SesionIndigo:
    motivo_consulta: str
    observacion: str
    cedula_actual: str = ""
    codigo_ese: str = ""
    # Si hay texto en columna E, se escribe; si no, se usa nº columna D con ↓ (1 = primera opción).
    motivo_opcion_numero: int | None = None


def _type_texto_campo(texto: str) -> None:
    keyboard.send_keys("^a{DELETE}{BACKSPACE}")
    time.sleep(0.06)
    keyboard.send_keys(texto or "", with_spaces=True, pause=0.03)


def _parse_fraccion_env(nombre_env: str, defecto: tuple[float, float]) -> tuple[float, float]:
    raw = (os.environ.get(nombre_env) or "").strip()
    if "," in raw:
        a, b = raw.split(",", 1)
        try:
            return max(0.0, min(1.0, float(a.strip()))), max(0.0, min(1.0, float(b.strip())))
        except ValueError:
            pass
    return defecto


def _normalizar_cedula_nombre_pdf(cedula: str) -> str:
    """Solo dígitos/letras del Excel columna A (pacientes); nunca motivo columna E."""
    return re.sub(r"[^\w\-.]", "", (cedula or "").strip())


def _establecer_portapapeles(texto: str) -> bool:
    try:
        import win32clipboard

        win32clipboard.OpenClipboard()
        try:
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32clipboard.CF_UNICODETEXT, texto or "")
        finally:
            win32clipboard.CloseClipboard()
        return True
    except Exception:
        return False


def _pegar_texto_portapapeles(texto: str) -> None:
    """Pega con Ctrl+V (texto largo, tildes y dos puntos; más fiable que teclear)."""
    try:
        import win32clipboard

        win32clipboard.OpenClipboard()
        try:
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32clipboard.CF_UNICODETEXT, texto or "")
        finally:
            win32clipboard.CloseClipboard()
    except Exception as ex:
        print(f"Flujo: aviso portapapeles ({ex}); se escribe con teclado.")
        _type_texto_campo(texto)
        return
    time.sleep(0.07)
    keyboard.send_keys("^a")
    time.sleep(0.06)
    keyboard.send_keys("^v")
    time.sleep(0.14)


def _leer_texto_portapapeles() -> str:
    try:
        import win32clipboard

        win32clipboard.OpenClipboard()
        try:
            data = win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
        finally:
            win32clipboard.CloseClipboard()
        return (data or "").strip()
    except Exception:
        return ""


def _plantilla_es_observacion(nombre_plantilla: str) -> bool:
    low = (nombre_plantilla or "").lower()
    return bool(re.match(r"^3\s+", (nombre_plantilla or "").strip())) and any(
        x in low for x in ("observacion", "observación", "comentario")
    )


def _resolver_plantilla_observacion() -> Path:
    num_map = _plantillas_png_por_numero()
    if "3" in num_map:
        return _resolver_ruta_plantilla_png(num_map["3"])
    if TEMPLATE_DIR.is_dir():
        for f in sorted(TEMPLATE_DIR.glob("3 *.png")):
            if f.is_file() and _plantilla_es_observacion(f.name):
                return f.resolve()
    raise RuntimeError(
        "No hay plantilla «3 …observación….png» en templates/. "
        "Coloque el recorte del campo Observación."
    )


def _clic_y_escribir_observacion(hwnd: int, texto: str) -> None:
    """Clic en plantilla 3 (caja de texto) y pega columna F del catálogo."""
    texto = (texto or "").strip()
    if not texto:
        print("Flujo: observación vacía (columna F del catálogo); no se escribe.")
        return
    tpl = _resolver_plantilla_observacion()
    nombre = tpl.name
    roi_def = (0.02, 0.02, 0.98, 0.98)
    rois_list = _rois_ampliados_plantilla_8_paciente(nombre, roi_def, roi_def)
    min_clic = _float_env("INDIGO_MATCH_MIN_PLANTILLA_3", 0.38)
    min_clic_fb = max(0.22, min_clic - 0.12)
    cy_min = _float_env("INDIGO_OBSERVACION_MATCH_CY_MIN", 0.11)
    orden = (os.environ.get("INDIGO_OBSERVACION_MATCH_ORDEN") or "top").strip().lower()
    fraccion = _parse_fraccion_env("INDIGO_OBSERVACION_CLICK_FRACCION_CAJA", (0.72, 0.5))
    print(
        f"Flujo: observación (columna F) → {texto[:90]!r}{'…' if len(texto) > 90 else ''} "
        f"(umbral ≥{min_clic:.2f}; fallback pantalla completa hasta {min_clic_fb:.2f})"
    )
    clic_plantilla_en_hwnd(
        hwnd,
        tpl,
        rois_list,
        orden if orden in ("left", "right", "top") else "top",
        min_clic,
        "IndiGO observación (caja)",
        pantalla_completa_si_falla=True,
        umbrales_pantalla_completa=(
            min_clic_fb,
            max(0.20, min_clic_fb - 0.04),
            max(0.18, min_clic_fb - 0.06),
            max(0.16, min_clic_fb - 0.08),
        ),
        fraccion_click_en_plantilla=fraccion,
        cy_min_frac=cy_min,
    )
    time.sleep(_float_env("INDIGO_OBSERVACION_TRAS_FOCUS_SEG", 0.32))
    if _bool_env("INDIGO_OBSERVACION_USAR_PORTAPAPELES", True):
        _pegar_texto_portapapeles(texto)
    else:
        _type_texto_campo(texto)


def _plantilla_es_motivo_consulta(nombre_plantilla: str) -> bool:
    low = (nombre_plantilla or "").lower()
    return bool(re.match(r"^2\s+", (nombre_plantilla or "").strip())) and any(
        x in low for x in ("motivo", "select", "consulta")
    )


def _fraccion_click_combo_motivo_flecha() -> tuple[float, float]:
    """Flecha del desplegable (puede abrir/cerrar en toggle; usar solo con INDIGO_MOTIVO_ABRIR=flecha)."""
    raw = (os.environ.get("INDIGO_MOTIVO_CLICK_FRACCION_FLECHA") or os.environ.get("INDIGO_MOTIVO_CLICK_FRACCION") or "0.93,0.5").strip()
    if "," in raw:
        a, b = raw.split(",", 1)
        try:
            return max(0.0, min(1.0, float(a.strip()))), max(0.0, min(1.0, float(b.strip())))
        except ValueError:
            pass
    return 0.93, 0.5


def _fraccion_click_combo_motivo_caja() -> tuple[float, float]:
    """Centro de la caja blanca del combo (foco + F4; evita doble toggle en la flecha)."""
    raw = (os.environ.get("INDIGO_MOTIVO_CLICK_FRACCION_CAJA") or "0.78,0.5").strip()
    if "," in raw:
        a, b = raw.split(",", 1)
        try:
            return max(0.0, min(1.0, float(a.strip()))), max(0.0, min(1.0, float(b.strip())))
        except ValueError:
            pass
    return 0.78, 0.5


def _resolver_plantilla_motivo_consulta() -> Path:
    num_map = _plantillas_png_por_numero()
    if "2" in num_map:
        return _resolver_ruta_plantilla_png(num_map["2"])
    if TEMPLATE_DIR.is_dir():
        for f in sorted(TEMPLATE_DIR.glob("2 *.png")):
            if _plantilla_es_motivo_consulta(f.name):
                return f.resolve()
    raise RuntimeError(
        "No hay plantilla «2 …motivo….png» en templates/. "
        "Coloque el recorte de «Motivo consulta» o defina el PNG en el flujo."
    )


def _clic_abrir_motivo_consulta(hwnd: int) -> None:
    """Enfoca el combo Motivo consulta y abre la lista (caja + F4 por defecto; sin clic en flecha)."""
    tpl = _resolver_plantilla_motivo_consulta()
    nombre = tpl.name
    roi_def = (0.02, 0.02, 0.98, 0.98)
    rois_list = _rois_ampliados_plantilla_8_paciente(nombre, roi_def, roi_def)
    min_clic = _float_env("INDIGO_MATCH_MIN_PLANTILLA_2", 0.42)
    # focus_alt (defecto): clic en caja + Alt+Abajo; suele ir mejor que F4/flecha en Vie/WebView.
    modo = (os.environ.get("INDIGO_MOTIVO_ABRIR") or "focus_alt").strip().lower()
    cy_max = _float_env(
        "INDIGO_MOTIVO_MATCH_CY_MAX",
        _float_env("INDIGO_MOTIVO_MATCH_CY_MAX_FRAC", 0.40),
    )
    orden_match = (os.environ.get("INDIGO_MOTIVO_MATCH_ORDEN") or "top").strip().lower()

    print(f"Flujo: abrir motivo consulta ({modo}; plantilla {tpl.name!r}; orden match {orden_match!r})")

    if modo in ("flecha", "arrow", "arrow_only"):
        clic_plantilla_en_hwnd(
            hwnd,
            tpl,
            rois_list,
            "right",
            min_clic,
            "IndiGO motivo (flecha)",
            pantalla_completa_si_falla=False,
            fraccion_click_en_plantilla=_fraccion_click_combo_motivo_flecha(),
            boton="left",
            cy_max_frac=cy_max,
        )
        time.sleep(_float_env("INDIGO_MOTIVO_TRAS_ABRIR_SEG", 0.45))
        return

    clic_plantilla_en_hwnd(
        hwnd,
        tpl,
        rois_list,
        orden_match if orden_match in ("left", "right", "top") else "top",
        min_clic,
        "IndiGO motivo (caja)",
        pantalla_completa_si_falla=False,
        fraccion_click_en_plantilla=_fraccion_click_combo_motivo_caja(),
        boton="left",
        cy_max_frac=cy_max,
    )
    time.sleep(_float_env("INDIGO_MOTIVO_TRAS_FOCUS_SEG", 0.30))
    if modo not in ("solo_click", "click_only", "solo_foco"):
        if modo in ("focus_f4", "f4"):
            keyboard.send_keys("{F4}")
        else:
            keyboard.send_keys("%{DOWN}")
        time.sleep(_float_env("INDIGO_MOTIVO_TRAS_ABRIR_SEG", 0.55))


def _type_texto_combo_sin_borrar(texto: str) -> None:
    """Escribe en un desplegable abierto sin Ctrl+A (evita que se cierre la lista)."""
    keyboard.send_keys(texto or "", with_spaces=True, pause=0.05)


def _seleccionar_motivo_consulta(sesion: SesionIndigo) -> None:
    """Tras abrir el combo: tecleo (E) o ↓ (D). Sin Ctrl+A ni HOME (suelen cerrar la lista o saltar a observación)."""
    texto = (sesion.motivo_consulta or "").strip()
    idx = sesion.motivo_opcion_numero
    prefer_d = _bool_env("INDIGO_MOTIVO_PREFERIR_COLUMNA_D", True)
    pausa_down = _float_env("INDIGO_MOTIVO_DOWN_PAUSE", 0.14)
    try:
        umbral_tecleo_d = int((os.environ.get("INDIGO_MOTIVO_TECLEO_SI_D_MAYOR_QUE") or "6").strip())
    except ValueError:
        umbral_tecleo_d = 6
    try:
        max_chars = int((os.environ.get("INDIGO_MOTIVO_TECLEO_MAX_CHARS") or "16").strip())
    except ValueError:
        max_chars = 16
    max_chars = max(4, min(40, max_chars))

    solo_down = _bool_env("INDIGO_MOTIVO_SOLO_DOWN_COLUMNA_D", False)
    usar_tecleo = bool(texto) and (
        not prefer_d or idx is None or idx > umbral_tecleo_d or not solo_down
    )
    usar_d = prefer_d and idx is not None and idx >= 1 and not usar_tecleo

    time.sleep(_float_env("INDIGO_MOTIVO_TRAS_ABRIR_ANTES_SELECCION", 0.20))

    if usar_tecleo:
        prefijo = texto[:max_chars]
        print(f"Flujo: motivo consulta (columna E, tecleo parcial) → {prefijo!r}")
        _type_texto_combo_sin_borrar(prefijo)
        time.sleep(_float_env("INDIGO_MOTIVO_TRAS_TECLEO_SEG", 0.45))
    elif usar_d:
        nd = max(0, idx - 1)
        print(f"Flujo: motivo consulta (columna D = opción {idx}, «Abajo» × {nd})")
        if _bool_env("INDIGO_MOTIVO_HOME_ANTES_DOWN", False):
            keyboard.send_keys("{HOME}")
            time.sleep(0.10)
        for _ in range(nd):
            keyboard.send_keys("{DOWN}")
            time.sleep(pausa_down)
    elif texto:
        prefijo = texto[:max_chars]
        print(f"Flujo: motivo consulta (solo E) → {prefijo!r}")
        _type_texto_combo_sin_borrar(prefijo)
        time.sleep(_float_env("INDIGO_MOTIVO_TRAS_TECLEO_SEG", 0.35))
    elif idx is not None and idx >= 1:
        nd = max(0, idx - 1)
        print(f"Flujo: motivo consulta (columna D = opción {idx}, «Abajo» × {nd})")
        for _ in range(nd):
            keyboard.send_keys("{DOWN}")
            time.sleep(pausa_down)
    else:
        raise RuntimeError(
            "Catálogo E.S.E.: rellene la columna E (texto del motivo) o la D (número de opción del desplegable)."
        )

    conf = (os.environ.get("INDIGO_MOTIVO_CONFIRMACION") or "enter").strip().lower()
    if _bool_env("INDIGO_MOTIVO_SIN_CONFIRMAR", False):
        conf = "none"
    if conf not in ("none", "no", "off", "0"):
        time.sleep(_float_env("INDIGO_MOTIVO_ANTES_CONFIRMAR_SEG", 0.18))
        if conf in ("space", "espacio"):
            keyboard.send_keys(" ")
        else:
            keyboard.send_keys("{ENTER}")
        time.sleep(_float_env("INDIGO_MOTIVO_TRAS_CONFIRMACION_SEG", 0.30))


def _plantilla_es_aceptar_tras_formulario(nombre_plantilla: str) -> bool:
    """Botón Aceptar tras motivo/observación (p. ej. plantilla «4 click en aceptar»)."""
    low = (nombre_plantilla or "").lower()
    if not re.match(r"^4\s+", (nombre_plantilla or "").strip()):
        return False
    return "aceptar" in low


def _plantilla_es_aceptar_entrada_ese(nombre_plantilla: str) -> bool:
    """Tras este clic el diálogo de empresa cierra y suele cambiar el HWND de trabajo."""
    low = (nombre_plantilla or "").lower()
    if "aceptar" not in low:
        return False
    return "ingresar" in low or "ingreso" in low or ("ese" in low and "ingres" in low)


def _plantilla_es_listado_folios(nombre_plantilla: str) -> bool:
    """Listado de folios tras Aceptar (plantilla 5): clic derecho en la primera fila."""
    low = (nombre_plantilla or "").lower()
    if not re.match(r"^5\s+", (nombre_plantilla or "").strip()):
        return False
    return any(
        x in low
        for x in ("folio", "folios", "listado", "derecho", "click derecho", "click_derecho")
    )


def _plantilla_es_opcion_menu_contexto_folios(nombre_plantilla: str) -> bool:
    """Opción del menú contextual tras clic derecho en folios (p. ej. plantilla 6 Consultar/Imprimir…)."""
    low = (nombre_plantilla or "").lower()
    if not re.match(r"^6\s+", (nombre_plantilla or "").strip()):
        return False
    return any(
        x in low
        for x in ("consultar", "imprimir", "filtro", "reg", "medico", "médico", "medicos", "médicos")
    )


def _tpl_bgr_recorte_izquierdo(tpl_bgr: np.ndarray, frac_ancho: float) -> np.ndarray:
    """Recorta iconos/expand de la fila (sin fechas ni diagnósticos que cambian por paciente)."""
    w = tpl_bgr.shape[1]
    cut = max(48, min(w - 8, int(round(w * frac_ancho))))
    return tpl_bgr[:, :cut].copy()


def _tpl_bgr_recorte_encabezado_grid(tpl_bgr: np.ndarray, frac_alto: float) -> np.ndarray:
    """Solo la franja de títulos de columna (Fecha, Diagnóstico…); estable entre pacientes y DPI cercanos."""
    h = tpl_bgr.shape[0]
    cut = max(14, min(h - 6, int(round(h * frac_alto))))
    return tpl_bgr[:cut, :].copy()


def _escalas_template_folios() -> tuple[float, ...]:
    raw = (os.environ.get("INDIGO_FOLIOS_TEMPLATE_SCALES") or "").strip().replace(",", " ")
    if raw:
        out: list[float] = []
        for p in raw.split():
            try:
                out.append(float(p.replace(",", ".")))
            except ValueError:
                pass
        if out:
            return tuple(out)
    return (0.62, 0.70, 0.78, 0.86, 0.92, 1.0, 1.08, 1.16, 1.24, 1.32)


def _variantes_tpl_listado_folios(tpl_full: np.ndarray) -> list[tuple[str, np.ndarray]]:
    frac_izq = _float_env("INDIGO_FOLIOS_RECORTE_IZQ_FRAC", 0.42)
    frac_hdr = _float_env("INDIGO_FOLIOS_RECORTE_ENCABEZADO_FRAC", 0.28)
    out: list[tuple[str, np.ndarray]] = [
        ("encabezado columnas", _tpl_bgr_recorte_encabezado_grid(tpl_full, frac_hdr)),
        (f"iconos izq. ({frac_izq:.0%})", _tpl_bgr_recorte_izquierdo(tpl_full, frac_izq)),
        ("completa", tpl_full),
    ]
    return out


def _max_similitud_listado_folios_detalle(
    hwnd: int,
    tpl_path: Path,
    rois: list[tuple[float, float, float, float]],
) -> tuple[float, str]:
    """Mayor similitud y nombre de variante (encabezado / iconos / completa)."""
    tpl_full = _cv2_imread_path(tpl_path)
    if tpl_full is None or tpl_full.size == 0:
        return 0.0, "sin plantilla"
    if tpl_full.ndim == 3 and tpl_full.shape[2] == 4:
        tpl_full = cv2.cvtColor(tpl_full, cv2.COLOR_BGRA2BGR)
    hay, _, _ = _origen_y_captura_para_match_hwnd(hwnd, "")
    escalas = _escalas_template_folios()
    mejor = 0.0
    mejor_etq = "ninguna"
    for etq, tpl_var in _variantes_tpl_listado_folios(tpl_full):
        for roi_rel in rois:
            hit = _mejor_match_multiescala(
                hay,
                tpl_var,
                roi_rel,
                "top",
                0.06,
                cy_min_frac=_float_env("INDIGO_FOLIOS_MATCH_CY_MIN", 0.14),
                cy_max_frac=_float_env("INDIGO_FOLIOS_MATCH_CY_MAX", 0.78),
                escalas=escalas,
            )
            if hit is not None and float(hit[2]) > mejor:
                mejor = float(hit[2])
                mejor_etq = etq
    return mejor, mejor_etq


def _max_similitud_listado_folios(
    hwnd: int,
    tpl_path: Path,
    rois: list[tuple[float, float, float, float]],
) -> float:
    s, _ = _max_similitud_listado_folios_detalle(hwnd, tpl_path, rois)
    return s


def _esperar_listado_folios(
    hwnd: int,
    tpl_path: Path,
    rois_list: list[tuple[float, float, float, float]],
    timeout_sec: float,
    umbral_ok: float,
    need_ok: int,
    nombre: str,
) -> None:
    """Espera el listado de folios. Si hay calibración guardada, no bloquea 60s por PNG que no matchea en otro PC."""
    cal = _cargar_calibracion_folios()
    path_cal = _ruta_calibracion_folios()

    if cal and _bool_env("INDIGO_FOLIOS_SALTAR_WAIT_SI_CALIBRACION", True):
        pausa = _float_env("INDIGO_FOLIOS_PAUSA_TRAS_LISTADO_SEG", 1.2)
        fx, fy = float(cal["frac_x"]), float(cal["frac_y"])
        print(
            f"Flujo: listado folios — calibración activa ({path_cal.name}); "
            f"no se exige coincidencia del PNG 5 (pausa {pausa:.1f}s). "
            f"Siguiente: clic derecho en ({fx:.3f}, {fy:.3f})."
        )
        restaurar_y_traer_al_frente(hwnd)
        time.sleep(pausa)
        return

    raw_fc = (os.environ.get("INDIGO_FOLIOS_WAIT_CONSECUTIVE") or "").strip()
    try:
        need_folios = int(raw_fc) if raw_fc else min(need_ok, 2)
    except ValueError:
        need_folios = min(need_ok, 2)
    need_folios = max(1, min(need_folios, 25))

    deadline = time.time() + timeout_sec
    ultima_s = 0.0
    ultima_var = ""
    racha = 0
    ultimo_info = time.time()
    print(
        f"Flujo: esperar listado folios {nombre!r} (máx {timeout_sec:.0f}s, umbral PNG ≥{umbral_ok:.2f}, "
        f"{need_folios} lecturas; encabezado/iconos/completa). "
        f"Si el listado ya está visible y no avanza: --calibrar-folios o "
        f"INDIGO_FOLIOS_SALTAR_WAIT_SI_CALIBRACION=1 con {path_cal.name}."
    )
    while time.time() < deadline:
        restaurar_y_traer_al_frente(hwnd)
        time.sleep(0.06)
        ultima_s, ultima_var = _max_similitud_listado_folios_detalle(hwnd, tpl_path, rois_list)
        if time.time() - ultimo_info >= 4.0:
            print(
                f"Flujo: … esperando listado folios (similitud {ultima_s:.2f} [{ultima_var}]; "
                f"umbral {umbral_ok:.2f}; racha {racha}/{need_folios}; "
                f"quedan {max(0, int(deadline - time.time()))}s)"
            )
            ultimo_info = time.time()
        if ultima_s >= umbral_ok:
            racha += 1
            if racha >= need_folios:
                print(
                    f"Flujo: listado folios detectado por PNG ({need_folios}× ≥{umbral_ok:.2f}; "
                    f"última {ultima_s:.2f} [{ultima_var}])"
                )
                return
        else:
            racha = 0
        time.sleep(0.32)
    if cal:
        print(
            f"Flujo: aviso — no hubo match PNG 5 en {timeout_sec:.0f}s (última {ultima_s:.2f} [{ultima_var}]); "
            f"se continúa con clic derecho por calibración ({path_cal.name})."
        )
        time.sleep(_float_env("INDIGO_FOLIOS_PAUSA_TRAS_LISTADO_SEG", 0.8))
        return
    raise TimeoutError(
        f"No apareció el listado de folios ({nombre!r}) en {timeout_sec:.0f}s "
        f"(última similitud {ultima_s:.2f} [{ultima_var}]; umbral {umbral_ok:.2f}). "
        "En otro PC/DPI use: py run_indigo_historias.py --calibrar-folios y deje "
        "INDIGO_FOLIOS_SALTAR_WAIT_SI_CALIBRACION=1, o comente wait_template 5 en el flujo."
    )


def _ruta_calibracion_folios() -> Path:
    raw = (os.environ.get("INDIGO_FOLIOS_CALIBRACION_JSON") or "").strip().strip('"')
    if raw:
        p = Path(raw)
        return p.expanduser().resolve() if p.is_absolute() else (ROOT / raw).resolve()
    return CALIBRACION_FOLIOS_DEFAULT.resolve()


def _cargar_calibracion_folios() -> dict | None:
    path = _ruta_calibracion_folios()
    if not path.is_file():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return None
    try:
        fx = float(data.get("frac_x"))
        fy = float(data.get("frac_y"))
    except (TypeError, ValueError):
        return None
    if not (0.0 <= fx <= 1.0 and 0.0 <= fy <= 1.0):
        return None
    data["frac_x"] = fx
    data["frac_y"] = fy
    return data


def _guardar_calibracion_folios(hwnd: int, screen_x: int, screen_y: int) -> Path:
    left, top, right, bottom = _rect_actual(hwnd)
    w, h = right - left, bottom - top
    fx = (screen_x - left) / w if w > 0 else 0.32
    fy = (screen_y - top) / h if h > 0 else 0.40
    fx = max(0.02, min(0.98, fx))
    fy = max(0.02, min(0.98, fy))
    path = _ruta_calibracion_folios()
    data = {
        "frac_x": round(fx, 6),
        "frac_y": round(fy, 6),
        "screen_x": int(screen_x),
        "screen_y": int(screen_y),
        "hwnd_al_guardar": int(hwnd),
        "ventana_titulo": (win32gui.GetWindowText(hwnd) or "").strip(),
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "nota": "Clic derecho listado folios (paso 5). Regenerar si cambia resolución o escala Windows.",
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    return path


def _coords_pantalla_clic_folios_calibrado(hwnd: int) -> tuple[int, int, float, float] | None:
    data = _cargar_calibracion_folios()
    if not data:
        return None
    left, top, right, bottom = _rect_actual(hwnd)
    w, h = right - left, bottom - top
    if w < 80 or h < 80:
        return None
    fx, fy = float(data["frac_x"]), float(data["frac_y"])
    cx = left + int(round(w * fx))
    cy = top + int(round(h * fy))
    cx, cy = _clamp_xy_evita_esquina_failsafe_pantalla(cx, cy)
    return cx, cy, fx, fy


def _clic_derecho_folios_calibrado(hwnd: int, *, etiqueta: str = "IndiGO folio 5 (calibración)") -> bool:
    """True si hubo archivo de calibración y se ejecutó el clic."""
    path = _ruta_calibracion_folios()
    if _bool_env("INDIGO_FOLIOS_IGNORAR_CALIBRACION", False):
        print(f"Flujo: calibración folios ignorada (INDIGO_FOLIOS_IGNORAR_CALIBRACION=1).")
        return False
    if not path.is_file():
        print(f"Flujo: no hay calibración folios en:\n  {path}\n  Ejecute: py run_indigo_historias.py --calibrar-folios")
        return False
    hit = _coords_pantalla_clic_folios_calibrado(hwnd)
    if hit is None:
        print(f"Flujo: calibración folios inválida o ventana IndiGO demasiado pequeña ({path}).")
        return False
    cx, cy, fx, fy = hit
    print(
        f"Flujo: {etiqueta} — moviendo a ({cx},{cy}) "
        f"[frac {fx:.4f},{fy:.4f}; archivo {path}]"
    )
    if not _clic_derecho_en_coordenadas_folio(hwnd, cx, cy, etiqueta):
        print(
            "Flujo: el clic derecho en folios no se aplicó (ratón ni mensaje HWND). "
            "Vea el diagnóstico arriba (AnyDesk solo lectura, calibración o ventana IndiGO)."
        )
        return False
    _programar_sin_traer_indigo_tras_menu_folios()
    return True


def _capturar_mouse_cuenta_atras(segundos: int | None = None) -> tuple[int, int]:
    """Cuenta atrás sin pulsar Enter (evita clic accidental en la consola)."""
    if segundos is None:
        segundos = int(_float_env("INDIGO_CALIB_FOLIOS_CUENTA_ATRAS", 6))
    segundos = max(3, min(20, segundos))
    print(
        "Coloque el mouse sobre la 1.ª fila del listado de folios.\n"
        "NO haga clic en la consola (guardaría mal el punto). Use Alt+Tab para cambiar de ventana.\n"
        f"Captura automática en {segundos} segundos…\n"
    )
    for n in range(segundos, 0, -1):
        mx, my = pyautogui.position()
        print(f"  {n:2d}s  mouse=({mx:4d}, {my:4d})     ", end="\r", flush=True)
        time.sleep(1.0)
    mx, my = pyautogui.position()
    print(f"\nCapturado: ({mx}, {my})")
    return int(mx), int(my)


def calibrar_clic_derecho_folios() -> None:
    """Asistente: guarda la posición del mouse como clic derecho en la 1.ª fila del listado."""
    print("\n=== Calibración — clic derecho en listado de folios (paso 5) ===\n")
    print(
        "1. Abra Vie/IndiGO con el listado de folios visible (Fecha historia, Diagnóstico…).\n"
        "2. Coloque el mouse sobre la 1.ª fila (donde haría clic derecho).\n"
    )
    modo = (os.environ.get("INDIGO_CALIB_FOLIOS_MODO") or "cuenta_atras").strip().lower()
    if modo in ("enter", "manual"):
        print("3. Sin hacer clic en la consola, pulse Enter (si hizo clic aquí, las coords. salen mal).\n")
        try:
            input("Enter cuando el mouse esté en el punto… ")
        except EOFError:
            raise SystemExit("Entrada cancelada.")
        mx, my = pyautogui.position()
    else:
        mx, my = _capturar_mouse_cuenta_atras()
    hwnd = encontrar_hwnd_indigo()
    restaurar_y_traer_al_frente(hwnd)
    time.sleep(0.25)
    path = _guardar_calibracion_folios(hwnd, int(mx), int(my))
    data = _cargar_calibracion_folios() or {}
    print(f"\nGuardado en:\n  {path}\n")
    print(f"  frac_x = {data.get('frac_x')}")
    print(f"  frac_y = {data.get('frac_y')}")
    if _bool_env("INDIGO_CALIB_FOLIOS_PROBAR_MOVIMIENTO", True):
        hit = _coords_pantalla_clic_folios_calibrado(hwnd)
        if hit:
            cx, cy, _, _ = hit
            print(f"\nPrueba de movimiento (sin clic) hacia ({cx},{cy})…")
            ax, ay, ok = _mover_cursor_a_pantalla(cx, cy)
            print(f"Mouse ahora en ({ax},{ay}). ¿Coincide con la 1.ª fila? movimiento_ok={ok}")
    print("\nPrueba clic derecho:\n  py run_indigo_historias.py --probar-folios-clic\n")
    print("El flujo normal usará este archivo antes que la plantilla PNG 5.\n")


def probar_clic_derecho_folios_calibrado() -> None:
    """Un solo clic derecho con las coordenadas guardadas (prueba rápida)."""
    if not _cargar_calibracion_folios():
        raise SystemExit(
            f"No existe calibración en {_ruta_calibracion_folios()}. "
            "Ejecute primero: py run_indigo_historias.py --calibrar-folios"
        )
    hwnd = encontrar_hwnd_indigo()
    restaurar_y_traer_al_frente(hwnd)
    time.sleep(0.3)
    if not _clic_derecho_folios_calibrado(hwnd, etiqueta="Prueba clic derecho folios"):
        raise SystemExit(
            "No se pudo aplicar la calibración (archivo inválido, ventana pequeña o "
            "el ratón no se movió — frecuente con AnyDesk en solo lectura)."
        )
    print(
        "Listo. Si abrió el menú contextual (aunque el ratón no se haya movido vía AnyDesk), "
        "la calibración es correcta."
    )


def _clic_derecho_pantalla_coords(hwnd: int, cx_img: int, cy_img: int, etiqueta: str) -> None:
    _pantalla, left, top = _origen_y_captura_para_match_hwnd(hwnd, etiqueta)
    cx = left + int(cx_img)
    cy = top + int(cy_img)
    cx, cy = _clamp_xy_evita_esquina_failsafe_pantalla(cx, cy)
    print(f"{etiqueta}: clic derecho → ({cx},{cy})")
    if not _clic_derecho_en_coordenadas_folio(hwnd, cx, cy, etiqueta):
        raise RuntimeError(f"{etiqueta}: no se pudo hacer clic derecho en ({cx},{cy})")


def _clic_derecho_fraccion_listado_folios(hwnd: int) -> None:
    """Clic derecho en zona del grid (1.ª fila) sin depender del texto del paciente."""
    if _clic_derecho_folios_calibrado(hwnd):
        return
    left, top, right, bottom = _rect_actual(hwnd)
    w, h = right - left, bottom - top
    if w < 80 or h < 80:
        raise RuntimeError("Ventana IndiGO demasiado pequeña para clic fracción en listado de folios.")
    fx = _float_env("INDIGO_FOLIOS_CLICK_FRACC_X", 0.32)
    fy = _float_env("INDIGO_FOLIOS_CLICK_FRACC_Y", 0.40)
    cx = left + int(w * fx)
    cy = top + int(h * fy)
    cx, cy = _clamp_xy_evita_esquina_failsafe_pantalla(cx, cy)
    print(
        f"Flujo: clic derecho listado folios por posición relativa "
        f"({fx:.2f}, {fy:.2f}) → ({cx},{cy}) [INDIGO_FOLIOS_CLICK_FRACC_X/Y]"
    )
    if not _clic_derecho_en_coordenadas_folio(
        hwnd, cx, cy, "IndiGO folio 5 (fracción listado)"
    ):
        raise RuntimeError(
            "No se pudo hacer clic derecho por fracción en el listado de folios "
            f"({fx:.2f}, {fy:.2f}). Calibre con: py run_indigo_historias.py --calibrar-folios"
        )


def _preparar_hwnd_antes_folios(hwnd: int) -> int:
    """Tras Aceptar (plantilla 4) la ventana puede cambiar; re-ancla antes del clic derecho."""
    pausa = _float_env("INDIGO_FOLIOS_PRECLICK_SEG", 0.45)
    restaurar_y_traer_al_frente(hwnd)
    time.sleep(pausa)
    try:
        hwnd_n = encontrar_hwnd_indigo(timeout=12.0)
        restaurar_y_traer_al_frente(hwnd_n)
        tit = (win32gui.GetWindowText(hwnd_n) or "").strip()
        print(f"Flujo: paso 5 — ventana IndiGO activa {tit!r} (hwnd={hwnd_n})")
        return hwnd_n
    except TimeoutError as ex:
        print(f"Flujo: paso 5 — aviso al re-localizar IndiGO: {ex}; se usa hwnd={hwnd}.")
        return hwnd


def _ejecutar_click_folios(hwnd: int) -> None:
    """Paso 5: clic derecho en la 1.ª fila (calibración → HWND → fracción; sin depender del PNG 5)."""
    hwnd = _preparar_hwnd_antes_folios(hwnd)
    path_cal = _ruta_calibracion_folios()
    tiene_cal = (
        _cargar_calibracion_folios() is not None
        and path_cal.is_file()
        and not _bool_env("INDIGO_FOLIOS_IGNORAR_CALIBRACION", False)
    )
    if tiene_cal:
        print(f"Flujo: click_folios — usando {path_cal.name}")
        if _clic_derecho_folios_calibrado(hwnd):
            _programar_sin_traer_indigo_tras_menu_folios()
            return
        if _bool_env("INDIGO_FOLIOS_SOLO_CALIBRACION", True):
            raise RuntimeError(
                f"Calibración en {path_cal} pero el clic derecho falló. "
                "Regenere con --calibrar-folios en la misma sesión/DPI, ejecute en el PC local "
                "(no AnyDesk solo lectura) o pruebe INDIGO_FOLIOS_CLIC_HWND=1 (defecto)."
            )
    print(
        "Flujo: click_folios — sin calibración; clic por fracción "
        "(INDIGO_FOLIOS_CLICK_FRACC_X/Y o --calibrar-folios)."
    )
    _clic_derecho_fraccion_listado_folios(hwnd)
    _programar_sin_traer_indigo_tras_menu_folios()


def _clic_opcion_menu_contexto_folios(
    nombre_tpl: str,
    roi: tuple[float, float, float, float],
    roi_def: tuple[float, float, float, float],
) -> None:
    """
    Clic en la opción del menú contextual (plantilla 6) sin traer IndiGO al frente.
    Busca el PNG en el monitor completo y hace clic en la ventana de menú (#32768).
    """
    tpl_path = _resolver_ruta_plantilla_png(nombre_tpl)
    tpl = _cv2_imread_path(tpl_path)
    if tpl is None or tpl.size == 0:
        raise RuntimeError(f"No se pudo leer la plantilla {tpl_path}")
    if tpl.ndim == 3 and tpl.shape[2] == 4:
        tpl = cv2.cvtColor(tpl, cv2.COLOR_BGRA2BGR)
    rois_list = _rois_ampliados_plantilla_8_paciente(nombre_tpl, roi, roi_def)
    timeout = _float_env("INDIGO_MENU_FOLIO_TIMEOUT", 12.0)
    pausa_ini = _float_env("INDIGO_MENU_FOLIO_ESPERA_INICIAL_SEG", 0.4)
    min_sc = _float_env("INDIGO_MENU_FOLIO_MATCH_MIN", 0.34)
    umbrales = (min_sc, min_sc - 0.06, min_sc - 0.10, 0.28, 0.24, 0.20)
    time.sleep(max(0.0, pausa_ini))
    print(
        f"Flujo: menú contextual folios — buscando {tpl_path.name!r} en pantalla completa "
        f"(sin traer IndiGO; máx {timeout:.0f}s; umbral ≥{min_sc:.2f})"
    )
    deadline = time.time() + timeout
    ultima_s = 0.0
    ultimo_info = time.time()
    while time.time() < deadline:
        hay, left, top = _captura_monitor_principal_bgr()
        hit = None
        for thr in umbrales:
            if thr < 0.14:
                break
            for roi_rel in rois_list:
                hit = _mejor_match_multiescala(
                    hay, tpl, roi_rel, "top", thr, escalas=(0.88, 0.94, 1.0, 1.06, 1.12)
                )
                if hit is None:
                    hit = _mejor_match_multiescala(
                        hay, tpl, (0.0, 0.0, 1.0, 1.0), "top", thr, escalas=(0.88, 0.94, 1.0, 1.06, 1.12)
                    )
                if hit is not None:
                    break
            if hit is not None:
                break
        if hit is not None:
            cx_img, cy_img, sc, tw, th = hit
            fxp = _float_env("INDIGO_MENU_FOLIO_FRACCION_CLICK_X", 0.5)
            fyp = _float_env("INDIGO_MENU_FOLIO_FRACCION_CLICK_Y", 0.5)
            tl_x = int(round(cx_img - tw / 2.0))
            tl_y = int(round(cy_img - th / 2.0))
            ox = max(0, min(tw - 1, int(round((tw - 1) * fxp))))
            oy = max(0, min(th - 1, int(round((th - 1) * fyp))))
            sx = left + tl_x + ox
            sy = top + tl_y + oy
            print(
                f"Flujo: menú contextual — match {sc:.2f} → clic en ({sx},{sy}) "
                f"[plantilla {tpl_path.name}]"
            )
            _clic_izquierdo_menu_contexto_en_pantalla(
                sx, sy, f"IndiGO menú folio ({tpl_path.name})"
            )
            time.sleep(_float_env("INDIGO_SLEEP_TRAS_MENU_FOLIO_SEG", 0.35))
            return
        for roi_rel in rois_list:
            h2 = _mejor_match_multiescala(
                hay, tpl, roi_rel, "top", 0.12, escalas=(0.88, 0.94, 1.0, 1.06, 1.12)
            )
            if h2 is not None:
                ultima_s = max(ultima_s, h2[2])
        if time.time() - ultimo_info >= 2.0:
            print(
                f"Flujo: … esperando menú contextual (mejor similitud {ultima_s:.2f}; "
                f"quedan {max(0, int(deadline - time.time()))}s)"
            )
            ultimo_info = time.time()
        time.sleep(0.14)
    if _bool_env("INDIGO_MENU_FOLIO_TECLADO_FALLBACK", True):
        print(
            "Flujo: menú contextual — sin match PNG; se envía Enter (primera opción del menú). "
            "Ajuste INDIGO_MENU_FOLIO_MATCH_MIN o el PNG 6 si no es la 1.ª opción."
        )
        keyboard.send_keys("{ENTER}")
        time.sleep(_float_env("INDIGO_SLEEP_TRAS_MENU_FOLIO_SEG", 0.35))
        return
    raise RuntimeError(
        f"No se encontró la opción del menú contextual ({nombre_tpl!r}) en {timeout:.0f}s "
        f"(última similitud {ultima_s:.2f}; umbral {min_sc:.2f}). "
        "Suba INDIGO_MENU_FOLIO_ESPERA_INICIAL_SEG tras click_folios o recorte el PNG 6."
    )


def _clic_derecho_primer_folio(
    hwnd: int,
    nombre_tpl: str,
    roi: tuple[float, float, float, float],
    roi_def: tuple[float, float, float, float],
) -> None:
    """Clic derecho en la 1.ª fila: calibración guardada, luego plantilla, luego fracción por defecto."""
    path_cal = _ruta_calibracion_folios()
    tiene_cal = (
        _cargar_calibracion_folios() is not None
        and path_cal.is_file()
        and not _bool_env("INDIGO_FOLIOS_IGNORAR_CALIBRACION", False)
    )
    solo_cal = _bool_env("INDIGO_FOLIOS_SOLO_CALIBRACION", tiene_cal)
    hwnd = _preparar_hwnd_antes_folios(hwnd)
    if tiene_cal and (solo_cal or _bool_env("INDIGO_FOLIOS_PREFERIR_CALIBRACION", True)):
        print(f"Flujo: paso 5 — calibración en {path_cal}")
        if _clic_derecho_folios_calibrado(hwnd):
            return
        if solo_cal:
            raise RuntimeError(
                f"Calibración en {path_cal} pero el clic derecho falló. "
                "Ejecute en sesión local, regenere --calibrar-folios o "
                "INDIGO_FOLIOS_PERMITIR_PLANTILLA_5=1 para intentar el PNG 5."
            )
        if not _bool_env("INDIGO_FOLIOS_PERMITIR_PLANTILLA_5", False):
            print("Flujo: calibración falló; se prueba fracción (sin plantilla PNG 5).")
            _clic_derecho_fraccion_listado_folios(hwnd)
            return
    elif solo_cal:
        raise RuntimeError(
            f"INDIGO_FOLIOS_SOLO_CALIBRACION=1 pero no hay {path_cal.name}. "
            "Ejecute: py run_indigo_historias.py --calibrar-folios"
        )

    tpl_path = _resolver_ruta_plantilla_png(nombre_tpl)
    tpl_full = _cv2_imread_path(tpl_path)
    if tpl_full is None or tpl_full.size == 0:
        raise RuntimeError(f"No se pudo leer la plantilla {tpl_path}")
    if tpl_full.ndim == 3 and tpl_full.shape[2] == 4:
        tpl_full = cv2.cvtColor(tpl_full, cv2.COLOR_BGRA2BGR)
    rois_list = _rois_ampliados_plantilla_8_paciente(nombre_tpl, roi, roi_def)
    min_sc = _float_env("INDIGO_MATCH_MIN_PLANTILLA_5", 0.32)
    cy_min = _float_env("INDIGO_FOLIOS_MATCH_CY_MIN", 0.14)
    cy_max = _float_env("INDIGO_FOLIOS_MATCH_CY_MAX", 0.78)
    offset_fila = int(_float_env("INDIGO_FOLIOS_OFFSET_PRIMERA_FILA_PX", 32))
    escalas = _escalas_template_folios()
    hay, _, _ = _origen_y_captura_para_match_hwnd(hwnd, "")
    umbrales_fb = (0.34, 0.28, 0.24, 0.20, 0.16, 0.14)

    print(
        f"Flujo: clic derecho 1.er folio (plantilla 5; umbral ≥{min_sc:.2f}; "
        f"encabezado/iconos/completa; escalas DPI; fallback fracción "
        f"INDIGO_FOLIOS_CLICK_FRACC_X/Y)"
    )
    ultimo_error: Exception | None = None
    for etiqueta, tpl_var in _variantes_tpl_listado_folios(tpl_full):
        thr = min_sc if etiqueta == "completa" else max(0.24, min_sc - 0.06)
        for roi_rel in rois_list:
            hit = _mejor_match_multiescala(
                hay,
                tpl_var,
                roi_rel,
                "top",
                thr,
                cy_min_frac=cy_min,
                cy_max_frac=cy_max,
                escalas=escalas,
            )
            if hit is None:
                for thr_fb in umbrales_fb:
                    hit = _mejor_match_multiescala(
                        hay,
                        tpl_var,
                        (0.0, 0.0, 1.0, 1.0),
                        "top",
                        thr_fb,
                        cy_min_frac=cy_min,
                        cy_max_frac=cy_max,
                        escalas=escalas,
                    )
                    if hit is not None:
                        break
            if hit is None:
                continue
            cx, cy, sc, _tw, th = hit
            if etiqueta.startswith("encabezado"):
                cy = int(cy) + int(th) + offset_fila
            print(
                f"IndiGO folio 5 ({etiqueta}): match {sc:.2f} → clic derecho en 1.ª fila "
                f"(coords. captura {cx},{cy})"
            )
            _clic_derecho_pantalla_coords(hwnd, int(cx), int(cy), f"IndiGO folio 5 ({etiqueta})")
            return
        ultimo_error = RuntimeError(f"sin match ({etiqueta}, umbral {thr:.2f})")
        print(f"Flujo: folio 5 ({etiqueta}) sin match; se prueba otra variante.")

    if _bool_env("INDIGO_FOLIOS_FRACCION_FALLBACK", True):
        print(
            "Flujo: plantilla 5 no coincidió (otro DPI/fuentes o filas distintas); "
            "clic derecho por posición en el área del listado."
        )
        _clic_derecho_fraccion_listado_folios(hwnd)
        return

    raise RuntimeError(
        f"No se encontró el listado de folios (plantilla 5) para clic derecho. {ultimo_error} "
        "Active INDIGO_FOLIOS_FRACCION_FALLBACK=1 (defecto) o ajuste INDIGO_FOLIOS_CLICK_FRACC_X/Y; "
        "recorte el PNG solo con encabezados o iconos de la 1.ª fila."
    ) from ultimo_error


def _plantilla_click_es_go_fab(nombre_plantilla: str) -> bool:
    """Botón flotante GO (menú horizontal): solo si el nombre de la plantilla menciona GO (no cualquier «5 …»)."""
    low = (nombre_plantilla or "").lower()
    if "go" not in low:
        return False
    m = re.match(r"^(\d+(?:\.\d*)?)\s+", (nombre_plantilla or "").strip())
    if not m:
        return False
    return m.group(1) in ("5", "5.")


def _roi_busqueda_go_fab() -> tuple[float, float, float, float]:
    """ROI por defecto en zona inferior derecha (FAB). Sobrescriba con INDIGO_GO_ROI=x0 y0 x1 y1 (cuatro números)."""
    raw = (os.environ.get("INDIGO_GO_ROI") or "").strip().replace(",", " ").split()
    if len(raw) >= 4:
        try:
            return (
                float(raw[0].replace(",", ".")),
                float(raw[1].replace(",", ".")),
                float(raw[2].replace(",", ".")),
                float(raw[3].replace(",", ".")),
            )
        except ValueError:
            pass
    return (0.54, 0.34, 0.996, 0.996)


def _rois_ampliados_plantilla_8_paciente(
    nombre_plantilla: str,
    roi: tuple[float, float, float, float],
    roi_def: tuple[float, float, float, float],
) -> list[tuple[float, float, float, float]]:
    """ROIs extra para plantillas de formulario: «1» cédula, «2» motivo, «3» observación, «8» Paciente (compat.).

    Para 2 y 3 se prioriza una franja superior del cliente: evita coincidencias falsas con la rejilla inferior
    o entre filas parecidas (label + caja blanca).
    """
    if roi != roi_def:
        return [roi]
    m = re.match(r"^(\d+(?:\.\d*)?)\s+", (nombre_plantilla or "").strip())
    if not m:
        return [roi_def]
    num = m.group(1)
    low = (nombre_plantilla or "").lower()
    es_campo_doc = num == "1" or (
        num == "8"
        and any(
            x in low
            for x in ("paciente", "cedula", "cédula", "excel", "input", "binocular", "documento")
        )
    )
    if es_campo_doc:
        out: list[tuple[float, float, float, float]] = []
        for r in (
            roi_def,
            (0.0, 0.12, 1.0, 1.0),
            (0.0, 0.18, 1.0, 1.0),
            (0.04, 0.22, 0.96, 0.98),
        ):
            if r not in out:
                out.append(r)
        return out
    if num == "2" and any(x in low for x in ("motivo", "select", "consulta")):
        out2: list[tuple[float, float, float, float]] = []
        for r in (
            (0.02, 0.04, 0.98, 0.36),
            (0.02, 0.06, 0.98, 0.40),
            (0.02, 0.08, 0.98, 0.44),
            roi_def,
        ):
            if r not in out2:
                out2.append(r)
        return out2
    if num == "3" and any(x in low for x in ("observacion", "observación", "comentario")):
        out3: list[tuple[float, float, float, float]] = []
        for r in (
            (0.02, 0.14, 0.98, 0.62),
            (0.02, 0.10, 0.98, 0.68),
            roi_def,
        ):
            if r not in out3:
                out3.append(r)
        return out3
    if num == "5" and any(x in low for x in ("folio", "folios", "listado")):
        out5: list[tuple[float, float, float, float]] = []
        for r in (
            (0.0, 0.16, 1.0, 0.92),
            (0.0, 0.20, 1.0, 0.88),
            (0.02, 0.24, 0.98, 0.85),
            roi_def,
        ):
            if r not in out5:
                out5.append(r)
        return out5
    if num == "9" and "imprimir" in low:
        out9: list[tuple[float, float, float, float]] = []
        for r in (
            (0.42, 0.66, 1.0, 0.98),
            (0.38, 0.60, 0.98, 0.96),
            (0.48, 0.70, 0.96, 0.95),
            roi_def,
        ):
            if r not in out9:
                out9.append(r)
        return out9
    if num == "10" and "imprimir" in low:
        out10: list[tuple[float, float, float, float]] = []
        for r in (
            (0.50, 0.76, 0.99, 0.99),
            (0.42, 0.70, 0.99, 0.99),
            (0.55, 0.82, 0.98, 0.98),
            roi_def,
        ):
            if r not in out10:
                out10.append(r)
        return out10
    return [roi_def]


def _plantilla_es_guardar_pdf_cedula(nombre_plantilla: str) -> bool:
    low = (nombre_plantilla or "").lower()
    if not re.match(r"^11\s+", (nombre_plantilla or "").strip()):
        return False
    return any(x in low for x in ("guardar", "pdf", "cedula", "cédula", "nombre"))


def _hwnd_guardar_pdf_si_visible() -> int | None:
    return _seleccionar_hwnd_dialogo_guardar_pdf()


def _fraccion_click_campo_nombre_guardar_pdf() -> tuple[float, float]:
    raw = (os.environ.get("INDIGO_PLANTILLA_11_FRACCION_CLICK") or "").strip().replace(",", " ").split()
    if len(raw) >= 2:
        try:
            return float(raw[0]), float(raw[1])
        except ValueError:
            pass
    # PNG 11 suele incluir la etiqueta «Nombre:» a la izquierda; clic en la caja de texto (derecha).
    return (
        _float_env("INDIGO_PLANTILLA_11_FRACCION_X", 0.82),
        _float_env("INDIGO_PLANTILLA_11_FRACCION_Y", 0.50),
    )


def _gettext_control_win32(hwnd_edit: int) -> str:
    """Lee el texto de un Edit/combobox (WM_GETTEXT); fiable frente a Ctrl+C en Guardar como."""
    if hwnd_edit <= 0 or not user32.IsWindow(hwnd_edit):
        return ""
    try:
        n = int(win32gui.SendMessage(hwnd_edit, win32con.WM_GETTEXTLENGTH, 0, 0))
    except Exception:
        n = 0
    if n <= 0:
        return (win32gui.GetWindowText(hwnd_edit) or "").strip()
    buf = ctypes.create_unicode_buffer(n + 4)
    try:
        win32gui.SendMessage(hwnd_edit, win32con.WM_GETTEXT, n + 1, buf)
    except Exception:
        return (win32gui.GetWindowText(hwnd_edit) or "").strip()
    return (buf.value or "").strip()


def _enum_edits_guardar_pdf_recursivo(hwnd_root: int) -> list[tuple[int, str]]:
    """Todos los controles Edit visibles dentro del diálogo Guardar (#32770)."""
    result: list[tuple[int, str]] = []

    def visit(h: int) -> None:
        cls = (win32gui.GetClassName(h) or "").lower()
        if cls == "edit" and user32.IsWindowVisible(h):
            result.append((int(h), _gettext_control_win32(int(h))))
        child = 0
        while True:
            try:
                child = int(win32gui.FindWindowEx(h, child, None, None) or 0)
            except Exception:
                break
            if not child:
                break
            visit(child)

    visit(int(hwnd_root))
    return result


def _hwnd_edit_en_combobox_ex(hwnd_dialog: int) -> int | None:
    """Edit hijo de ComboBoxEx32 (campo «Nombre de archivo» en Guardar como clásico)."""
    encontrados: list[int] = []

    def visit(h: int) -> None:
        cls = (win32gui.GetClassName(h) or "").lower()
        if cls in ("comboboxex32", "combobox"):
            ch = 0
            while True:
                try:
                    ch = int(win32gui.FindWindowEx(h, ch, "Edit", None) or 0)
                except Exception:
                    break
                if not ch:
                    break
                if user32.IsWindowVisible(ch):
                    encontrados.append(ch)
        child = 0
        while True:
            try:
                child = int(win32gui.FindWindowEx(h, child, None, None) or 0)
            except Exception:
                break
            if not child:
                break
            visit(child)

    visit(int(hwnd_dialog))
    if not encontrados:
        return None
    if len(encontrados) == 1:
        return encontrados[0]
    mejor = max(encontrados, key=lambda h: len(_gettext_control_win32(h)))
    return int(mejor)


def _hwnd_campo_nombre_guardar_pdf(hwnd_dialog: int) -> int | None:
    """Edit del nombre de archivo (p. ej. «Verificación Tic.pdf» por defecto del spooler)."""
    h_edt = _hwnd_edt1_guardar_pdf(hwnd_dialog)
    if h_edt is not None:
        txt = _gettext_control_win32(h_edt)
        print(f"Flujo: campo nombre PDF — edt1 hwnd={h_edt} texto actual={txt!r}")
        return int(h_edt)
    h_combo = _hwnd_edit_en_combobox_ex(int(hwnd_dialog))
    if h_combo is not None:
        txt = _gettext_control_win32(h_combo)
        print(f"Flujo: campo nombre PDF — ComboBoxEx Edit hwnd={h_combo} texto actual={txt!r}")
        return int(h_combo)

    edits = _enum_edits_guardar_pdf_recursivo(hwnd_dialog)
    if not edits:
        return None

    def puntaje(item: tuple[int, str]) -> float:
        _h, texto = item
        t = texto or ""
        s = float(len(t))
        if ".pdf" in t.lower():
            s += 90.0
        if re.search(r"\d{5,}", t):
            s += 55.0
        if re.search(r"\.[a-z0-9]{2,5}$", t, flags=re.IGNORECASE):
            s += 25.0
        return s

    edits.sort(key=puntaje, reverse=True)
    h_mejor, txt = edits[0]
    print(f"Flujo: campo nombre PDF — Edit hwnd={h_mejor} texto actual={txt!r}")
    return int(h_mejor)


# Diálogo «Guardar impresión como» / Guardar como clásico (#32770): IDs de dlgs.h
_GUARDAR_PDF_CTRL_EDT1 = 1152
_GUARDAR_PDF_CTRL_CMB13 = 1148
_CDM_FIRST = win32con.WM_USER + 100
_CDM_SETCONTROLTEXT = _CDM_FIRST + 4


def _hwnd_edt1_guardar_pdf(hwnd_dialog: int) -> int | None:
    """Edit «Nombre de archivo» por ID de recurso (más fiable que enumerar hijos)."""
    for ctrl_id in (_GUARDAR_PDF_CTRL_EDT1, _GUARDAR_PDF_CTRL_CMB13):
        try:
            h = int(win32gui.GetDlgItem(int(hwnd_dialog), ctrl_id) or 0)
        except Exception:
            h = 0
        if h <= 0:
            continue
        cls = (win32gui.GetClassName(h) or "").lower()
        if cls in ("comboboxex32", "combobox"):
            h_edit = _hwnd_edit_en_combobox_ex(h)
            if h_edit is not None:
                return int(h_edit)
            continue
        if cls == "edit":
            return h
    return None


def _escribir_nombre_guardar_pdf_via_cdm(hwnd_dialog: int, nombre: str) -> bool:
    """CDM_SETCONTROLTEXT en edt1/cmb13 — sustituye «Verificación Tic» del spooler por la cédula."""
    nom = _normalizar_cedula_nombre_pdf(nombre)
    if not nom or hwnd_dialog <= 0:
        return False
    nom_pdf = nom if nom.lower().endswith(".pdf") else f"{nom}.pdf"
    _activar_dialogo_guardar_pdf_forzado(int(hwnd_dialog))
    time.sleep(0.08)
    ok = False
    for ctrl_id, candidato in (
        (_GUARDAR_PDF_CTRL_EDT1, nom_pdf),
        (_GUARDAR_PDF_CTRL_EDT1, nom),
        (_GUARDAR_PDF_CTRL_CMB13, nom_pdf),
        (_GUARDAR_PDF_CTRL_CMB13, nom),
    ):
        try:
            r = int(
                win32gui.SendMessage(
                    int(hwnd_dialog),
                    _CDM_SETCONTROLTEXT,
                    int(ctrl_id),
                    candidato,
                )
                or 0
            )
            if r:
                ok = True
        except Exception as ex:
            print(f"Flujo: CDM_SETCONTROLTEXT id={ctrl_id} falló ({ex})")
    h_edit = _hwnd_edt1_guardar_pdf(hwnd_dialog)
    if h_edit is not None:
        for candidato in (nom, nom_pdf):
            if _escribir_texto_en_edit_guardar(int(h_edit), candidato):
                ok = True
    leido = _leer_nombre_archivo_guardar_pdf(hwnd_dialog)
    if _nombre_archivo_coincide_cedula(leido, nom) and not _nombre_es_titulo_spooler_motivo(leido, nom):
        print(f"Flujo: nombre PDF vía CDM/Edit → {leido!r} (cédula {nom!r})")
        return True
    if ok:
        print(
            f"Flujo: CDM/Edit enviado pero el campo muestra {leido!r} "
            f"(esperada cédula {nom!r}); se probará teclado."
        )
    return False


def _preparar_campo_nombre_guardar_pdf_para_escritura(hwnd_indigo: int = 0) -> None:
    """Foco en el cuadro Nombre (plantilla 11 si existe, si no clic en Edit por HWND)."""
    _traer_dialogos_guardar_pdf_al_frente()
    hwnd_g = _buscar_hwnd_guardar_pdf_abierto(restore=True)
    if hwnd_g is None:
        return
    num_map = _plantillas_png_por_numero()
    tpl11 = num_map.get(11)
    if tpl11 and int(hwnd_indigo or 0) > 0:
        try:
            _clic_campo_guardar_pdf_cedula(
                int(hwnd_indigo),
                tpl11,
                (0.02, 0.02, 0.98, 0.98),
                (0.02, 0.02, 0.98, 0.98),
            )
            return
        except Exception as ex:
            print(f"Flujo: aviso — clic plantilla 11 antes de escribir cédula: {ex}")
    _activar_dialogo_guardar_pdf_forzado(int(hwnd_g))
    _foco_campo_nombre_guardar_pdf(int(hwnd_g))
    _clic_fisico_edit_nombre_guardar_pdf(int(hwnd_g))


def _escribir_texto_en_edit_guardar(hwnd_edit: int, texto: str) -> bool:
    """Sustituye todo el texto del Edit (EM_REPLACESEL); no usa portapapeles."""
    if hwnd_edit <= 0 or not user32.IsWindow(h_edit := hwnd_edit):
        return False
    t = texto or ""
    try:
        win32gui.SendMessage(h_edit, win32con.EM_SETSEL, 0, -1)
        win32gui.SendMessage(h_edit, win32con.EM_REPLACESEL, True, t)
    except Exception as ex:
        print(f"Flujo: EM_REPLACESEL falló ({ex}); se prueba WM_SETTEXT.")
        try:
            win32gui.SendMessage(h_edit, win32con.WM_SETTEXT, 0, t)
        except Exception as ex2:
            print(f"Flujo: WM_SETTEXT falló ({ex2}).")
            return False
    time.sleep(0.08)
    return True


def _leer_nombre_archivo_guardar_pdf(hwnd_dialog: int) -> str:
    h = _hwnd_campo_nombre_guardar_pdf(hwnd_dialog)
    if h is None:
        return ""
    return _gettext_control_win32(h)


def _nombre_archivo_coincide_cedula(texto_campo: str, ced: str) -> bool:
    if not ced:
        return False
    raw = (texto_campo or "").strip()
    base = re.sub(r"\.pdf$", "", raw, flags=re.IGNORECASE).strip()
    ced_l = ced.strip().lower()
    return base.lower() == ced_l or ced_l in base.lower()


def _escribir_nombre_en_todos_edits_guardar(hwnd_dialog: int, nombre: str) -> bool:
    """Prueba todos los Edit del diálogo (el nombre por defecto del spooler puede no ser el ComboBoxEx)."""
    nom = _normalizar_cedula_nombre_pdf(nombre)
    if not nom:
        return False
    nom_pdf = nom if nom.lower().endswith(".pdf") else f"{nom}.pdf"
    edits = _enum_edits_guardar_pdf_recursivo(int(hwnd_dialog))
    if not edits:
        return False
    ok = False
    for h_edit, txt in edits:
        nt = _norm_etiqueta(txt or "")
        es_nombre = (
            not (txt or "").strip()
            or ".pdf" in (txt or "").lower()
            or any(x in nt for x in ("verificacion", "verificación"))
            or "tic" in nt
            or re.search(r"\d{5,}", txt or "")
        )
        if not es_nombre:
            continue
        for candidato in (nom, nom_pdf):
            if not _escribir_texto_en_edit_guardar(int(h_edit), candidato):
                continue
            leido = _gettext_control_win32(int(h_edit))
            if _nombre_archivo_coincide_cedula(leido, nom):
                print(f"Flujo: cédula en Edit hwnd={h_edit} → {leido!r}")
                ok = True
                break
        if ok:
            break
    return ok


def _escribir_nombre_archivo_guardar_pdf_wm(hwnd_dialog: int, nombre: str) -> bool:
    """Escribe la cédula en el Edit del diálogo (sin portapapeles)."""
    nom = _normalizar_cedula_nombre_pdf(nombre)
    if not nom:
        return False
    nom_pdf = nom if nom.lower().endswith(".pdf") else f"{nom}.pdf"
    nom_base = re.sub(r"\.pdf$", "", nom_pdf, flags=re.IGNORECASE)
    if _escribir_nombre_guardar_pdf_via_cdm(hwnd_dialog, nom):
        return True
    if _escribir_nombre_en_todos_edits_guardar(hwnd_dialog, nom):
        return True
    h_edit = _hwnd_campo_nombre_guardar_pdf(hwnd_dialog)
    if h_edit is None:
        return False
    _activar_dialogo_guardar_pdf_forzado(int(hwnd_dialog))
    time.sleep(0.06)
    try:
        win32gui.SetFocus(h_edit)
    except Exception:
        pass
    time.sleep(0.05)
    ok = False
    leido = ""
    for candidato in (nom_base, nom_pdf):
        if not _escribir_texto_en_edit_guardar(h_edit, candidato):
            continue
        leido = _gettext_control_win32(h_edit)
        if _nombre_archivo_coincide_cedula(leido, nom_base):
            ok = True
            break
    if ok:
        print(f"Flujo: nombre PDF en Edit OK — {leido!r} (cédula {nom_base!r})")
    else:
        print(
            f"Flujo: Edit del Guardar — esperada cédula {nom_base!r}, "
            f"campo sigue en {leido!r} (no usar texto del motivo columna E)."
        )
    return ok


def _clic_fisico_edit_nombre_guardar_pdf(hwnd_dialog: int) -> bool:
    """Clic en el cuadro «Nombre» (requiere Guardar PDF al frente, no el modal Vie)."""
    h_edit = _hwnd_campo_nombre_guardar_pdf(hwnd_dialog)
    if h_edit is None:
        return False
    try:
        l, t, r, b = win32gui.GetWindowRect(int(h_edit))
        cx, cy = (l + r) // 2, (t + b) // 2
        pyautogui.click(cx, cy)
        time.sleep(0.1)
        return True
    except Exception as ex:
        print(f"Flujo: clic físico en campo Nombre PDF falló ({ex})")
        return False


def _escribir_nombre_pdf_solo_teclado(hwnd_dialog: int, cedula: str) -> bool:
    """Borra el nombre por defecto y teclea la cédula (sin leer/pegar portapapeles)."""
    nom = _normalizar_cedula_nombre_pdf(cedula)
    if not nom:
        return False
    _activar_dialogo_guardar_pdf_forzado(int(hwnd_dialog))
    _foco_campo_nombre_guardar_pdf(hwnd_dialog)
    _clic_fisico_edit_nombre_guardar_pdf(hwnd_dialog)
    time.sleep(0.1)
    _borrar_texto_campo_activo()
    _type_texto_campo(nom)
    time.sleep(_float_env("INDIGO_GUARDAR_PDF_TRAS_ESCRIBIR_SEG", 0.22))
    return _campo_nombre_guardar_pdf_contiene_cedula(hwnd_dialog, nom)


def _foco_campo_nombre_guardar_pdf(hwnd_dialog: int | None = None) -> None:
    """Enfoca el cuadro «Nombre:» (Edit directo o Alt+N)."""
    if hwnd_dialog is not None:
        _activar_dialogo_guardar_pdf_forzado(int(hwnd_dialog))
        h = _hwnd_edt1_guardar_pdf(int(hwnd_dialog)) or _hwnd_campo_nombre_guardar_pdf(hwnd_dialog)
        if h is not None:
            try:
                _force_foreground_window(int(hwnd_dialog))
                win32gui.SetFocus(int(hwnd_dialog))
                win32gui.SetFocus(h)
            except Exception:
                pass
            time.sleep(_float_env("INDIGO_GUARDAR_PDF_ALT_N_PAUSA", 0.12))
            return
    if not _bool_env("INDIGO_GUARDAR_PDF_ALT_N", True):
        return
    keyboard.send_keys("%n")
    time.sleep(_float_env("INDIGO_GUARDAR_PDF_ALT_N_PAUSA", 0.14))


def _borrar_texto_campo_activo(repeticiones: int | None = None) -> None:
    """Quita el nombre por defecto (p. ej. «Verificación Tic») antes de pegar la cédula."""
    n = repeticiones if repeticiones is not None else int(_float_env("INDIGO_GUARDAR_PDF_BORRAR_REPETICIONES", 3))
    n = max(1, min(6, n))
    for _ in range(n):
        keyboard.send_keys("^a")
        time.sleep(0.05)
        keyboard.send_keys("{DELETE}")
        time.sleep(0.04)
        keyboard.send_keys("{BACKSPACE}")
        time.sleep(0.04)
        keyboard.send_keys("{HOME}")
        time.sleep(0.03)
        keyboard.send_keys("+{END}")
        time.sleep(0.03)
        keyboard.send_keys("{DELETE}")
        time.sleep(0.04)


def _esperar_dialogo_guardar_pdf_estable(
    timeout_sec: float | None = None,
    pausa_estabilizar: float | None = None,
) -> int:
    """Espera el diálogo Guardar PDF y una pausa para que el campo Nombre acepte texto."""
    timeout = max(10.0, float(timeout_sec or _float_env("INDIGO_GUARDAR_PDF_APARECER_TIMEOUT", 90.0)))
    pausa = max(
        0.5,
        float(
            pausa_estabilizar
            if pausa_estabilizar is not None
            else _float_env("INDIGO_GUARDAR_PDF_ESTABILIZAR_SEG", 3.5)
        ),
    )
    need = int(_float_env("INDIGO_GUARDAR_PDF_ESTABLE_LECTURAS", 3))
    need = max(2, min(8, need))
    deadline = time.time() + timeout
    racha = 0
    ultimo_log = time.time()
    print(
        f"Flujo: esperar diálogo Guardar PDF estable (máx {timeout:.0f}s; "
        f"pausa {pausa:.1f}s; {need} lecturas; INDIGO_GUARDAR_PDF_ESTABILIZAR_SEG)"
    )
    while time.time() < deadline:
        _traer_dialogos_guardar_pdf_al_frente()
        h = _hwnd_guardar_pdf_si_visible()
        if h is not None:
            racha += 1
            if racha >= need:
                _activar_dialogo_guardar_pdf_forzado(int(h))
                time.sleep(pausa)
                tit = (win32gui.GetWindowText(int(h)) or "").strip()
                print(
                    f"Flujo: Guardar PDF listo — «{tit}» (hwnd={h}); "
                    f"pausa estabilización {pausa:.1f}s (modal Vie detrás)"
                )
                _prefijar_nombre_guardar_pdf_si_posible(int(h))
                return int(h)
        else:
            racha = 0
        if time.time() - ultimo_log >= 8.0:
            print(
                f"Flujo: … esperando Guardar PDF (racha {racha}/{need}; "
                f"quedan {max(0, int(deadline - time.time()))}s)"
            )
            ultimo_log = time.time()
        time.sleep(0.35)
    raise TimeoutError(
        f"No apareció el diálogo Guardar PDF estable en {timeout:.0f}s. "
        "Revise plantilla 10, INDIGO_GUARDAR_PDF_APARECER_TIMEOUT o la impresora PDF."
    )


def _campo_nombre_guardar_pdf_contiene_cedula(hwnd_dialog: int, ced: str) -> bool:
    """Comprueba el Edit del diálogo (WM_GETTEXT), no el portapapeles (evita falsos positivos)."""
    if not _bool_env("INDIGO_GUARDAR_PDF_VERIFICAR_PEGADO", True):
        return True
    leido = _leer_nombre_archivo_guardar_pdf(hwnd_dialog)
    ok = _nombre_archivo_coincide_cedula(leido, ced)
    if not ok:
        print(f"Flujo: verificación nombre PDF — campo={leido!r}, cédula esperada={ced!r}")
    return ok


def _escribir_cedula_en_guardar_pdf(cedula: str) -> None:
    """Nombre del PDF = cédula (columna A del Excel de pacientes), no motivo columna E."""
    ced = _normalizar_cedula_nombre_pdf(cedula)
    if not ced:
        raise RuntimeError(
            "No hay cédula para el nombre del PDF. Revise la columna A del Excel de pacientes "
            "(p. ej. Pacientes_Guarne.xlsx), no el catálogo de motivos."
        )
    print(
        f"Flujo: guardar PDF — cédula columna A (Excel pacientes) → {ced!r}; "
        "no usar el texto del motivo (columna E del catálogo de consultas)."
    )
    reintentos = int(_float_env("INDIGO_GUARDAR_PDF_ESCRIBIR_REINTENTOS", 4))
    reintentos = max(1, min(8, reintentos))
    ultimo_hwnd: int | None = None
    for intento in range(1, reintentos + 1):
        hwnd = _traer_dialogos_guardar_pdf_al_frente()
        if hwnd is None:
            if intento < reintentos:
                time.sleep(0.6)
                continue
            raise RuntimeError(
                "No hay diálogo «Guardar» / Guardar PDF visible. "
                "Espere wait_template 11 / esperar_guardar_pdf o suba INDIGO_GUARDAR_PDF_APARECER_TIMEOUT."
            )
        ultimo_hwnd = int(hwnd)
        tit_hwnd = (win32gui.GetWindowText(ultimo_hwnd) or "").strip()
        print(f"Flujo: escribir cédula en diálogo «{tit_hwnd}» (hwnd={ultimo_hwnd})")
        _activar_dialogo_guardar_pdf_forzado(ultimo_hwnd)
        time.sleep(_float_env("INDIGO_GUARDAR_PDF_PREESCRITURA_SEG", 0.45))
        if _escribir_nombre_archivo_guardar_pdf_wm(ultimo_hwnd, ced):
            print(
                f"Flujo: nombre del PDF → {ced!r} "
                f"(hwnd={ultimo_hwnd}; Edit API; intento {intento}/{reintentos})"
            )
            return
        if _escribir_nombre_pdf_solo_teclado(ultimo_hwnd, ced):
            print(
                f"Flujo: nombre del PDF (teclado, sin portapapeles) → {ced!r} "
                f"(intento {intento}/{reintentos})"
            )
            return
        _activar_dialogo_guardar_pdf_forzado(ultimo_hwnd)
        _foco_campo_nombre_guardar_pdf(ultimo_hwnd)
        _clic_fisico_edit_nombre_guardar_pdf(ultimo_hwnd)
        time.sleep(0.12)
        _borrar_texto_campo_activo()
        if _establecer_portapapeles(ced):
            keyboard.send_keys("^a")
            time.sleep(0.06)
            keyboard.send_keys("^v")
            time.sleep(_float_env("INDIGO_GUARDAR_PDF_TRAS_ESCRIBIR_SEG", 0.22))
        else:
            _type_texto_campo(ced)
            time.sleep(_float_env("INDIGO_GUARDAR_PDF_TRAS_ESCRIBIR_SEG", 0.22))
        if _campo_nombre_guardar_pdf_contiene_cedula(ultimo_hwnd, ced):
            print(
                f"Flujo: nombre del PDF (último recurso pegado) → {ced!r} "
                f"(intento {intento}/{reintentos})"
            )
            return
        leido = _leer_nombre_archivo_guardar_pdf(ultimo_hwnd)
        print(
            f"Flujo: reintento escribir cédula en Guardar PDF ({intento}/{reintentos}); "
            f"campo={leido!r}…"
        )
        time.sleep(0.35)
    raise RuntimeError(
        f"No se pudo poner la cédula {ced!r} en el campo Nombre del PDF "
        f"({reintentos} intentos). El nombre por defecto del spooler (motivo de consulta) "
        "no debe usarse; revise INDIGO_PLANTILLA_11_FRACCION_CLICK o click 11."
    )


def _enum_hwnds_32770_visibles() -> list[tuple[int, str]]:
    out: list[tuple[int, str]] = []

    @ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
    def cb(hwnd, _):
        if not user32.IsWindowVisible(hwnd):
            return True
        try:
            hi = int(hwnd)
        except (TypeError, ValueError):
            return True
        if (_clase_ventana(hwnd) or "").lower() == "#32770":
            out.append((hi, (win32gui.GetWindowText(hwnd) or "").strip()))
        return True

    user32.EnumWindows(cb, 0)
    return out


def _textos_hijos_ventana(hwnd: int) -> str:
    partes: list[str] = []

    @ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
    def child_cb(ch, _):
        if user32.IsWindowVisible(ch):
            t = (win32gui.GetWindowText(ch) or "").strip()
            if t:
                partes.append(t)
        return True

    try:
        win32gui.EnumChildWindows(hwnd, child_cb, None)
    except Exception:
        pass
    return " ".join(partes)


def _es_dialogo_reemplazar_archivo(hwnd: int, titulo: str) -> bool:
    blob = _norm_etiqueta(f"{titulo or ''} {_textos_hijos_ventana(hwnd)}")
    return any(
        x in blob
        for x in (
            "reemplaz",
            "replace",
            "sobrescrib",
            "overwrite",
            "ya existe",
            "already exists",
            "desea reemplazar",
            "want to replace",
            "existe un archivo",
        )
    )


def _responder_dialogo_reemplazar_archivo() -> bool:
    """Responde «¿Desea reemplazar…?» (defecto: Sí)."""
    hwnd_guardar = _hwnd_guardar_pdf_si_visible()
    for h, tit in _enum_hwnds_32770_visibles():
        if hwnd_guardar is not None and h == hwnd_guardar:
            continue
        if not _es_dialogo_reemplazar_archivo(h, tit):
            continue
        activar_ventana_modal(h)
        time.sleep(0.1)
        resp = (os.environ.get("INDIGO_REEMPLAZAR_PDF") or "si").strip().lower()
        if resp in ("no", "n", "0", "false"):
            print(
                f"Flujo: diálogo reemplazar → No ({tit!r}); "
                "se reescribe la cédula en Guardar (no Enter con nombre del motivo)."
            )
            keyboard.send_keys("n")
            time.sleep(0.08)
            keyboard.send_keys("{ESC}")
        else:
            print(f"Flujo: diálogo reemplazar → Sí ({tit!r})")
            keyboard.send_keys("s")
            time.sleep(0.04)
            keyboard.send_keys("y")
            time.sleep(0.04)
            keyboard.send_keys("%s")
        time.sleep(_float_env("INDIGO_REEMPLAZAR_PDF_PAUSA", 0.35))
        return True
    return False


def _esperar_cierre_dialogo_guardar_pdf(timeout_sec: float) -> None:
    deadline = time.time() + max(2.0, timeout_sec)
    while time.time() < deadline:
        _responder_dialogo_reemplazar_archivo()
        if _hwnd_guardar_pdf_si_visible() is None:
            return
        time.sleep(0.22)
    print(
        f"Flujo: aviso — el diálogo Guardar PDF sigue visible tras {timeout_sec:.0f}s; "
        "se intenta cerrar ventanas de impresión."
    )


def _cerrar_modales_impresion_post_pdf(hwnd_indigo: int) -> None:
    """Cierra modal Vie «Consultar / Imprimir Historias» y diálogos Imprimir tras guardar el PDF."""
    global _pasos_sin_traer_indigo
    _pasos_sin_traer_indigo = 0
    veces = int(_float_env("INDIGO_POST_PDF_ESC_VECES", 5))
    veces = max(2, min(12, veces))
    print(
        f"Flujo: cerrar ventanas de impresión / modal Vie tras guardar PDF "
        f"(hasta {veces}× ESC; INDIGO_POST_PDF_ESC_VECES)"
    )
    for n in range(veces):
        cerrado_algo = False
        for h, tit in _enum_hwnds_modal_vie_consultar_imprimir():
            activar_ventana_modal(h)
            keyboard.send_keys("{ESC}")
            time.sleep(0.18)
            print(f"Flujo: ESC en modal Vie «{tit}» (hwnd={h})")
            cerrado_algo = True
        for h, tit in _enum_hwnds_dialogo_impresion_sistema():
            activar_ventana_modal(h)
            keyboard.send_keys("{ESC}")
            time.sleep(0.15)
            print(f"Flujo: ESC en diálogo impresión «{tit}»")
            cerrado_algo = True
        for h, tit in _enum_hwnds_guardar_pdf_impresion():
            if _hwnd_guardar_pdf_si_visible():
                activar_ventana_modal(h)
                keyboard.send_keys("{ESC}")
                time.sleep(0.12)
                cerrado_algo = True
        if not cerrado_algo:
            break
        time.sleep(0.12)
    restaurar_y_traer_al_frente(hwnd_indigo)
    time.sleep(_float_env("INDIGO_POST_PDF_TRAS_CERRAR_SEG", 0.35))


def _confirmar_guardado_pdf_cedula(hwnd_indigo: int, cedula: str = "") -> None:
    """Enter en Guardar solo si el Edit del diálogo contiene la cédula (no «Verificación Tic»)."""
    ced = _normalizar_cedula_nombre_pdf(cedula)
    if not ced:
        raise RuntimeError(
            "No se confirma el guardado PDF sin cédula (columna A del Excel de pacientes)."
        )
    hwnd = _traer_dialogos_guardar_pdf_al_frente()
    if hwnd is None:
        raise RuntimeError("No hay diálogo Guardar PDF para confirmar el guardado.")
    _activar_dialogo_guardar_pdf_forzado(int(hwnd))
    time.sleep(0.1)
    actual = _leer_nombre_archivo_guardar_pdf(int(hwnd))
    if not _nombre_archivo_coincide_cedula(actual, ced):
        if not _escribir_nombre_archivo_guardar_pdf_wm(int(hwnd), ced):
            if not _escribir_nombre_pdf_solo_teclado(int(hwnd), ced):
                raise RuntimeError(
                    f"No se pulsa Enter: el nombre del PDF sigue siendo {actual!r} "
                    f"(se requiere la cédula {ced!r} del Excel de pacientes). "
                    "«Verificación Tic.pdf» es el título por defecto del motivo de consulta, no la cédula."
                )
        actual = _leer_nombre_archivo_guardar_pdf(int(hwnd))
    if not _nombre_archivo_coincide_cedula(actual, ced):
        raise RuntimeError(
            f"No se pulsa Enter: el campo Nombre sigue en {actual!r} (cédula esperada {ced!r})."
        )
    if _nombre_es_titulo_spooler_motivo(actual, ced):
        raise RuntimeError(
            f"No se pulsa Enter: el nombre {actual!r} es el título del motivo de consulta "
            f"(spooler), no la cédula {ced!r} del Excel de pacientes."
        )
    print(f"Flujo: confirmar guardado PDF — nombre en campo: {actual!r} (cédula {ced!r})")
    time.sleep(0.06)
    print("Flujo: confirmar guardado PDF (botón Guardar / Enter + diálogo reemplazar si aparece)")
    if not _pulsar_boton_guardar_en_dialogo(int(hwnd)):
        keyboard.send_keys("{ENTER}")
    time.sleep(_float_env("INDIGO_GUARDAR_PDF_TRAS_ENTER_SEG", 0.5))
    reps = int(_float_env("INDIGO_REEMPLAZAR_PDF_INTENTOS", 12))
    for i in range(max(3, min(25, reps))):
        if _responder_dialogo_reemplazar_archivo():
            time.sleep(0.25)
            hwnd_g = _hwnd_guardar_pdf_si_visible()
            if hwnd_g is not None:
                _activar_dialogo_guardar_pdf_forzado(int(hwnd_g))
                _escribir_cedula_en_guardar_pdf(ced)
                actual2 = _leer_nombre_archivo_guardar_pdf(int(hwnd_g))
                if (
                    _nombre_archivo_coincide_cedula(actual2, ced)
                    and not _nombre_es_titulo_spooler_motivo(actual2, ced)
                ):
                    keyboard.send_keys("{ENTER}")
                    time.sleep(0.35)
                else:
                    print(
                        f"Flujo: tras reemplazar — nombre sigue en {actual2!r}; "
                        "no se pulsa Enter hasta tener la cédula."
                    )
            continue
        if _hwnd_guardar_pdf_si_visible() is None:
            break
        time.sleep(0.2)
    _esperar_cierre_dialogo_guardar_pdf(_float_env("INDIGO_GUARDAR_PDF_CIERRE_TIMEOUT", 18.0))
    time.sleep(0.25)
    _cerrar_modales_impresion_post_pdf(hwnd_indigo)


def _guardar_pdf_con_cedula_completo(hwnd_indigo: int, cedula: str) -> None:
    """Paso 11 completo: esperar Guardar estable, nombre = cédula (Excel pacientes), confirmar."""
    ced = _normalizar_cedula_nombre_pdf(cedula)
    if not ced:
        raise RuntimeError(
            "guardar_pdf_cedula: sin cédula en sesión. El bucle debe asignar sesion.cedula_actual "
            "desde la columna A del Excel de la carpeta E.S.E. (Pacientes_*.xlsx)."
        )
    print(f"Flujo: guardar_pdf_cedula — cédula del Excel de pacientes (columna A): {ced!r}")
    h0 = _hwnd_guardar_pdf_si_visible()
    if h0 is not None:
        _activar_dialogo_guardar_pdf_forzado(int(h0))
    _esperar_dialogo_guardar_pdf_estable()
    time.sleep(_float_env("INDIGO_GUARDAR_PDF_TRAS_ESTABLE_EXTRA_SEG", 0.35))
    _preparar_campo_nombre_guardar_pdf_para_escritura(hwnd_indigo)
    _escribir_cedula_en_guardar_pdf(ced)
    _confirmar_guardado_pdf_cedula(hwnd_indigo, ced)


def _clic_campo_guardar_pdf_cedula(
    hwnd_vie: int,
    nombre_tpl: str,
    roi: tuple[float, float, float, float],
    roi_def: tuple[float, float, float, float],
) -> None:
    """Clic en el campo Nombre del diálogo Guardar PDF (plantilla 11), en la ventana correcta."""
    tpl = _resolver_ruta_plantilla_png(nombre_tpl)
    _traer_dialogos_guardar_pdf_al_frente()
    hwnd_g = _hwnd_guardar_pdf_si_visible()
    hwnd_u = int(hwnd_g or hwnd_vie)
    rois_list = [(0.0, 0.0, 1.0, 1.0)] if hwnd_g else _rois_ampliados_plantilla_8_paciente(nombre_tpl, roi, roi_def)
    min_sc = _float_env("INDIGO_MATCH_MIN_PLANTILLA_11", 0.34)
    fraccion = _fraccion_click_campo_nombre_guardar_pdf()
    activar_ventana_modal(hwnd_u)
    time.sleep(0.12)
    dest = "Guardar PDF" if hwnd_g else "pantalla Vie"
    print(
        f"Flujo: clic plantilla 11 en {dest} (hwnd={hwnd_u}; campo Nombre fracción "
        f"{fraccion[0]:.2f},{fraccion[1]:.2f})"
    )
    clic_plantilla_en_hwnd(
        hwnd_u,
        tpl,
        rois_list,
        "left",
        min_sc,
        f"IndiGO guardar PDF {tpl.name}",
        pantalla_completa_si_falla=bool(hwnd_g),
        umbrales_pantalla_completa=(0.32, 0.28, 0.24),
        fraccion_click_en_plantilla=fraccion,
        boton="left",
    )
    time.sleep(_float_env("INDIGO_GUARDAR_PDF_TRAS_CLIC_CAMPO_SEG", 0.22))
    _foco_campo_nombre_guardar_pdf(hwnd_u)


def _plantilla_es_deshacer_post_impresion(nombre_plantilla: str) -> bool:
    low = (nombre_plantilla or "").lower()
    if not re.match(r"^12\s+", (nombre_plantilla or "").strip()):
        return False
    return "deshacer" in low


def _clic_deshacer_post_impresion(
    hwnd: int,
    nombre_tpl: str,
    roi: tuple[float, float, float, float],
    roi_def: tuple[float, float, float, float],
) -> None:
    """Deshacer en Vie tras guardar el PDF (plantilla 12) para preparar la siguiente cédula."""
    tpl = _resolver_ruta_plantilla_png(nombre_tpl)
    restaurar_y_traer_al_frente(hwnd)
    time.sleep(_float_env("INDIGO_DESHACER_PRECLICK_SEG", 0.25))
    rois_list = _rois_ampliados_plantilla_8_paciente(nombre_tpl, roi, roi_def)
    min_sc = _float_env("INDIGO_MATCH_MIN_PLANTILLA_12", 0.28)
    print(f"Flujo: clic Deshacer (plantilla 12; similitud ≥{min_sc:.2f}) en ventana IndiGO")
    clic_plantilla_en_hwnd(
        hwnd,
        tpl,
        rois_list,
        "left",
        min_sc,
        f"IndiGO deshacer {tpl.name}",
        pantalla_completa_si_falla=True,
        umbrales_pantalla_completa=(0.24, 0.20, 0.16),
        boton="left",
    )
    time.sleep(_float_env("INDIGO_SLEEP_TRAS_DESHACER", 0.45))


def _max_similitud_plantilla_guardar_pdf(
    hwnd_vie: int,
    tpl_path: Path,
    rois_vie: list[tuple[float, float, float, float]],
) -> float:
    _traer_dialogos_guardar_pdf_al_frente()
    h = _hwnd_guardar_pdf_si_visible()
    if h is not None:
        return _max_similitud_plantilla_hwnd(h, tpl_path, [(0.0, 0.0, 1.0, 1.0)])
    return _max_similitud_plantilla_hwnd(hwnd_vie, tpl_path, rois_vie)


def _plantilla_es_imprimir_primero(nombre_plantilla: str) -> bool:
    low = (nombre_plantilla or "").lower()
    return bool(re.match(r"^9\s+", (nombre_plantilla or "").strip())) and "imprimir" in low


def _plantilla_es_imprimir_modal_segundo(nombre_plantilla: str) -> bool:
    low = (nombre_plantilla or "").lower()
    return bool(re.match(r"^10\s+", (nombre_plantilla or "").strip())) and "imprimir" in low


def _params_exclusion_imprimir_9() -> tuple[tuple[int, int] | None, int]:
    radio = int(_float_env("INDIGO_EXCLUIR_RADIO_IMPRIMIR_9", 110))
    if _xy_imprimir_9_en_captura is None or radio <= 0:
        return None, 0
    return _xy_imprimir_9_en_captura, radio


def _nombre_plantilla_imprimir_modal_segundo() -> str:
    num_map = _plantillas_png_por_numero()
    nombre = num_map.get(10)
    if nombre:
        return nombre
    cands = sorted(TEMPLATE_DIR.glob("10*.png"))
    if not cands:
        raise RuntimeError(
            "No hay plantilla «10 …Imprimir….png» en templates/. "
            "Añádala o use wait_template 10 en el flujo."
        )
    return cands[0].name


def _advertir_plantillas_9_10_muy_similares(tpl10_path: Path) -> None:
    """Aviso si los PNG 9 y 10 son casi iguales (falsos positivos / mismo recorte)."""
    if not _bool_env("INDIGO_IMP10_AVISO_SIMILAR_9", True):
        return
    num_map = _plantillas_png_por_numero()
    n9 = num_map.get(9)
    if not n9:
        return
    p9 = TEMPLATE_DIR / n9
    g9 = _cv2_imread_path(p9)
    g10 = _cv2_imread_path(tpl10_path)
    if g9 is None or g10 is None or g9.size == 0 or g10.size == 0:
        return
    if g9.ndim == 3 and g9.shape[2] == 4:
        g9 = cv2.cvtColor(g9, cv2.COLOR_BGRA2BGR)
    if g10.ndim == 3 and g10.shape[2] == 4:
        g10 = cv2.cvtColor(g10, cv2.COLOR_BGRA2BGR)
    s9 = cv2.resize(cv2.cvtColor(g9, cv2.COLOR_BGR2GRAY), (72, 72))
    s10 = cv2.resize(cv2.cvtColor(g10, cv2.COLOR_BGR2GRAY), (72, 72))
    a = s9.astype(np.float64).ravel()
    b = s10.astype(np.float64).ravel()
    if np.std(a) < 1e-6 or np.std(b) < 1e-6:
        return
    corr = float(np.corrcoef(a, b)[0, 1])
    thr = _float_env("INDIGO_IMP10_CORR_AVISO_MAX", 0.92)
    if corr >= thr:
        print(
            f"Aviso: plantillas 9 ({n9}) y 10 ({tpl10_path.name}) parecen muy similares "
            f"(corr≈{corr:.2f}; umbral aviso {thr:.2f}). Use recortes distintos (DPI, hover, zoom)."
        )


def _guardar_captura_debug_imp10(hwnd_cap: int, ultima_s: float, sc_win: float) -> None:
    """Guarda la captura usada para match plantilla 10 si la similitud cae en banda dudosa (p. ej. carga lenta / DPI)."""
    global _imp10_debug_last_save
    if not _bool_env("INDIGO_IMP10_DEBUG_SAVE", False):
        return
    lo = _float_env("INDIGO_IMP10_DEBUG_SIM_LO", 0.68)
    hi = _float_env("INDIGO_IMP10_DEBUG_SIM_HI", 0.78)
    mx = max(ultima_s, sc_win)
    if not (lo <= mx <= hi):
        return
    iv = float(_float_env("INDIGO_IMP10_DEBUG_INTERVAL", 18.0))
    now = time.time()
    if now - _imp10_debug_last_save < iv:
        return
    if hwnd_cap <= 0 or not user32.IsWindow(hwnd_cap):
        return
    _imp10_debug_last_save = now
    raw_dir = (os.environ.get("INDIGO_IMP10_DEBUG_DIR") or "").strip()
    d = Path(raw_dir).expanduser() if raw_dir else (ROOT / "debug_imp10_capturas")
    try:
        d.mkdir(parents=True, exist_ok=True)
    except OSError:
        return
    try:
        hay, _, _ = _origen_y_captura_para_match_hwnd(hwnd_cap, "")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        p = d / f"imp10_{ts}_mv{ultima_s:.2f}_win{sc_win:.2f}.png"
        cv2.imwrite(str(p), hay)
        print(f"Flujo: depuración INDIGO_IMP10_DEBUG_SAVE — guardada {p.name} (banda similitud {lo:.2f}–{hi:.2f}).")
    except Exception as ex:
        print(f"Flujo: no se pudo guardar depuración IMP10 ({ex}).")


def _clic_imprimir_primero(
    hwnd: int,
    nombre_tpl: str,
    roi: tuple[float, float, float, float],
    roi_def: tuple[float, float, float, float],
) -> None:
    global _ts_clic_imprimir_9, _xy_imprimir_9_en_captura
    tpl = _resolver_ruta_plantilla_png(nombre_tpl)
    rois_ampliados = _rois_ampliados_plantilla_8_paciente(nombre_tpl, roi, roi_def)
    min_sc = _float_env("INDIGO_MATCH_MIN_PLANTILLA_9", 0.70)
    raw_clic = (os.environ.get("INDIGO_MATCH_MIN_PLANTILLA_9_CLIC") or "").strip().replace(",", ".")
    if raw_clic:
        try:
            min_sc_clic = max(0.45, min(0.99, float(raw_clic)))
        except ValueError:
            min_sc_clic = max(0.52, min_sc - 0.08)
    else:
        min_sc_clic = max(0.52, min_sc - 0.08)
    min_espera = _float_env("INDIGO_ESPERA_MIN_TRAS_IMPRIMIR_9", 75.0)
    timeout_modal = _float_env("INDIGO_MODAL_VIE_ESPERA_TIMEOUT", 30.0)

    found = _esperar_hwnd_modal_vie_consultar_imprimir(timeout_modal)
    hwnd_clic = hwnd
    tit_clic = (win32gui.GetWindowText(hwnd) or "").strip()
    if found:
        hwnd_clic, tit_clic = found
        activar_ventana_modal(hwnd_clic)
        time.sleep(_float_env("INDIGO_MODAL_VIE_PRECLICK_SEG", 0.22))
        print(
            f"Flujo: 1.er Imprimir (plantilla 9) en modal Vie «{tit_clic}» "
            f"(hwnd={hwnd_clic}; similitud clic ≥{min_sc_clic:.2f})"
        )
    else:
        print(
            "Flujo: aviso — no se detectó el modal «Consultar / Imprimir Historias»; "
            "plantilla 9 en la ventana IndiGO principal (revise paso 8 o suba "
            "INDIGO_MODAL_VIE_ESPERA_TIMEOUT)."
        )
        restaurar_y_traer_al_frente(hwnd)

    orden, cx_min, cy_min, rois_ctx = _match_boton_imprimir_10_contexto(hwnd_clic, tit_clic)
    rois_list: list[tuple[float, float, float, float]] = []
    for r in rois_ctx + rois_ampliados:
        if r not in rois_list:
            rois_list.append(r)
    fraccion_btn = _fraccion_click_boton_imprimir_plantilla_10(hwnd_clic, tpl)
    frac_txt = f"; fracción clic {fraccion_btn}" if fraccion_btn else ""
    roi_txt = (
        f"cx≥{cx_min:.2f} cy≥{cy_min:.2f}"
        if cx_min is not None and cy_min is not None
        else "ROI pantalla completa (modal Vie/Imprimir no clasificado)"
    )
    print(
        f"Flujo: plantilla 9 — ROI inferior-derecha del modal; orden={orden}; "
        f"{roi_txt}{frac_txt}; luego espera mín. {min_espera:.0f}s antes del 2.º"
    )
    ultimo_error: Exception | None = None
    try:
        cx, cy, sc = clic_plantilla_en_hwnd(
            hwnd_clic,
            tpl,
            rois_list,
            orden,
            min_sc_clic,
            "IndiGO imprimir modal Vie (1.º)",
            pantalla_completa_si_falla=True,
            umbrales_pantalla_completa=(0.62, 0.55, 0.48, 0.40, 0.34),
            fraccion_click_en_plantilla=fraccion_btn,
            boton="left",
            cx_min_frac=cx_min,
            cy_min_frac=cy_min,
        )
    except RuntimeError as ex:
        ultimo_error = ex
        if not _bool_env("INDIGO_PLANTILLA_9_FRACCION_FALLBACK", True):
            raise
        left, top, right, bottom = rect_pantalla(hwnd_clic)
        w, h = right - left, bottom - top
        if w < 80 or h < 80:
            raise RuntimeError(
                f"Plantilla 9 sin match y ventana demasiado pequeña ({w}×{h}). {ex}"
            ) from ex
        fx = _float_env("INDIGO_PLANTILLA_9_FRACCION_X", 0.78)
        fy = _float_env("INDIGO_PLANTILLA_9_FRACCION_Y", 0.92)
        cx = left + int(w * fx)
        cy = top + int(h * fy)
        cx, cy = _clamp_xy_evita_esquina_failsafe_pantalla(cx, cy)
        print(
            f"Flujo: plantilla 9 sin match PNG ({ex}); clic por posición en el modal "
            f"({fx:.2f}, {fy:.2f}) → ({cx},{cy})"
        )
        _clic_izquierdo_menu_contexto_en_pantalla(cx, cy, "IndiGO imprimir modal Vie (9 fracción)")
        sc = 0.0

    left, top, _, _ = rect_pantalla(hwnd_clic)
    _xy_imprimir_9_en_captura = (cx - left, cy - top)
    _ts_clic_imprimir_9 = time.time()
    print(
        f"Flujo: registrado 1.er Imprimir en ({_xy_imprimir_9_en_captura[0]},{_xy_imprimir_9_en_captura[1]}) "
        f"coords. captura (similitud {sc:.2f}); radio exclusión "
        f"{int(_float_env('INDIGO_EXCLUIR_RADIO_IMPRIMIR_9', 110))} px."
    )
    time.sleep(_float_env("INDIGO_SLEEP_TRAS_IMPRIMIR_9", 0.35))
    _activar_bloqueo_foco_impresion_9_10()
    _mantener_foco_flujo_imprimir_9_10(int(hwnd_clic) if found else None)


def _esperar_modal_imprimir_segundo(
    hwnd: int,
    timeout_sec: float,
    min_sim: float | None = None,
) -> None:
    """Espera tras plantilla 9 y detección estable de plantilla 10 (2.º Imprimir).

    Tras detectar el modal «Imprimir» de Windows, espera INDIGO_IMP10_ESPERA_TRAS_MODAL_IMPRIMIR_SEG
    antes de aceptar el match (render/DPI/RDP; el botón puede habilitarse un instante después).
    Con INDIGO_IMP10_ADAPTIVE=1 admite similitud intermedia si el centro del match es estable N lecturas
    (sincronía con sistema lento). El matching usa escala múltiple (TEMPLATE_SCALES) y grises; bordes
    opcionales con INDIGO_IMP10_MATCH_EDGES=1.
    """
    global _hwnd_destino_imprimir_10, _omitir_clic_plantilla_10
    _hwnd_destino_imprimir_10 = None
    _omitir_clic_plantilla_10 = False
    _activar_bloqueo_foco_impresion_9_10()
    if _imp10_saltar_si_guardar_pdf_abierto():
        return
    nombre = _nombre_plantilla_imprimir_modal_segundo()
    tpl = _resolver_ruta_plantilla_png(nombre)
    _advertir_plantillas_9_10_muy_similares(tpl)
    roi_def = (0.02, 0.02, 0.98, 0.98)
    rois_list = _rois_ampliados_plantilla_8_paciente(nombre, roi_def, roi_def)
    min_espera = _float_env("INDIGO_ESPERA_MIN_TRAS_IMPRIMIR_9", 75.0)
    umbral_ok = (
        float(min_sim)
        if min_sim is not None
        else _float_env(
            "INDIGO_WAIT_MATCH_MIN_IMPRIMIR_10",
            _float_env("INDIGO_WAIT_MATCH_MIN_MODAL_IMPRIMIR", 0.58),
        )
    )
    umbral_ok = max(0.45, min(0.99, umbral_ok))
  # En diálogo #32770 el mismo PNG suele puntuar menos (0,55–0,70) que en el modal Vie.
    umbral_win_env = (os.environ.get("INDIGO_WAIT_MATCH_MIN_IMPRIMIR_10_WINDOWS") or "").strip()
    umbral_win_def = (
        float(umbral_win_env.replace(",", "."))
        if umbral_win_env
        else _float_env("INDIGO_WAIT_MATCH_MIN_IMPRIMIR_10_WINDOWS", 0.55)
    )
    if _imp10_requiere_dialogo_imprimir_windows():
        umbral_win = min(umbral_ok, max(0.45, min(0.99, umbral_win_def)))
    else:
        umbral_win = max(0.45, min(0.99, umbral_win_def))
    exclude_xy, exclude_radio = _params_exclusion_imprimir_9()
    raw_c = (os.environ.get("INDIGO_WAIT_CONSECUTIVE") or "").strip()
    try:
        need_ok = int(raw_c) if raw_c else 3
    except ValueError:
        need_ok = 3
    need_ok = max(1, min(need_ok, 25))

    delay_modal = _float_env("INDIGO_IMP10_ESPERA_TRAS_MODAL_IMPRIMIR_SEG", 8.0)
    extra_modal = _float_env("INDIGO_IMP10_ESPERA_TRAS_MODAL_EXTRA_SEG", 0.0)
    adaptive_on = _bool_env("INDIGO_IMP10_ADAPTIVE", True)
    adapt_min = max(0.45, _float_env("INDIGO_IMP10_ADAPTIVE_SCORE_MIN", 0.52))
    umbral_win_piso = max(0.45, _float_env("INDIGO_IMP10_WIN_PISO_ESTABLE", 0.50))
    stab_px = int(_float_env("INDIGO_IMP10_ADAPTIVE_STABLE_PX", 12))
    need_adapt = int(_float_env("INDIGO_IMP10_ADAPTIVE_CONSECUTIVE", 0))
    if need_adapt <= 0:
        need_adapt = need_ok

    t_hwin_first: float | None = None
    last_ap_pos: tuple[int, int] | None = None
    adapt_streak = 0

    if _ts_clic_imprimir_9 > 0:
        faltan = min_espera - (time.time() - _ts_clic_imprimir_9)
        if faltan > 0:
            print(
                f"Flujo: espera mínima {min_espera:.0f}s tras plantilla 9 "
                f"(faltan ~{faltan:.0f}s; no se busca plantilla 10 aún)…"
            )
            ultimo_log = time.time()
            ultimo_foco_min = time.time()
            intervalo_foco_min = _float_env("INDIGO_IMP10_REFRESCO_FOCO_MIN_ESPERA_SEG", 3.0)
            while True:
                faltan = min_espera - (time.time() - _ts_clic_imprimir_9)
                if faltan <= 0:
                    break
                if _imp10_saltar_si_guardar_pdf_abierto():
                    return
                if intervalo_foco_min > 0 and time.time() - ultimo_foco_min >= intervalo_foco_min:
                    _mantener_foco_flujo_imprimir_9_10()
                    ultimo_foco_min = time.time()
                if _imp10_requiere_dialogo_imprimir_windows():
                    h_early = _buscar_hwnd_dialogo_imprimir_windows_abierto()
                    if h_early is not None:
                        print(
                            f"Flujo: diálogo Imprimir Windows detectado durante espera tras 9 "
                            f"(hwnd={h_early}; faltaban ~{max(0, int(faltan))}s de espera mínima)"
                        )
                        _enviar_modal_vie_detras_dialogo_imprimir_windows(int(h_early))
                        _activar_dialogo_impresion_windows(int(h_early))
                if time.time() - ultimo_log >= 15.0:
                    print(f"Flujo: … espera mínima tras 9 (faltan ~{max(0, int(faltan))}s)")
                    ultimo_log = time.time()
                time.sleep(min(2.0, max(0.25, faltan)))

    deadline = time.time() + timeout_sec
    ultima_s = 0.0
    racha = 0
    ultimo_info = time.time()
    print(
        f"Flujo: esperar 2.º Imprimir {nombre!r} (máx {timeout_sec:.0f}s, "
        f"similitud ≥{umbral_ok:.2f} o Win ≥{umbral_win:.2f}; "
        f"adaptativa ≥{adapt_min:.2f} estable; piso Win {umbral_win_piso:.2f}; "
        f"{delay_modal:.1f}s post-modal; {need_ok} lecturas; INDIGO_IMP10_*)"
    )
    ultima_etq = ""
    ultimo_refresco_foco = 0.0
    intervalo_foco = _float_env("INDIGO_IMP10_REFRESCO_FOCO_SEG", 4.0)
    while time.time() < deadline:
        if _imp10_saltar_si_guardar_pdf_abierto():
            return
        # No traer IndiGO al frente: quita foco al diálogo «Imprimir».
        if intervalo_foco > 0 and time.time() - ultimo_refresco_foco >= intervalo_foco:
            _mantener_foco_flujo_imprimir_9_10()
            ultimo_refresco_foco = time.time()
        h_win = _buscar_hwnd_dialogo_imprimir_windows_abierto()
        if h_win is not None:
            if t_hwin_first is None:
                t_hwin_first = time.time()
                print(
                    f"Flujo: ventana «Imprimir» (#32770) detectada; "
                    f"esperando {delay_modal:.1f}s + extra {extra_modal:.1f}s antes de validar el botón "
                    f"(render, escalado DPI/RDP, botón puede habilitarse)…"
                )
        else:
            t_hwin_first = None

        ultima_s, hwnd_hit, ultima_etq = _max_similitud_imprimir_segundo_multiventana(
            hwnd, tpl, rois_list, exclude_xy, exclude_radio
        )

        sc_win = 0.0
        pos_win: tuple[int, int] | None = None
        tit_w = ""
        if h_win is not None:
            tit_w = (win32gui.GetWindowText(h_win) or "").strip()
            sc_win, pos_win = _detalle_match_boton_imprimir_10_en_hwnd(h_win, tpl, tit_w)

        render_ok = t_hwin_first is None or (
            time.time() >= t_hwin_first + delay_modal + extra_modal
        )
        debounce = h_win is not None and not render_ok

        cap_hw = int(h_win) if h_win else int(hwnd)
        _guardar_captura_debug_imp10(cap_hw, ultima_s, sc_win)

        if time.time() - ultimo_info >= 12.0:
            extra = f" en {ultima_etq}" if ultima_etq else ""
            sw = f"; Win {sc_win:.2f}" if h_win is not None else ""
            db = "; debounce render" if debounce else ""
            ad = f"; adapt {adapt_streak}/{need_adapt}" if adaptive_on and h_win else ""
            print(
                f"Flujo: … esperando {nombre!r} (mejor {ultima_s:.2f}{extra}{sw}{db}{ad}; "
                f"racha {racha}/{need_ok}; quedan {max(0, int(deadline - time.time()))}s)"
            )
            ultimo_info = time.time()

        if debounce:
            pasa_normal = False
            pasa_piso = False
        elif _imp10_requiere_dialogo_imprimir_windows():
            # Plantilla 10 = botón del diálogo #32770 «Imprimir», NO el del modal Vie (misma PNG que 9).
            pasa_normal = h_win is not None and render_ok and sc_win >= umbral_win
            pasa_piso = (
                h_win is not None
                and render_ok
                and sc_win >= umbral_win_piso
                and sc_win >= adapt_min
            )
        else:
            pasa_normal = ultima_s >= umbral_ok or (h_win is not None and sc_win >= umbral_win)
            pasa_piso = False

        pasa_adaptive = False
        if adaptive_on and h_win is not None and render_ok and pos_win is not None:
            if adapt_min <= sc_win < umbral_win or (pasa_piso and sc_win < umbral_win):
                if last_ap_pos is None or (
                    abs(pos_win[0] - last_ap_pos[0]) <= stab_px
                    and abs(pos_win[1] - last_ap_pos[1]) <= stab_px
                ):
                    adapt_streak += 1
                else:
                    adapt_streak = 1
                last_ap_pos = pos_win
                pasa_adaptive = adapt_streak >= need_adapt
            else:
                adapt_streak = 0
                last_ap_pos = None
        else:
            adapt_streak = 0
            last_ap_pos = None

        pasa = pasa_normal or pasa_adaptive or (
            _imp10_requiere_dialogo_imprimir_windows() and pasa_piso and pasa_adaptive
        )
        if pasa:
            racha += 1
            if racha >= need_ok:
                dest: int | None = None
                if h_win is not None and (
                    sc_win >= umbral_win or pasa_adaptive or (pasa_piso and sc_win >= umbral_win_piso)
                ):
                    _hwnd_destino_imprimir_10 = int(h_win)
                    dest = int(h_win)
                elif (
                    not _imp10_requiere_dialogo_imprimir_windows()
                    and hwnd_hit is not None
                ):
                    _hwnd_destino_imprimir_10 = int(hwnd_hit)
                    dest = int(hwnd_hit)
                if dest is not None:
                    tit_hit = (win32gui.GetWindowText(dest) or "").strip()
                    if _hwnd_es_dialogo_impresion_sistema(dest, tit_hit):
                        _enviar_modal_vie_detras_dialogo_imprimir_windows(int(dest))
                        print(
                            "Flujo: diálogo «Imprimir» de Windows — destino del clic 10 "
                            f"(«{tit_hit}» hwnd={dest})"
                        )
                        _activar_dialogo_impresion_windows(dest)
                        _programar_sin_traer_indigo_para_flujo_impresion_pdf()
                    elif _imp10_requiere_dialogo_imprimir_windows():
                        print(
                            f"Flujo: aviso — match en {tit_hit!r} (hwnd={dest}) no es diálogo Windows; "
                            "se ignora (use solo el #32770 «Imprimir»)."
                        )
                        _hwnd_destino_imprimir_10 = None
                        racha = 0
                        continue
                elif _imp10_requiere_dialogo_imprimir_windows():
                    racha = 0
                    continue
                modo = "adaptativo (centro estable)" if pasa_adaptive and not pasa_normal else "umbral"
                print(
                    f"Flujo: 2.º Imprimir listo — {modo} ({need_ok}× OK; multiventana {ultima_s:.2f}; "
                    f"Win {sc_win:.2f}; destino hwnd={_hwnd_destino_imprimir_10})"
                )
                pausa_listo = _float_env("INDIGO_IMP10_PAUSA_TRAS_LISTO_SEG", 0.6)
                if pausa_listo > 0:
                    print(
                        f"Flujo: pausa {pausa_listo:.1f}s tras detectar plantilla 10 "
                        "(antes del clic; INDIGO_IMP10_PAUSA_TRAS_LISTO_SEG)"
                    )
                    time.sleep(pausa_listo)
                _mantener_foco_flujo_imprimir_9_10()
                return
        else:
            racha = 0
        time.sleep(0.32)
    if _bool_env("INDIGO_IMP10_TIMEOUT_ACEPTAR_WIN_ABIERTO", True) and _imp10_requiere_dialogo_imprimir_windows():
        h_fin = _buscar_hwnd_dialogo_imprimir_windows_abierto()
        if h_fin is not None:
            tit_f = (win32gui.GetWindowText(int(h_fin)) or "").strip()
            sc_f, _pos_f = _detalle_match_boton_imprimir_10_en_hwnd(int(h_fin), tpl, tit_f)
            min_timeout = max(0.45, _float_env("INDIGO_IMP10_TIMEOUT_WIN_MIN", 0.48))
            if sc_f >= min_timeout:
                _hwnd_destino_imprimir_10 = int(h_fin)
                _enviar_modal_vie_detras_dialogo_imprimir_windows(int(h_fin))
                _activar_dialogo_impresion_windows(int(h_fin))
                _programar_sin_traer_indigo_para_flujo_impresion_pdf()
                print(
                    f"Flujo: 2.º Imprimir — timeout pero diálogo Windows abierto "
                    f"(Win {sc_f:.2f} ≥ {min_timeout:.2f}; hwnd={h_fin}); se continúa al clic 10."
                )
                return
    if _imp10_saltar_si_guardar_pdf_abierto():
        return
    raise TimeoutError(
        f"No apareció el 2.º Imprimir modal ({nombre!r}) en {timeout_sec:.0f}s "
        f"(multiventana {ultima_s:.2f}; umbral Vie {umbral_ok:.2f}; "
        f"umbral Win {umbral_win:.2f}; último Win {sc_win:.2f}; adapt ≥{adapt_min:.2f}). "
        "Si «Guardar impresión como» ya estaba abierto (aunque minimizado), active "
        "INDIGO_IMP10_SALTAR_SI_GUARDAR_PDF=1. Baje el umbral, renueve plantilla 10, o revise Print to PDF."
    )


def _clic_imprimir_dialogo_windows_idok_fallback(hwnd_imprimir: int) -> bool:
    """Pulsa el botón predeterminado del diálogo Imprimir si el match PNG falla."""
    if not _bool_env("INDIGO_IMP10_IDOK_FALLBACK", True):
        return False
    _activar_dialogo_impresion_windows(int(hwnd_imprimir))
    time.sleep(0.12)
    try:
        win32gui.SendMessage(int(hwnd_imprimir), win32con.WM_COMMAND, 1, 0)
        print(f"Flujo: plantilla 10 — WM_COMMAND IDOK en diálogo Imprimir (hwnd={hwnd_imprimir})")
        time.sleep(0.35)
        return True
    except Exception as ex:
        print(f"Flujo: IDOK en diálogo Imprimir falló ({ex})")
        return False


def _resolver_hwnd_para_clic_imprimir_10(
    hwnd_fallback: int,
    timeout_win: float | None = None,
) -> tuple[int, str]:
    """Obliga a usar el diálogo «Imprimir» de Windows, no el modal Vie."""
    global _hwnd_destino_imprimir_10
    h = _hwnd_destino_imprimir_10
    if h is not None:
        tit = (win32gui.GetWindowText(int(h)) or "").strip()
        if _hwnd_es_dialogo_impresion_sistema(int(h), tit):
            _enviar_modal_vie_detras_dialogo_imprimir_windows(int(h))
            _activar_dialogo_impresion_windows(int(h))
            return int(h), tit
    h2 = _buscar_hwnd_dialogo_imprimir_windows_abierto()
    if h2 is not None:
        tit2 = (win32gui.GetWindowText(int(h2)) or "").strip()
        _hwnd_destino_imprimir_10 = int(h2)
        _enviar_modal_vie_detras_dialogo_imprimir_windows(int(h2))
        _activar_dialogo_impresion_windows(int(h2))
        return int(h2), tit2
    if _imp10_requiere_dialogo_imprimir_windows():
        to = float(
            timeout_win
            if timeout_win is not None
            else _float_env("INDIGO_IMP10_ESPERA_WIN_ANTES_CLIC", 90.0)
        )
        h3 = _esperar_dialogo_imprimir_windows_aparecer(to, "antes del clic 10")
        tit3 = (win32gui.GetWindowText(h3) or "").strip()
        _hwnd_destino_imprimir_10 = h3
        return h3, tit3
    tit_f = (win32gui.GetWindowText(int(hwnd_fallback)) or "").strip()
    return int(hwnd_fallback), tit_f


def _clic_imprimir_modal_segundo_y_confirmar_dialogo_windows(
    hwnd: int,
    nombre_tpl: str,
    roi: tuple[float, float, float, float],
    roi_def: tuple[float, float, float, float],
) -> None:
    """Clic plantilla 10 solo en el diálogo «Imprimir» de Windows (#32770), no en el modal Vie."""
    global _hwnd_destino_imprimir_10
    pausa_antes = _float_env("INDIGO_IMP10_PAUSA_ANTES_CLIC_SEG", 0.5)
    if pausa_antes > 0:
        print(
            f"Flujo: pausa {pausa_antes:.1f}s antes del clic plantilla 10 "
            "(INDIGO_IMP10_PAUSA_ANTES_CLIC_SEG; evite pausas largas que pierden foco)"
        )
        time.sleep(pausa_antes)
    hwnd_win, tit_win = _resolver_hwnd_para_clic_imprimir_10(int(hwnd))
    _hwnd_destino_imprimir_10 = hwnd_win
    print(
        f"Flujo: clic 10 — ventana destino fijada al diálogo Windows «{tit_win}» (hwnd={hwnd_win}); "
        "el modal Vie queda detrás."
    )
    tpl = _resolver_ruta_plantilla_png(nombre_tpl)
    rois_list = _rois_ampliados_plantilla_8_paciente(nombre_tpl, roi, roi_def)
    min_sc = _float_env("INDIGO_MATCH_MIN_PLANTILLA_10", 0.55)
    raw_clic = (os.environ.get("INDIGO_MATCH_MIN_PLANTILLA_10_CLIC") or "").strip().replace(",", ".")
    if raw_clic:
        try:
            min_sc_clic = max(0.45, min(0.99, float(raw_clic)))
        except ValueError:
            min_sc_clic = max(0.48, min_sc - 0.08)
    else:
        min_sc_clic = max(0.48, min_sc - 0.08)
    max_try = int(_float_env("INDIGO_PLANTILLA_10_MAX_INTENTOS", 5))
    pausa = _float_env("INDIGO_PLANTILLA_10_PAUSA_TRAS_CLIC", 1.2)
    exclude_xy, exclude_radio = _params_exclusion_imprimir_9()
    nudge_x = int(_float_env("INDIGO_CLICK10_NUDGE_X", 0))
    nudge_y = int(_float_env("INDIGO_CLICK10_NUDGE_Y", 0))
    click_nudge = (nudge_x, nudge_y) if (nudge_x != 0 or nudge_y != 0) else None

    print(
        f"Flujo: clic plantilla 10 (similitud ≥{min_sc_clic:.2f} en clic, referencia {min_sc:.2f}; "
        f"botón abajo-derecha; hasta {max_try} intentos; INDIGO_PLANTILLA_10_FRACCION_PEQUENO_* si el PNG tiene márgenes)"
    )
    for intento in range(1, max_try + 1):
        hwnd_clic, tit_clic = _resolver_hwnd_para_clic_imprimir_10(int(hwnd))
        es_dialogo = _hwnd_es_dialogo_impresion_sistema(hwnd_clic, tit_clic)
        if _imp10_requiere_dialogo_imprimir_windows() and not es_dialogo:
            raise RuntimeError(
                f"El clic 10 debe ir al diálogo «Imprimir» de Windows, no a «{tit_clic}» (hwnd={hwnd_clic}). "
                "Revise que tras la espera larga del PDF aparezca la ventana del sistema."
            )
        orden_10, cx_min_10, cy_min_10, rois_uso = _match_boton_imprimir_10_contexto(hwnd_clic, tit_clic)
        if orden_10 == "left" and rois_uso == [(0.02, 0.02, 0.98, 0.98)]:
            rois_uso = rois_list
        es_vie_modal = False
        excl_xy = None
        excl_r = 0
        _activar_dialogo_impresion_windows(hwnd_clic)
        time.sleep(_float_env("INDIGO_IMP10_PRECLICK_EXTRA_SEG", 0.35))
        fraccion_btn = _fraccion_click_boton_imprimir_plantilla_10(hwnd_clic, tpl)
        dest = f"diálogo «{tit_clic}»" if es_dialogo else ("modal Vie" if es_vie_modal else "Vie/IndiGO")
        frac_txt = f"; clic en fracción tpl {fraccion_btn}" if fraccion_btn is not None else ""
        roi_txt_10 = (
            f"cx≥{cx_min_10:.2f} cy≥{cy_min_10:.2f}"
            if cx_min_10 is not None and cy_min_10 is not None
            else "ROI pantalla completa"
        )
        print(
            f"Flujo: plantilla 10 intento {intento}/{max_try} en {dest} "
            f"(hwnd={hwnd_clic}; orden={orden_10}; ROI inferior-derecha; "
            f"{roi_txt_10}{frac_txt})"
        )
        clic_ok = False
        try:
            clic_plantilla_en_hwnd(
                hwnd_clic,
                tpl,
                rois_uso,
                orden_10,
                min_sc_clic,
                f"IndiGO imprimir Windows intento {intento}/{max_try}",
                pantalla_completa_si_falla=False,
                umbrales_pantalla_completa=(0.75, 0.70, 0.65, 0.58),
                fraccion_click_en_plantilla=fraccion_btn,
                click_nudge_px=click_nudge,
                boton="left",
                exclude_xy=excl_xy,
                exclude_radio_px=excl_r,
                cx_min_frac=cx_min_10,
                cy_min_frac=cy_min_10,
            )
            clic_ok = True
        except RuntimeError as ex:
            print(f"Flujo: plantilla 10 sin match en diálogo Windows ({ex})")
            if _clic_imprimir_dialogo_windows_idok_fallback(hwnd_clic):
                clic_ok = True
            elif intento >= max_try:
                raise
        time.sleep(pausa)
        if clic_ok and _clic_10_produjo_avance_impresion():
            print(
                f"Flujo: tras plantilla 10 (intento {intento}/{max_try}) avanzó impresión/guardar PDF; "
                "esperando diálogo Guardar estable…"
            )
            try:
                h_g = _esperar_dialogo_guardar_pdf_estable(
                    timeout_sec=_float_env("INDIGO_GUARDAR_PDF_APARECER_TIMEOUT", 90.0),
                    pausa_estabilizar=_float_env("INDIGO_GUARDAR_PDF_TRAS_IMP10_SEG", 2.5),
                )
                _prefijar_nombre_guardar_pdf_si_posible(int(h_g))
            except TimeoutError as ex:
                print(f"Flujo: aviso — {ex}; se continúa con guardar_pdf_cedula.")
            _programar_sin_traer_indigo_para_flujo_impresion_pdf()
            _desactivar_bloqueo_foco_impresion_9_10()
            return
        print(
            f"Flujo: tras clic plantilla 10 (intento {intento}/{max_try}) el diálogo Imprimir sigue "
            "sin avanzar; se reintenta (revise fracción del botón o use PNG solo del botón)."
        )
    raise RuntimeError(
        f"Tras {max_try} clics en plantilla 10 ({tpl.name}) no avanzó impresión/guardar PDF. "
        "Baje INDIGO_MATCH_MIN_PLANTILLA_10_CLIC o alinee el PNG con el DPI de la sesión; si el clic cae al lado del botón, "
        "use INDIGO_PLANTILLA_10_FRACCION_PEQUENO_WIN_X/Y o un recorte centrado solo en «Imprimir»."
    )


def ejecutar_pasos_flujo_indigo(hwnd: int, pasos: list[tuple], sesion: SesionIndigo) -> None:
    global _pasos_sin_traer_indigo
    _set_cedula_pdf_iteracion(sesion.cedula_actual or "")
    roi_def = (0.02, 0.02, 0.98, 0.98)
    hwnd_w = int(hwnd)
    for p in pasos:
        k = p[0]
        omitir_restaurar_indigo = False
        if _pasos_sin_traer_indigo > 0:
            _pasos_sin_traer_indigo -= 1
            omitir_restaurar_indigo = True
            if k == "wait_impresion_windows":
                print(
                    "Flujo: sin traer IndiGO durante espera de impresión/guardar PDF "
                    f"(pasos sin foco IndiGO restantes ~{_pasos_sin_traer_indigo})."
                )
        elif k == "click" and len(p) >= 2 and _plantilla_es_imprimir_modal_segundo(p[1]):
            omitir_restaurar_indigo = True
            print(
                "Flujo: clic plantilla 10 — sin traer IndiGO antes (evita cerrar «Imprimir» de Windows)."
            )
        elif k == "click" and len(p) >= 2 and _plantilla_es_opcion_menu_contexto_folios(p[1]):
            omitir_restaurar_indigo = True
            print(
                "Flujo: clic plantilla 6 (menú contextual) — sin traer IndiGO "
                "(si se trae al frente, el menú se cierra)."
            )
        elif k == "click" and len(p) >= 2 and _plantilla_es_imprimir_primero(p[1]):
            if _enum_hwnds_modal_vie_consultar_imprimir():
                omitir_restaurar_indigo = True
                print(
                    "Flujo: clic plantilla 9 — sin traer IndiGO al frente "
                    "(el botón Imprimir está en el modal «Consultar / Imprimir Historias»)."
                )
        elif k == "click_menu_folio":
            omitir_restaurar_indigo = True
        elif k == "click_folios":
            pass
        elif k == "wait_template" and len(p) >= 2 and _plantilla_es_guardar_pdf_cedula(p[1]):
            omitir_restaurar_indigo = True
            print(
                "Flujo: wait_template Guardar PDF — sin traer IndiGO (el diálogo «Guardar» debe permanecer al frente)."
            )
        elif k in (
            "type_cedula_guardar_pdf",
            "guardar_pdf_cedula",
            "guardar_pdf",
            "confirmar_guardar_pdf",
            "cerrar_impresion_tras_pdf",
            "cerrar_impresion",
            "esperar_guardar_pdf",
        ) or k in ("esperar_modal_imprimir", "wait_impresion_windows") or (
            k == "click"
            and len(p) >= 2
            and (
                _plantilla_es_imprimir_primero(p[1])
                or _plantilla_es_imprimir_modal_segundo(p[1])
            )
        ) or (
            k == "type_cedula"
            and _hwnd_guardar_pdf_si_visible() is not None
            and _bool_env("INDIGO_GUARDAR_PDF_CEDULA_DEDICADO", True)
        ):
            omitir_restaurar_indigo = True
        elif k == "click" and len(p) >= 2 and _plantilla_es_deshacer_post_impresion(p[1]):
            _pasos_sin_traer_indigo = 0
            print("Flujo: clic plantilla 12 (Deshacer) — se trae IndiGO al frente.")
        if (
            (not omitir_restaurar_indigo)
            and _bool_env("INDIGO_FOCO_GUARDAR_PDF", True)
            and _hwnd_guardar_pdf_si_visible() is not None
        ):
            omitir_restaurar_indigo = True
            _traer_dialogos_guardar_pdf_al_frente()
            time.sleep(0.08)
            if k not in ("wait_template", "wait_impresion_windows", "esperar_guardar_pdf"):
                print(
                    f"Flujo: «Guardar impresión como» visible — no traer IndiGO al frente "
                    f"(paso {k!r}; modal Vie detrás)."
                )
            if k == "key" and len(p) >= 2 and str(p[1]).lower() in ("enter", "return"):
                _confirmar_guardado_pdf_cedula(hwnd_w, sesion.cedula_actual or "")
                continue
        if not omitir_restaurar_indigo:
            restaurar_y_traer_al_frente(hwnd_w)
            time.sleep(0.08)
        if k == "click_menu_folio":
            nombre, roi = p[1], p[2]
            _clic_opcion_menu_contexto_folios(nombre, roi, roi_def)
            continue
        if k == "click":
            nombre, roi = p[1], p[2]
            if _plantilla_es_opcion_menu_contexto_folios(nombre):
                _clic_opcion_menu_contexto_folios(nombre, roi, roi_def)
                continue
            tpl = _resolver_ruta_plantilla_png(nombre)
            min_clic = MATCH_THRESHOLD
            es_motivo = _plantilla_es_motivo_consulta(nombre)
            es_imprimir_primero = _plantilla_es_imprimir_primero(nombre)
            es_imprimir_modal = _plantilla_es_imprimir_modal_segundo(nombre)
            if es_imprimir_primero:
                _clic_imprimir_primero(hwnd_w, nombre, roi, roi_def)
                continue
            if es_imprimir_modal:
                global _omitir_clic_plantilla_10
                if _omitir_clic_plantilla_10:
                    print(
                        "Flujo: clic plantilla 10 omitido — «Guardar impresión como» ya está abierto "
                        "(Microsoft Print to PDF sin 2.º Imprimir visible)."
                    )
                    _omitir_clic_plantilla_10 = False
                else:
                    _clic_imprimir_modal_segundo_y_confirmar_dialogo_windows(hwnd_w, nombre, roi, roi_def)
                continue
            if _plantilla_es_guardar_pdf_cedula(nombre):
                _clic_campo_guardar_pdf_cedula(hwnd_w, nombre, roi, roi_def)
                continue
            if _plantilla_es_deshacer_post_impresion(nombre):
                _clic_deshacer_post_impresion(hwnd_w, nombre, roi, roi_def)
                continue
            if es_motivo:
                min_clic = _float_env("INDIGO_MATCH_MIN_PLANTILLA_2", 0.42)
            rois_list = _rois_ampliados_plantilla_8_paciente(nombre, roi, roi_def)
            es_go_fab = _plantilla_click_es_go_fab(nombre)
            fraccion_motivo = _fraccion_click_combo_motivo_flecha() if es_motivo else None
            cy_max_motivo = (
                _float_env("INDIGO_MOTIVO_MATCH_CY_MAX", _float_env("INDIGO_MOTIVO_MATCH_CY_MAX_FRAC", 0.40))
                if es_motivo
                else None
            )
            if es_go_fab:
                fab = _roi_busqueda_go_fab()
                min_clic = _float_env("INDIGO_MATCH_MIN_GO", 0.54)
                if roi == roi_def:
                    rois_list = [fab, roi_def]
                else:
                    rois_list = [roi, fab, roi_def]
                print(
                    f"Flujo: clic plantilla {tpl.name!r} (GO: ROI FAB + similitud ≥{min_clic:.2f} + orden «right»; "
                    "INDIGO_GO_ROI / INDIGO_MATCH_MIN_GO)"
                )
            else:
                print(f"Flujo: clic plantilla {tpl.name!r}")
            # Con orden «left», en ROI ancha se prefieren coincidencias a la izquierda: mal para el FAB GO.
            # Motivo consulta: «right» + clic en la flecha del combo (fracción derecha de la plantilla).
            orden_clic = "right" if (es_go_fab or es_motivo) else "left"
            clic_plantilla_en_hwnd(
                hwnd_w,
                tpl,
                rois_list,
                orden_clic,
                min_clic,
                f"IndiGO {tpl.name}",
                pantalla_completa_si_falla=not es_go_fab,
                umbrales_pantalla_completa=(0.42, 0.36, 0.30, 0.24, 0.20),
                fraccion_click_en_plantilla=fraccion_motivo,
                cy_max_frac=cy_max_motivo,
                boton="left",
            )
            if es_motivo:
                print(
                    "Flujo: aviso — use type_motivo_consulta (abre con caja+F4) en lugar de click 2 "
                    "para evitar que el desplegable se cierre."
                )
                time.sleep(_float_env("INDIGO_SLEEP_TRAS_CLICK_MOTIVO", 0.55))
            if _plantilla_es_aceptar_entrada_ese(nombre) or _plantilla_es_aceptar_tras_formulario(
                nombre
            ):
                extra = _float_env("INDIGO_SLEEP_TRAS_ACEPTAR_SEG", 2.5)
                time.sleep(extra)
                try:
                    hwnd_w = encontrar_hwnd_indigo(timeout=28.0)
                    tit = win32gui.GetWindowText(hwnd_w)
                    print(
                        f"Flujo: ventana IndiGO re-localizada tras Aceptar (+{extra:.1f}s) → "
                        f"{tit!r} (hwnd={hwnd_w})"
                    )
                except TimeoutError as ex:
                    print(f"Flujo: aviso tras Aceptar — {ex}; se sigue con hwnd={hwnd_w}.")
            elif _plantilla_click_es_go_fab(nombre):
                extra_go = _float_env("INDIGO_SLEEP_TRAS_GO_SEG", 1.0)
                time.sleep(extra_go)
                try:
                    hwnd_w = encontrar_hwnd_indigo(timeout=22.0)
                    tit = win32gui.GetWindowText(hwnd_w)
                    print(
                        f"Flujo: ventana IndiGO re-localizada tras GO (+{extra_go:.1f}s) → "
                        f"{tit!r} (hwnd={hwnd_w})"
                    )
                except TimeoutError as ex:
                    print(f"Flujo: aviso tras GO — {ex}; se sigue con hwnd={hwnd_w}.")
        elif k == "click_folios":
            _ejecutar_click_folios(hwnd_w)
            _programar_sin_traer_indigo_tras_menu_folios()
        elif k == "click_right":
            nombre, roi = p[1], p[2]
            if _plantilla_es_listado_folios(nombre):
                _clic_derecho_primer_folio(hwnd_w, nombre, roi, roi_def)
            else:
                tpl = _resolver_ruta_plantilla_png(nombre)
                rois_list = _rois_ampliados_plantilla_8_paciente(nombre, roi, roi_def)
                print(f"Flujo: clic derecho plantilla {tpl.name!r}")
                clic_plantilla_en_hwnd(
                    hwnd_w,
                    tpl,
                    rois_list,
                    "left",
                    MATCH_THRESHOLD,
                    f"IndiGO {tpl.name}",
                    pantalla_completa_si_falla=True,
                    umbrales_pantalla_completa=(0.42, 0.36, 0.30, 0.24, 0.20),
                    boton="right",
                )
            if _plantilla_es_aceptar_entrada_ese(nombre) or _plantilla_es_aceptar_tras_formulario(
                nombre
            ):
                extra = _float_env("INDIGO_SLEEP_TRAS_ACEPTAR_SEG", 2.5)
                time.sleep(extra)
                try:
                    hwnd_w = encontrar_hwnd_indigo(timeout=28.0)
                    tit = win32gui.GetWindowText(hwnd_w)
                    print(
                        f"Flujo: ventana IndiGO re-localizada tras Aceptar (+{extra:.1f}s) → "
                        f"{tit!r} (hwnd={hwnd_w})"
                    )
                except TimeoutError as ex:
                    print(f"Flujo: aviso tras Aceptar — {ex}; se sigue con hwnd={hwnd_w}.")
            elif _plantilla_click_es_go_fab(nombre):
                extra_go = _float_env("INDIGO_SLEEP_TRAS_GO_SEG", 1.0)
                time.sleep(extra_go)
                try:
                    hwnd_w = encontrar_hwnd_indigo(timeout=22.0)
                    tit = win32gui.GetWindowText(hwnd_w)
                    print(
                        f"Flujo: ventana IndiGO re-localizada tras GO (+{extra_go:.1f}s) → "
                        f"{tit!r} (hwnd={hwnd_w})"
                    )
                except TimeoutError as ex:
                    print(f"Flujo: aviso tras GO — {ex}; se sigue con hwnd={hwnd_w}.")
        elif k == "click_codigo_ese":
            code = (sesion.codigo_ese or "").strip()
            if not code:
                raise RuntimeError(
                    "No hay código numérico de E.S.E.: el nombre de carpeta debe empezar por dígitos "
                    "(p. ej. «009 Guarne»)."
                )
            tpl = _resolver_ruta_plantilla_codigo_ese(code)
            rois_list = [roi_def]
            print(f"Flujo: clic en código E.S.E. {code!r} → plantilla {tpl.name!r}")
            clic_plantilla_en_hwnd(
                hwnd_w,
                tpl,
                rois_list,
                "left",
                MATCH_THRESHOLD,
                f"IndiGO codigo ESE {tpl.name}",
                pantalla_completa_si_falla=True,
                umbrales_pantalla_completa=(0.42, 0.36, 0.30, 0.24, 0.20),
                boton="left",
            )
        elif k == "esperar_modal_imprimir":
            _activar_bloqueo_foco_impresion_9_10()
            timeout_sec = float(p[1])
            min_sim = p[2] if len(p) > 2 else None
            _esperar_modal_imprimir_segundo(hwnd_w, timeout_sec, min_sim)
        elif k == "wait_template":
            nombre, roi = p[1], p[2]
            timeout_sec = float(p[3])
            min_sim = p[4] if len(p) > 4 else None
            if _plantilla_es_imprimir_modal_segundo(nombre):
                _esperar_modal_imprimir_segundo(hwnd_w, timeout_sec, min_sim)
                continue
            tpl = _resolver_ruta_plantilla_png(nombre)
            rois_list = _rois_ampliados_plantilla_8_paciente(nombre, roi, roi_def)
            es_guardar_pdf = _plantilla_es_guardar_pdf_cedula(nombre)
            es_folios = _plantilla_es_listado_folios(nombre)
            umbral_ok = float(min_sim) if min_sim is not None else _float_env("INDIGO_WAIT_MATCH_MIN", 0.32)
            raw_c = (os.environ.get("INDIGO_WAIT_CONSECUTIVE") or "").strip()
            try:
                need_ok = int(raw_c) if raw_c else 3
            except ValueError:
                need_ok = 3
            need_ok = max(1, min(need_ok, 25))
            if es_folios:
                _esperar_listado_folios(hwnd_w, tpl, rois_list, timeout_sec, umbral_ok, need_ok, nombre)
                continue
            deadline = time.time() + timeout_sec
            ultima_s = 0.0
            racha = 0
            ultimo_info = time.time()
            print(
                f"Flujo: esperar plantilla {nombre!r} (máx {timeout_sec:.0f}s, "
                f"similitud mínima {umbral_ok:.2f}, {need_ok} lecturas seguidas ≥ umbral; "
                f"INDIGO_WAIT_CONSECUTIVE para cambiar)"
            )
            if es_guardar_pdf:
                print(
                    f"Flujo: esperar {nombre!r} en ventana Guardar PDF (restaura si está minimizada)"
                )
            while time.time() < deadline:
                if es_guardar_pdf:
                    _traer_dialogos_guardar_pdf_al_frente()
                else:
                    restaurar_y_traer_al_frente(hwnd_w)
                time.sleep(0.06)
                if es_guardar_pdf:
                    ultima_s = _max_similitud_plantilla_guardar_pdf(hwnd_w, tpl, rois_list)
                else:
                    ultima_s = _max_similitud_plantilla_hwnd(hwnd_w, tpl, rois_list)
                if time.time() - ultimo_info >= 12.0:
                    print(
                        f"Flujo: … esperando {nombre!r} (última similitud {ultima_s:.2f}; "
                        f"racha {racha}/{need_ok}; quedan {max(0, int(deadline - time.time()))}s)"
                    )
                    ultimo_info = time.time()
                if ultima_s >= umbral_ok:
                    racha += 1
                    if racha >= need_ok:
                        print(f"Flujo: plantilla estable ({need_ok}× similitud ≥ {umbral_ok:.2f}; última {ultima_s:.2f})")
                        if es_guardar_pdf:
                            ced_w = _cedula_pdf_actual()
                            if ced_w:
                                print(
                                    f"Flujo: tras plantilla 11 — sustituir nombre spooler por cédula {ced_w!r}"
                                )
                                try:
                                    _preparar_campo_nombre_guardar_pdf_para_escritura(hwnd_w)
                                    _escribir_cedula_en_guardar_pdf(ced_w)
                                except Exception as ex_w:
                                    print(
                                        f"Flujo: aviso — no se pudo prefijar cédula tras plantilla 11: {ex_w}"
                                    )
                        break
                else:
                    racha = 0
                time.sleep(0.32)
            else:
                raise TimeoutError(
                    f"No apareció la plantilla {nombre!r} en {timeout_sec:.0f}s "
                    f"(última similitud {ultima_s:.2f}; umbral {umbral_ok:.2f}; "
                    f"se pidieron {need_ok} lecturas seguidas). "
                    "Baje el umbral en la línea wait_template, use INDIGO_WAIT_MATCH_MIN / INDIGO_WAIT_CONSECUTIVE, "
                    "o recorte la plantilla más pequeña y estable."
                )
        elif k == "wait_template_any":
            specs_w = p[1]
            timeout_sec = float(p[2])
            min_sim_w = p[3] if len(p) > 3 else None
            umbral_ok = float(min_sim_w) if min_sim_w is not None else _float_env("INDIGO_WAIT_MATCH_MIN", 0.32)
            raw_c = (os.environ.get("INDIGO_WAIT_CONSECUTIVE") or "").strip()
            try:
                need_ok = int(raw_c) if raw_c else 3
            except ValueError:
                need_ok = 3
            need_ok = max(1, min(need_ok, 25))
            deadline = time.time() + timeout_sec
            ultima_s = 0.0
            racha = 0
            ultimo_info = time.time()
            etiquetas = ", ".join(repr(sn) for sn, _roi in specs_w)
            print(
                f"Flujo: esperar cualquiera de las plantillas {etiquetas} (máx {timeout_sec:.0f}s, "
                f"similitud máx. ≥ {umbral_ok:.2f}, {need_ok} lecturas seguidas; INDIGO_WAIT_CONSECUTIVE)"
            )
            while time.time() < deadline:
                restaurar_y_traer_al_frente(hwnd_w)
                time.sleep(0.06)
                max_s = 0.0
                for nombre_w, roi_w in specs_w:
                    tpl_w = _resolver_ruta_plantilla_png(nombre_w)
                    rois_w = _rois_ampliados_plantilla_8_paciente(nombre_w, roi_w, roi_def)
                    max_s = max(max_s, _max_similitud_plantilla_hwnd(hwnd_w, tpl_w, rois_w))
                ultima_s = max_s
                if time.time() - ultimo_info >= 12.0:
                    print(
                        f"Flujo: … esperando cualquiera de {etiquetas} (mejor similitud {ultima_s:.2f}; "
                        f"racha {racha}/{need_ok}; quedan {max(0, int(deadline - time.time()))}s)"
                    )
                    ultimo_info = time.time()
                if ultima_s >= umbral_ok:
                    racha += 1
                    if racha >= need_ok:
                        print(
                            f"Flujo: alguna plantilla estable ({need_ok}× similitud máx. ≥ {umbral_ok:.2f}; "
                            f"última {ultima_s:.2f})"
                        )
                        break
                else:
                    racha = 0
                time.sleep(0.32)
            else:
                raise TimeoutError(
                    f"No apareció ninguna de las plantillas {etiquetas} en {timeout_sec:.0f}s "
                    f"(mejor similitud {ultima_s:.2f}; umbral {umbral_ok:.2f}; {need_ok} lecturas seguidas). "
                    "Use wait_template_any 0 1 90 0.20 (umbral más bajo), INDIGO_WAIT_MATCH_MIN, "
                    "o actualice los PNG 0 y 1."
                )
        elif k == "click_optional":
            nombre, roi = p[1], p[2]
            tpl = _resolver_ruta_plantilla_png(nombre)
            rois_list = _rois_ampliados_plantilla_8_paciente(nombre, roi, roi_def)
            print(f"Flujo: clic opcional {tpl.name!r}")
            clic_plantilla_en_hwnd(
                hwnd_w,
                tpl,
                rois_list,
                "left",
                MATCH_THRESHOLD,
                f"IndiGO opcional {tpl.name}",
                pantalla_completa_si_falla=True,
                umbrales_pantalla_completa=(0.42, 0.36, 0.30, 0.24, 0.20),
                boton="left",
            )
        elif k in ("guardar_pdf_cedula", "guardar_pdf"):
            _guardar_pdf_con_cedula_completo(hwnd_w, sesion.cedula_actual or "")
        elif k == "confirmar_guardar_pdf":
            _confirmar_guardado_pdf_cedula(hwnd_w, sesion.cedula_actual or "")
        elif k in ("cerrar_impresion_tras_pdf", "cerrar_impresion"):
            _cerrar_modales_impresion_post_pdf(hwnd_w)
        elif k == "type_cedula_guardar_pdf":
            _escribir_cedula_en_guardar_pdf(sesion.cedula_actual or "")
        elif k == "type_cedula":
            ced = (sesion.cedula_actual or "").strip()
            if (
                _hwnd_guardar_pdf_si_visible() is not None
                and _bool_env("INDIGO_GUARDAR_PDF_CEDULA_DEDICADO", True)
            ):
                _escribir_cedula_en_guardar_pdf(ced)
            else:
                print(f"Flujo: escribir cédula/documento {ced!r}")
                tabs_pdf = int(_float_env("INDIGO_GUARDAR_PDF_TABS_ANTES_CEDULA", 0))
                if tabs_pdf > 0 and _hwnd_guardar_pdf_si_visible() is not None:
                    for _ in range(min(8, max(0, tabs_pdf))):
                        keyboard.send_keys("{TAB}")
                        time.sleep(0.05)
                _type_texto_campo(ced)
        elif k == "type_motivo_consulta":
            _clic_abrir_motivo_consulta(hwnd_w)
            _seleccionar_motivo_consulta(sesion)
        elif k == "type_observacion":
            _clic_y_escribir_observacion(hwnd_w, sesion.observacion)
        elif k == "wait_impresion_windows":
            _esperar_dialogo_impresion_windows(float(p[1]))
        elif k == "esperar_guardar_pdf":
            _esperar_dialogo_guardar_pdf_estable(timeout_sec=float(p[1]))
        elif k == "sleep":
            time.sleep(float(p[1]))
        elif k == "key":
            nombre_tecla = p[1]
            mapa = {
                "tab": "{TAB}",
                "enter": "{ENTER}",
                "return": "{ENTER}",
                "esc": "{ESC}",
                "escape": "{ESC}",
            }
            tecla = nombre_tecla.lower()
            if tecla in ("enter", "return") and _hwnd_guardar_pdf_si_visible() is not None:
                _confirmar_guardado_pdf_cedula(hwnd_w, sesion.cedula_actual or "")
            else:
                send = mapa.get(nombre_tecla, "{" + nombre_tecla.upper() + "}")
                keyboard.send_keys(send)
        time.sleep(0.1)


def _listar_subcarpetas_ese() -> list[Path]:
    if not INSUMOS_ROOT.is_dir():
        return []
    return sorted([p for p in INSUMOS_ROOT.iterdir() if p.is_dir()], key=lambda p: p.name.lower())


def _resolver_excel_insumo(carpeta_ese: Path) -> Path:
    for pat in ("*.xlsx", "*.xlsm", "*.csv"):
        cands = sorted(carpeta_ese.glob(pat))
        if cands:
            return cands[0]
    raise FileNotFoundError(
        f"No hay Excel/CSV de cédulas en {carpeta_ese}. Coloque un .xlsx (primera columna = cédula)."
    )


def _normalizar_texto_cedula_desde_excel(v: object) -> str:
    """Quita apóstrofo de texto Excel ('1035…) y comillas que encierren el valor."""
    s = str(v if v is not None else "").strip()
    if not s or s.startswith("#"):
        return ""
    for _ in range(8):
        t = s.strip()
        if not t:
            break
        if t[0] in ("'", '"', "\u2018", "\u2019", "\u201c", "\u201d"):
            s = t[1:].strip()
            continue
        if t[-1] in ('"', "\u201d") and len(t) > 1:
            s = t[:-1].strip()
            continue
        break
    return s.strip()


def cargar_cedulas_desde_insumo(path: Path) -> list[str]:
    ext = path.suffix.lower()
    cedulas: list[str] = []
    if ext == ".csv":
        raw = path.read_bytes()
        texto = None
        for enc in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                texto = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if texto is None:
            texto = raw.decode("utf-8", errors="replace")
        try:
            dialect = csv.Sniffer().sniff(texto[:4096], delimiters=";,\t")
            rows = csv.reader(io.StringIO(texto), dialect=dialect)
        except csv.Error:
            rows = csv.reader(io.StringIO(texto), delimiter=";", skipinitialspace=True)
        for parts in rows:
            if not parts:
                continue
            v = _normalizar_texto_cedula_desde_excel(parts[0] if parts else "")
            if not v or v.startswith("#"):
                continue
            if not any(ch.isdigit() for ch in v):
                continue
            cedulas.append(v)
    else:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise RuntimeError("Para leer .xlsx instale: pip install openpyxl") from e
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
                v = row[0]
                s = _normalizar_texto_cedula_desde_excel(v)
                if not s or s.startswith("#"):
                    continue
                if not any(ch.isdigit() for ch in s):
                    continue
                cedulas.append(s)
        finally:
            wb.close()
    if not cedulas:
        raise ValueError(f"{path.name}: no se leyeron cédulas en la primera columna.")
    return cedulas


def _prompt_sesion_interactiva(tabla_motivo_obs: Path | None = None) -> tuple[Path, str, str, str, int | None]:
    print("\n=== RPA historias clínicas — Vie Cloud IndiGO ===\n")
    print(
        "Deje IndiGO en «Consulta historias» con el campo Paciente/cédula visible (plantilla 1; la plantilla 0 es "
        "opcional). El RPA arranca como el H40/Stone: primer foco con la plantilla 1.\n"
        "Motivo y observación se leen del Excel de catálogo (A=código; D=nº opción motivo; E=texto motivo; F=observación).\n"
        "Solo elija la carpeta de la E.S.E. que contiene el Excel de cédulas.\n"
    )
    carpetas = _listar_subcarpetas_ese()
    if carpetas:
        print(f"Insumos (subcarpetas por E.S.E.) en:\n  {INSUMOS_ROOT}\n")
        for i, p in enumerate(carpetas, start=1):
            print(f"  [{i}] {p.name}")
        sel = input("\nNúmero de la E.S.E. a procesar (Enter = 1): ").strip() or "1"
        try:
            idx = int(sel)
        except ValueError:
            idx = 1
        idx = max(1, min(idx, len(carpetas)))
        carpeta_ese = carpetas[idx - 1]
    else:
        print(
            f"No hay subcarpetas en {INSUMOS_ROOT}.\n"
            "Cree esa carpeta y dentro una subcarpeta por E.S.E. con el Excel de cédulas.\n"
        )
        ruta = input("Ruta completa a la carpeta de la E.S.E. (con el Excel): ").strip().strip('"')
        if not ruta:
            raise SystemExit("Operación cancelada: sin carpeta de insumos.")
        carpeta_ese = Path(ruta).expanduser().resolve()
        if not carpeta_ese.is_dir():
            raise SystemExit(f"No es una carpeta válida: {carpeta_ese}")

    excel = _resolver_excel_insumo(carpeta_ese)
    print(f"\nArchivo de cédulas: {excel}")
    codigo = extraer_codigo_numerico_carpeta_ese(carpeta_ese.name)
    motivo, obs, mot_n = motivo_y_observacion_desde_excel_ese(codigo, carpeta_ese.name, tabla_motivo_obs)
    tabla = _ruta_tabla_motivo_obs_ese(tabla_motivo_obs)
    info_d = f"D={mot_n}" if mot_n is not None else "D=—"
    print(f"Catálogo {tabla.name}: {info_d}; E (motivo)={motivo[:55]!r}{'…' if len(motivo) > 55 else ''}")
    print(f"F (observación): {obs[:80]!r}{'…' if len(obs) > 80 else ''}\n")
    return excel, motivo, obs, carpeta_ese.name, mot_n


def _nuevo_log_resultados(nombre_ese: str) -> Path:
    SALIDAS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    slug = re.sub(r"[^\w\-]+", "_", _norm_etiqueta(nombre_ese))[:80].strip("_") or "ese"
    return SALIDAS_DIR / f"resultados_{slug}_{ts}.csv"


def _escribir_fila_resultado(
    path: Path,
    nombre_ese: str,
    cedula: str,
    estado: str,
    inicio: str,
    fin: str,
    detalle: str = "",
) -> None:
    nuevo = not path.is_file()
    with path.open("a", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        if nuevo:
            w.writerow(["cedula", "estado", "inicio_iso", "fin_iso", "detalle", "ese"])
        w.writerow([cedula, estado, inicio, fin, detalle, nombre_ese])


def indigo_historias_main(argv: list[str] | None = None) -> None:
    if argv is None:
        argv = sys.argv[1:]
    import argparse

    p = argparse.ArgumentParser(description="RPA Vie Cloud IndiGO — historias por lista de cédulas.")
    p.add_argument(
        "--calibrar-folios",
        action="store_true",
        help="Asistente: guarda coordenadas del clic derecho en el listado (paso 5) en indigo_calibracion_folios.json.",
    )
    p.add_argument(
        "--probar-folios-clic",
        action="store_true",
        help="Prueba un clic derecho con las coordenadas guardadas (--calibrar-folios).",
    )
    p.add_argument("--flujo", type=Path, default=None, help="Archivo de flujo por cédula (defecto: flujo_indigo_historias.txt).")
    p.add_argument(
        "--flujo-inicio",
        type=Path,
        default=None,
        help="Si existe: flujo que corre una vez al inicio (defecto: flujo_indigo_inicio.txt si el archivo está presente).",
    )
    p.add_argument("--insumos-root", type=Path, default=None, help="Carpeta raíz de insumos por E.S.E.")
    p.add_argument("--no-preguntar", action="store_true", help="Usar solo argumentos (sin menú interactivo).")
    p.add_argument(
        "--carpeta-ese",
        type=Path,
        default=None,
        help="Carpeta con el Excel/.csv de cédulas (columna A). Con --no-preguntar es obligatoria si no pasa CARPETA_ESE.",
    )
    p.add_argument(
        "carpeta_pos",
        nargs="?",
        type=Path,
        default=None,
        metavar="CARPETA_ESE",
        help="(Opcional) Misma carpeta que --carpeta-ese: modo consola sin menú. Ej.: IndigoHistorias.exe \"C:\\…\\MiESE\"",
    )
    p.add_argument(
        "--tabla-motivo-obs-xlsx",
        type=Path,
        default=None,
        help="Excel catálogo: A=código; D=nº opción select motivo; E=texto motivo; F=observación.",
    )
    p.add_argument(
        "--motivo",
        default="",
        help="Con --no-preguntar: anula el texto E del catálogo y no usa D con flechas.",
    )
    p.add_argument(
        "--observacion",
        default="",
        help="Con --no-preguntar: anula la observación leída del catálogo Excel si se indica texto.",
    )
    args, _rest = p.parse_known_args(argv)

    if args.calibrar_folios:
        calibrar_clic_derecho_folios()
        return
    if args.probar_folios_clic:
        probar_clic_derecho_folios_calibrado()
        return

    global INSUMOS_ROOT
    if args.insumos_root:
        INSUMOS_ROOT = args.insumos_root.expanduser().resolve()

    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    INSUMOS_ROOT.mkdir(parents=True, exist_ok=True)

    path_flujo = args.flujo or FLUJO_DEFAULT
    pasos = cargar_pasos_flujo_indigo(path_flujo)
    path_inicio = args.flujo_inicio
    if path_inicio is None:
        path_inicio = FLUJO_INICIO_DEFAULT

    tabla_arg = args.tabla_motivo_obs_xlsx.expanduser().resolve() if args.tabla_motivo_obs_xlsx else None

    carpeta_cli = args.carpeta_ese if args.carpeta_ese is not None else args.carpeta_pos
    modo_sin_menu = args.no_preguntar or (args.carpeta_pos is not None) or (args.carpeta_ese is not None)

    if modo_sin_menu:
        if not carpeta_cli:
            exe = Path(sys.argv[0]).name
            raise SystemExit(
                f"Indique la carpeta de la E.S.E. con el Excel de cédulas.\n"
                f"  {exe} \"C:\\ruta\\carpeta_ESE\"\n"
                f"  {exe} --carpeta-ese \"C:\\ruta\\carpeta_ESE\" [--no-preguntar]\n"
                "La carpeta debe existir y contener un .xlsx o .csv (cédulas en la primera columna)."
            )
        if not carpeta_cli.is_dir():
            raise SystemExit(f"No es una carpeta válida: {carpeta_cli}")
        carpeta = carpeta_cli.expanduser().resolve()
        excel = _resolver_excel_insumo(carpeta)
        nombre_ese = carpeta.name
        codigo = extraer_codigo_numerico_carpeta_ese(nombre_ese)
        motivo, obs, mot_num = motivo_y_observacion_desde_excel_ese(codigo, nombre_ese, tabla_arg)
        if (args.motivo or "").strip():
            motivo = args.motivo.strip()
            mot_num = None
        if (args.observacion or "").strip():
            obs = args.observacion.strip()
    else:
        excel, motivo, obs, nombre_ese, mot_num = _prompt_sesion_interactiva(tabla_arg)

    cedulas = cargar_cedulas_desde_insumo(excel)
    log_path = _nuevo_log_resultados(nombre_ese)

    print(f"\nCédulas a procesar: {len(cedulas)} | Flujo por cédula: {path_flujo.name} ({len(pasos)} pasos)")
    print(f"Registro: {log_path}\n")

    if (os.environ.get("INDIGO_VENTANA_SOLO_PLANTILLA_8") or "").strip():
        print(
            "Aviso: INDIGO_VENTANA_SOLO_PLANTILLA_8 está definida y se ignora: se usan «1 …» y, si existen, "
            "«0 …» / *Pantalla inicial* en templates/ (puede borrar esa variable).\n"
        )

    sesion = SesionIndigo(motivo_consulta=motivo, observacion=obs, motivo_opcion_numero=mot_num)
    sesion.codigo_ese = extraer_codigo_numerico_carpeta_ese(nombre_ese)
    if sesion.codigo_ese:
        print(f"Código E.S.E. tomado de la carpeta «{nombre_ese}»: {sesion.codigo_ese}")
    else:
        print(
            f"Aviso: en «{nombre_ese}» no hay prefijo numérico (use p. ej. «009 Nombre»). "
            "click_codigo_ese no tendrá código."
        )

    if path_inicio.is_file():
        print(f"\n--- Fase inicial (una vez): {path_inicio.name} ---")
        pasos_inicio = cargar_pasos_flujo_indigo(path_inicio)
        print(f"Pasos inicio: {len(pasos_inicio)}")
        hwnd0 = encontrar_hwnd_indigo()
        print(f"Ventana: {win32gui.GetWindowText(hwnd0)!r} (hwnd={hwnd0})")
        ejecutar_pasos_flujo_indigo(hwnd0, pasos_inicio, sesion)
        print("--- Fin fase inicial; comienza el bucle por cédula ---\n")
    else:
        print(
            f"\n(Aviso) No hay fase inicial: no existe {path_inicio.name}. "
            "Créelo con wait_template 1 y click 1 (como H40/Stone), o wait_template_any 0 1 si usa también PNG «0».\n"
        )

    for i, ced in enumerate(cedulas, start=1):
        sesion.cedula_actual = _normalizar_cedula_nombre_pdf(ced) or ced
        _set_cedula_pdf_iteracion(sesion.cedula_actual)
        t_ini = datetime.now().isoformat(timespec="seconds")
        estado = "OK"
        detalle = ""
        print("\n" + "=" * 60)
        print(f"[{i}/{len(cedulas)}] Cédula {ced!r}")
        try:
            hwnd = encontrar_hwnd_indigo()
            titulo = win32gui.GetWindowText(hwnd)
            print(f"Ventana: {titulo!r} (hwnd={hwnd})")
            ejecutar_pasos_flujo_indigo(hwnd, pasos, sesion)
        except Exception as ex:
            msg = str(ex).lower()
            es_timeout = type(ex).__name__ == "TimeoutError"
            estado = "NO_ENCONTRADO" if ("sin coincidencia" in msg or es_timeout) else "ERROR"
            detalle = str(ex)[:500]
            print(f"Fallo: {ex}")
            if (
                "fail-safe" in msg
                or "imprimir" in detalle.lower()
                or _enum_hwnds_dialogo_impresion_sistema()
            ):
                _preparar_siguiente_cedula_tras_fallo_impresion()
        t_fin = datetime.now().isoformat(timespec="seconds")
        _escribir_fila_resultado(log_path, nombre_ese, ced, estado, t_ini, t_fin, detalle)

    print("\nProcesamiento terminado.")


if __name__ == "__main__":
    indigo_historias_main()
